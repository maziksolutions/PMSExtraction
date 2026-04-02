from __future__ import annotations

import uuid
from datetime import datetime
from typing import Annotated, Any, List, Optional

import io

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.models.component import Component, ComponentTemplate, QCStatus
from app.models.feedback import CorrectionType, FeedbackEntry
from app.models.ingestion import Manual
from app.models.job import Job
from app.models.spare import Spare
from app.models.user import User
from app.models.vessel import VesselProject
from app.schemas.component import ComponentCreate, ComponentOut, ComponentUpdate
from app.services.component_matcher import merge_component_into_target
from app.services.review_workflow import (
    broadcast_activity,
    log_activity,
    sync_components_to_global_library,
)
from app.services.vessel_library import ensure_vessel_library_baseline

router = APIRouter()


async def _get_vessel_or_404(vessel_id: uuid.UUID, db: AsyncSession) -> VesselProject:
    result = await db.execute(
        select(VesselProject).where(
            VesselProject.id == vessel_id, VesselProject.is_deleted == False
        )
    )
    vessel = result.scalar_one_or_none()
    if vessel is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vessel not found")
    return vessel


async def _normalize_pending_extracted_components(
    *,
    vessel_id: uuid.UUID,
    tenant_id: uuid.UUID,
    vessel_updated_at: Optional[datetime],
    db: AsyncSession,
) -> None:
    manual_result = await db.execute(
        select(Manual.id).where(
            Manual.vessel_id == vessel_id,
            Manual.tenant_id == tenant_id,
            Manual.is_deleted == False,
        )
    )
    current_vessel_manual_ids = {manual_id for manual_id in manual_result.scalars().all()}

    comp_result = await db.execute(
        select(Component).where(
            Component.vessel_id == vessel_id,
            Component.tenant_id == tenant_id,
            Component.is_deleted == False,
            Component.qc_status == QCStatus.pending,
            Component.source_manual_id.is_not(None),
        )
    )
    components = comp_result.scalars().all()

    changed = False
    for comp in components:
        should_be_unmapped = comp.source_manual_id in current_vessel_manual_ids
        if (
            should_be_unmapped
            and vessel_updated_at is not None
            and comp.created_at is not None
            and comp.created_at < vessel_updated_at
        ):
            comp.is_deleted = True
            db.add(comp)
            changed = True
            continue
        if comp.is_unmapped != should_be_unmapped:
            comp.is_unmapped = should_be_unmapped
            db.add(comp)
            changed = True

    if changed:
        await db.commit()


@router.get("/{vessel_id}/components", summary="List components with filters")
async def list_components(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    group1: Optional[str] = Query(None),
    group2: Optional[str] = Query(None),
    main_machinery: Optional[str] = Query(None),
    qc_status: Optional[str] = Query(None),
    min_confidence: Optional[int] = Query(None),
    is_unmapped: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    maker_filter: Optional[str] = Query(None),
    model_filter: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=5000),
) -> dict[str, Any]:
    vessel = await _get_vessel_or_404(vessel_id, db)
    await ensure_vessel_library_baseline(
        db=db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        vessel_type_name=vessel.vessel_type,
    )
    await _normalize_pending_extracted_components(
        vessel_id=vessel_id,
        tenant_id=current_user.tenant_id,
        vessel_updated_at=vessel.updated_at,
        db=db,
    )
    base_where = [
        Component.vessel_id == vessel_id,
        Component.tenant_id == current_user.tenant_id,
        Component.is_deleted == False,
    ]
    if group1:
        base_where.append(Component.group1 == group1)
    if group2:
        base_where.append(Component.group2 == group2)
    if main_machinery:
        base_where.append(Component.main_machinery == main_machinery)
    if qc_status:
        try:
            base_where.append(Component.qc_status == QCStatus(qc_status))
        except ValueError:
            pass
    if min_confidence is not None:
        base_where.append(Component.confidence_score >= min_confidence)
    if is_unmapped is not None:
        base_where.append(Component.is_unmapped == is_unmapped)
    if search:
        search_term = f"%{search}%"
        from sqlalchemy import or_
        base_where.append(
            or_(
                Component.component_name.ilike(search_term),
                Component.maker.ilike(search_term),
                Component.model.ilike(search_term),
                Component.main_machinery.ilike(search_term),
            )
        )
    if maker_filter:
        base_where.append(Component.maker.ilike(f"%{maker_filter}%"))
    if model_filter:
        base_where.append(Component.model.ilike(f"%{model_filter}%"))

    total_result = await db.execute(select(func.count()).select_from(Component).where(*base_where))
    total: int = total_result.scalar_one()

    query = (
        select(Component)
        .where(*base_where)
        .order_by(Component.group1, Component.group2, Component.main_machinery, Component.component_name)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(query)
    components = result.scalars().all()
    return {"items": [ComponentOut.model_validate(c) for c in components], "page": page, "page_size": page_size, "total": total}


@router.post(
    "/{vessel_id}/components/auto-merge-extracted",
    summary="Auto-merge extracted components into matching library components",
)
async def auto_merge_extracted(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """
    For each extracted component (has source_manual_id), find the best matching
    library component (source_manual_id IS NULL) by component name similarity.
    If similarity >= 70%, merge extracted data into library component and delete
    the extracted duplicate.
    """
    from app.services.component_matcher import auto_merge_extracted_components
    merged, unmatched = await auto_merge_extracted_components(
        db=db,
        vessel_id=vessel_id,
        tenant_id=current_user.tenant_id,
    )
    return {"merged": merged, "unmatched": unmatched}


@router.post(
    "/{vessel_id}/components",
    response_model=ComponentOut,
    status_code=status.HTTP_201_CREATED,
    summary="Create a component manually",
)
async def create_component(
    vessel_id: uuid.UUID,
    body: ComponentCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ComponentOut:
    await _get_vessel_or_404(vessel_id, db)
    comp = Component(
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        **body.model_dump(),
    )
    db.add(comp)
    await db.flush()
    activity = await log_activity(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        action_type="component.created",
        entity_type="component",
        entity_id=comp.id,
        description=f"Created component '{comp.component_name}'.",
        metadata={"group1": comp.group1, "main_machinery": comp.main_machinery},
    )
    if comp.qc_status == QCStatus.accepted:
        await sync_components_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            components=[comp],
        )
    await db.commit()
    await db.refresh(comp)
    await broadcast_activity(activity)
    return ComponentOut.model_validate(comp)


@router.patch(
    "/{vessel_id}/components/{component_id}",
    response_model=ComponentOut,
    summary="Update a component",
)
async def update_component(
    vessel_id: uuid.UUID,
    component_id: uuid.UUID,
    body: ComponentUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ComponentOut:
    await _get_vessel_or_404(vessel_id, db)
    result = await db.execute(
        select(Component).where(
            Component.id == component_id,
            Component.vessel_id == vessel_id,
            Component.is_deleted == False,
        )
    )
    comp = result.scalar_one_or_none()
    if comp is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Component not found")

    original = {
        "component_name": comp.component_name,
        "maker": comp.maker,
        "model": comp.model,
        "qc_status": comp.qc_status.value if comp.qc_status else None,
    }
    original_qc_status = comp.qc_status

    update_data = body.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(comp, field, value)
    db.add(comp)

    if comp.confidence_score and comp.confidence_score > 0 and update_data:
        manual_id = comp.source_manual_id
        if manual_id:
            feedback = FeedbackEntry(
                tenant_id=current_user.tenant_id,
                manual_id=manual_id,
                entity_type="component",
                original_value=original,
                corrected_value=update_data,
                correction_type=CorrectionType.wrong_value,
                created_by=current_user.id,
            )
            db.add(feedback)

    activity = None
    if update_data:
        activity = await log_activity(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            user_id=current_user.id,
            action_type="component.corrected" if comp.source_manual_id else "component.modified",
            entity_type="component",
            entity_id=comp.id,
            description=f"Updated component '{comp.component_name}'.",
            metadata={"fields": sorted(update_data.keys())},
        )
    if comp.qc_status == QCStatus.accepted and (
        original_qc_status != QCStatus.accepted or update_data
    ):
        await sync_components_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            components=[comp],
        )
    await db.commit()
    await db.refresh(comp)
    if activity:
        await broadcast_activity(activity)
    return ComponentOut.model_validate(comp)


@router.delete("/{vessel_id}/components/{component_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response, response_model=None)
async def delete_component(
    vessel_id: uuid.UUID,
    component_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    result = await db.execute(
        select(Component).where(
            Component.id == component_id, Component.vessel_id == vessel_id
        )
    )
    comp = result.scalar_one_or_none()
    if comp:
        comp.is_deleted = True
        db.add(comp)
        await db.commit()


@router.post("/{vessel_id}/components/bulk-update", summary="Bulk update fields on selected components")
async def bulk_update_components(
    vessel_id: uuid.UUID,
    body: dict,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """
    Apply the same field values to all specified component IDs.
    Body: { "ids": [...], "updates": { field: value, ... } }
    Only fields present in updates (with non-empty string / non-None values) are applied.
    """
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    updates: dict = body.get("updates", {})
    if not ids or not updates:
        return {"updated": 0}

    # Whitelist of patchable fields
    ALLOWED = {
        "qc_status", "criticality", "is_critical",
        "maker", "model", "specification", "serial_number",
        "location", "machinery_particulars",
        "job_pages", "spare_pages", "pdf_reference",
    }
    safe_updates = {k: v for k, v in updates.items() if k in ALLOWED and v is not None and v != ""}

    if not safe_updates:
        return {"updated": 0}

    # Convert qc_status string → enum
    if "qc_status" in safe_updates:
        try:
            safe_updates["qc_status"] = QCStatus(safe_updates["qc_status"])
        except ValueError:
            safe_updates.pop("qc_status")

    result = await db.execute(
        select(Component).where(
            Component.id.in_(ids),
            Component.vessel_id == vessel_id,
            Component.tenant_id == current_user.tenant_id,
            Component.is_deleted == False,
        )
    )
    components = result.scalars().all()
    for comp in components:
        for field, value in safe_updates.items():
            setattr(comp, field, value)
        db.add(comp)
    await db.commit()
    return {"updated": len(components)}


@router.post("/{vessel_id}/components/bulk-accept", summary="Bulk accept components")
async def bulk_accept_components(
    vessel_id: uuid.UUID,
    body: dict[str, List[str]],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    result = await db.execute(
        select(Component).where(
            Component.id.in_(ids), Component.vessel_id == vessel_id
        )
    )
    components = result.scalars().all()
    activities = []
    for comp in components:
        comp.qc_status = QCStatus.accepted
        db.add(comp)
        activities.append(
            await log_activity(
                db,
                tenant_id=current_user.tenant_id,
                vessel_id=vessel_id,
                user_id=current_user.id,
                action_type="component.accepted",
                entity_type="component",
                entity_id=comp.id,
                description=f"Accepted component '{comp.component_name}'.",
            )
        )
    if components:
        await sync_components_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            components=components,
        )
    await db.commit()
    for activity in activities:
        await broadcast_activity(activity)
    return {"accepted": len(components)}


@router.post("/{vessel_id}/components/bulk-reject", summary="Bulk reject components")
async def bulk_reject_components(
    vessel_id: uuid.UUID,
    body: dict[str, List[str]],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    result = await db.execute(
        select(Component).where(
            Component.id.in_(ids), Component.vessel_id == vessel_id
        )
    )
    components = result.scalars().all()
    activities = []
    for comp in components:
        comp.qc_status = QCStatus.rejected
        db.add(comp)
        activities.append(
            await log_activity(
                db,
                tenant_id=current_user.tenant_id,
                vessel_id=vessel_id,
                user_id=current_user.id,
                action_type="component.rejected",
                entity_type="component",
                entity_id=comp.id,
                description=f"Rejected component '{comp.component_name}'.",
            )
        )
    await db.commit()
    for activity in activities:
        await broadcast_activity(activity)
    return {"rejected": len(components)}


@router.post("/{vessel_id}/components/{component_id}/remap", summary="Remap component hierarchy")
async def remap_component(
    vessel_id: uuid.UUID,
    component_id: uuid.UUID,
    body: dict[str, str],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ComponentOut:
    result = await db.execute(
        select(Component).where(
            Component.id == component_id,
            Component.vessel_id == vessel_id,
            Component.is_deleted == False,
        )
    )
    comp = result.scalar_one_or_none()
    if comp is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Component not found")

    if "component_name" in body and body["component_name"].strip():
        comp.component_name = body["component_name"].strip()
    if "group1" in body:
        comp.group1 = body["group1"]
    if "group2" in body:
        comp.group2 = body["group2"]
    if "main_machinery" in body:
        comp.main_machinery = body["main_machinery"]
    requested_qc = body.get("qc_status")
    if requested_qc:
        try:
            comp.qc_status = QCStatus(requested_qc)
        except ValueError:
            pass
    elif comp.source_manual_id and comp.qc_status == QCStatus.pending:
        comp.qc_status = QCStatus.modified
    comp.is_unmapped = False
    db.add(comp)
    if comp.qc_status == QCStatus.accepted:
        await sync_components_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            components=[comp],
        )
    activity = await log_activity(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        action_type="component.remapped",
        entity_type="component",
        entity_id=comp.id,
        description=f"Mapped extracted component '{comp.component_name}' into the vessel structure.",
        metadata={"group1": comp.group1, "group2": comp.group2, "main_machinery": comp.main_machinery},
    )
    await db.commit()
    await db.refresh(comp)
    await broadcast_activity(activity)
    return ComponentOut.model_validate(comp)


@router.post("/{vessel_id}/components/{component_id}/merge-into", summary="Merge an extracted unmapped component into an existing vessel component")
async def merge_component_into_existing(
    vessel_id: uuid.UUID,
    component_id: uuid.UUID,
    body: dict[str, str],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ComponentOut:
    target_id_raw = body.get("target_component_id")
    if not target_id_raw:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="target_component_id is required")

    target_component_id = uuid.UUID(target_id_raw)
    result = await db.execute(
        select(Component).where(
            Component.id.in_([component_id, target_component_id]),
            Component.vessel_id == vessel_id,
            Component.tenant_id == current_user.tenant_id,
            Component.is_deleted == False,
        )
    )
    components = {component.id: component for component in result.scalars().all()}
    source = components.get(component_id)
    target = components.get(target_component_id)

    if source is None or target is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Component not found")
    if source.id == target.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Source and target cannot be the same component")

    merge_component_into_target(source, target)
    target.qc_status = QCStatus.accepted
    db.add(target)

    jobs_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == current_user.tenant_id,
            Job.component_id == source.id,
            Job.is_deleted == False,
        )
    )
    for job in jobs_result.scalars().all():
        job.component_id = target.id
        job.is_unmapped = False
        db.add(job)

    spares_result = await db.execute(
        select(Spare).where(
            Spare.vessel_id == vessel_id,
            Spare.tenant_id == current_user.tenant_id,
            Spare.component_id == source.id,
            Spare.is_deleted == False,
        )
    )
    for spare in spares_result.scalars().all():
        spare.component_id = target.id
        db.add(spare)

    source.is_deleted = True
    source.is_unmapped = False
    source.qc_status = QCStatus.accepted
    db.add(source)
    await sync_components_to_global_library(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        components=[target],
    )
    activity = await log_activity(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        action_type="component.merged",
        entity_type="component",
        entity_id=target.id,
        description=f"Merged extracted component '{source.component_name}' into '{target.component_name}'.",
        metadata={"source_component_id": str(source.id)},
    )

    await db.commit()
    await db.refresh(target)
    await broadcast_activity(activity)
    return ComponentOut.model_validate(target)


@router.get("/components/import-template", summary="Download blank Excel template for component import")
async def download_components_template() -> StreamingResponse:
    """Return a pre-formatted .xlsx template for the Components page import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Components"

    headers = [
        "Group", "Sub-Group", "Main Machinery", "Component Name",
        "Maker", "Model", "Serial Number", "Specification",
        "Location/Sets Installed", "Machinery Particulars",
        "Critical", "Job Pages", "Spare Pages", "PDF Reference",
    ]
    sample_rows = [
        ["Propulsion", "Main Engine", "Main Engine", "Cylinder Head",
         "MAN B&W", "S60MC-C", "SN-12345", "6-cylinder, 2-stroke diesel",
         "Engine Room", "Machinery Particulars Rev.3",
         "Yes", "12-15", "45-48", "ME-Manual-p12"],
        ["Propulsion", "Main Engine", "Main Engine", "Fuel Injection Pump",
         "MAN B&W", "S60MC-C", "", "High-pressure fuel injection",
         "Engine Room", "",
         "Yes", "18-20", "50-52", "ME-Manual-p18"],
        ["Ballast System", "Ballast Pumps", "Ballast Pump No.1", "Ballast Pump",
         "Shinko", "SBP-500", "BP-001", "500 m³/h, 2.5 bar",
         "Pump Room", "Machinery Particulars Rev.3",
         "No", "", "60-62", "BP-Manual-p5"],
        ["Deck Machinery", "Mooring", "Anchor Windlass", "Anchor Windlass",
         "Rolls-Royce", "AW-250", "", "250 kN pull",
         "Forecastle Deck", "",
         "No", "8", "70", ""],
    ]

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(sample_rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 16)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=components_import_template.xlsx"},
    )


@router.get("/{vessel_id}/components/export", summary="Export QC-accepted components as Excel")
async def export_components(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    qc_status: Optional[str] = Query("accepted"),
) -> StreamingResponse:
    """Export components (default: QC accepted) as .xlsx in the import format."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    await _get_vessel_or_404(vessel_id, db)

    base_where = [
        Component.vessel_id == vessel_id,
        Component.tenant_id == current_user.tenant_id,
        Component.is_deleted == False,
    ]
    if qc_status:
        try:
            base_where.append(Component.qc_status == QCStatus(qc_status))
        except ValueError:
            pass

    result = await db.execute(
        select(Component)
        .where(*base_where)
        .order_by(Component.group1, Component.group2, Component.main_machinery, Component.component_name)
    )
    components = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Components"

    headers = [
        "Group", "Sub-Group", "Main Machinery", "Component Name",
        "Maker", "Model", "Serial Number", "Specification",
        "Location/Sets Installed", "Machinery Particulars",
        "Critical", "Job Pages", "Spare Pages", "PDF Reference",
    ]

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, comp in enumerate(components, start=2):
        ws.cell(row=row_idx, column=1, value=comp.group1)
        ws.cell(row=row_idx, column=2, value=comp.group2)
        ws.cell(row=row_idx, column=3, value=comp.main_machinery)
        ws.cell(row=row_idx, column=4, value=comp.component_name)
        ws.cell(row=row_idx, column=5, value=comp.maker or "")
        ws.cell(row=row_idx, column=6, value=comp.model or "")
        ws.cell(row=row_idx, column=7, value=comp.serial_number or "")
        ws.cell(row=row_idx, column=8, value=comp.specification or "")
        ws.cell(row=row_idx, column=9, value=comp.location or "")
        ws.cell(row=row_idx, column=10, value=comp.machinery_particulars or "")
        ws.cell(row=row_idx, column=11, value="Yes" if comp.is_critical else "No")
        ws.cell(row=row_idx, column=12, value=comp.job_pages or "")
        ws.cell(row=row_idx, column=13, value=comp.spare_pages or "")
        ws.cell(row=row_idx, column=14, value=comp.pdf_reference or "")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 16)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=components_export_{vessel_id}.xlsx"},
    )


@router.post("/{vessel_id}/components/import-excel", summary="Import component hierarchy from Excel/CSV")
async def import_components_excel(
    vessel_id: uuid.UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    """
    Parse an Excel (.xlsx) or CSV file and bulk-create components.

    Expected columns (case-insensitive, order flexible):
    Group | Sub-Group | Main Machinery | Component Name | Maker | Model |
    Serial Number | Specification | Critical | Job Pages | Spare Pages | PDF Reference
    """
    await _get_vessel_or_404(vessel_id, db)
    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []

    try:
        if filename.endswith(".csv"):
            import csv, io as _io
            reader = csv.DictReader(_io.StringIO(content.decode("utf-8", errors="replace")))
            rows = [dict(r) for r in reader]
        else:
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
            wb.close()
    except Exception as exc:
        from fastapi import HTTPException as _HTTPException
        raise _HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    # Normalise header aliases
    ALIASES = {
        "group": "group1",
        "group 1": "group1",
        "group1": "group1",
        "sub-group": "group2",
        "sub group": "group2",
        "subgroup": "group2",
        "group 2": "group2",
        "group2": "group2",
        "main machinery": "main_machinery",
        "machinery": "main_machinery",
        "component name": "component_name",
        "component": "component_name",
        "name": "component_name",
        "serial number": "serial_number",
        "serial no": "serial_number",
        "serialnumber": "serial_number",
        "critical": "is_critical",
        "job pages": "job_pages",
        "jobpages": "job_pages",
        "spare pages": "spare_pages",
        "sparepages": "spare_pages",
        "pdf reference": "pdf_reference",
        "pdf ref": "pdf_reference",
        "pdfreference": "pdf_reference",
        "specification": "specification",
        "spec": "specification",
        "maker": "maker",
        "manufacturer": "maker",
        "model": "model",
        "location/sets installed": "location",
        "location": "location",
        "sets installed": "location",
        "machinery particulars": "machinery_particulars",
        "mp reference": "machinery_particulars",
        "machinery_particulars": "machinery_particulars",
    }

    def _normalise(row: dict) -> dict:
        return {ALIASES.get(k.lower().strip(), k.lower().strip()): v for k, v in row.items()}

    created = 0
    skipped = 0
    for raw_row in rows:
        r = _normalise(raw_row)
        g1 = r.get("group1") or ""
        g2 = r.get("group2") or ""
        mm = r.get("main_machinery") or ""
        name = r.get("component_name") or ""
        if not name:
            skipped += 1
            continue

        critical_raw = str(r.get("is_critical") or "").lower()
        is_critical = critical_raw in {"yes", "true", "1", "y"}

        comp = Component(
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            group1=g1 or "Uncategorised",
            group2=g2 or "Uncategorised",
            main_machinery=mm or "Unknown",
            component_name=name,
            maker=r.get("maker") or None,
            model=r.get("model") or None,
            specification=r.get("specification") or None,
            serial_number=r.get("serial_number") or None,
            location=r.get("location") or None,
            machinery_particulars=r.get("machinery_particulars") or None,
            is_critical=is_critical,
            job_pages=r.get("job_pages") or None,
            spare_pages=r.get("spare_pages") or None,
            pdf_reference=r.get("pdf_reference") or None,
            confidence_score=100,
        )
        db.add(comp)
        created += 1

    await db.commit()
    return {"imported": created, "skipped": skipped}


@router.post(
    "/{vessel_id}/components/auto-link-pages",
    summary="Auto-populate job/spare pages on components from matched manuals",
)
async def auto_link_pages(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """
    For each component that has no job_pages/spare_pages, find classified manuals for
    the same vessel and copy the page ranges from the manual's classification data.
    Matches by component name keywords against manual filenames.
    """
    await _get_vessel_or_404(vessel_id, db)

    # Get classified manuals
    manuals_res = await db.execute(
        select(Manual).where(
            Manual.vessel_id == vessel_id,
            Manual.tenant_id == current_user.tenant_id,
            Manual.is_deleted == False,
            Manual.category != None,
        )
    )
    manuals = manuals_res.scalars().all()

    # Build a lookup: manual filename -> (job_pages, spare_pages, pdf_reference)
    manual_map = [
        {
            "name": m.original_filename,
            "job_pages": m.pages_with_jobs or "",
            "spare_pages": m.pages_with_spares or "",
            "pdf_ref": m.original_filename,
            "id": m.id,
        }
        for m in manuals
        if m.pages_with_jobs or m.pages_with_spares
    ]

    if not manual_map:
        return {"updated": 0, "message": "No classified manuals with page data found."}

    # Build a fast lookup: manual_id -> manual entry
    manual_by_id = {m["id"]: m for m in manual_map}

    # Get components missing page data
    comps_res = await db.execute(
        select(Component).where(
            Component.vessel_id == vessel_id,
            Component.tenant_id == current_user.tenant_id,
            Component.is_deleted == False,
        )
    )
    components = comps_res.scalars().all()

    updated = 0
    for comp in components:
        if comp.job_pages and comp.spare_pages and comp.pdf_reference:
            continue  # already fully filled

        matched_manual = None

        # Priority 1: component already has a source_manual_id — use that manual's page data
        if comp.source_manual_id and comp.source_manual_id in manual_by_id:
            matched_manual = manual_by_id[comp.source_manual_id]
        else:
            # Priority 2: keyword overlap between component name/machinery and manual filename
            # Only accept a genuine match (score > 0) — never fall back to an arbitrary manual
            search_text = (comp.main_machinery + " " + comp.component_name).lower()
            best_score = 0
            for m in manual_map:
                manual_words = set(m["name"].lower().replace("_", " ").replace("-", " ").split())
                comp_words = set(search_text.split())
                score = len(manual_words & comp_words)
                if score > best_score:
                    best_score = score
                    matched_manual = m

            if best_score == 0:
                matched_manual = None  # no genuine match — skip this component

        if matched_manual:
            if comp.source_manual_id and matched_manual["id"] == comp.source_manual_id:
                comp.job_pages = matched_manual["job_pages"] or None
                comp.spare_pages = matched_manual["spare_pages"] or None
                comp.pdf_reference = matched_manual["pdf_ref"] or None
            else:
                if not comp.job_pages:
                    comp.job_pages = matched_manual["job_pages"]
                if not comp.spare_pages:
                    comp.spare_pages = matched_manual["spare_pages"]
                if not comp.pdf_reference:
                    comp.pdf_reference = matched_manual["pdf_ref"]
            if not comp.source_manual_id:
                comp.source_manual_id = matched_manual["id"]
            db.add(comp)
            updated += 1

    await db.commit()
    return {"updated": updated}


@router.post(
    "/{vessel_id}/components/clear-extraction-links",
    summary="Clear extraction-linked fields (pdf_reference, source_manual_id, job_pages, spare_pages, page_reference) from all components",
)
async def clear_extraction_links(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    component_ids: Optional[list[str]] = None,
) -> dict[str, Any]:
    """
    Resets pdf_reference, source_manual_id, page_reference, job_pages, and spare_pages
    to NULL for all components on the vessel (or a specific subset if component_ids provided).
    Use this to undo an incorrect auto-merge or auto-link-pages run.
    """
    from sqlalchemy import update as _update
    await _get_vessel_or_404(vessel_id, db)

    base_where = [
        Component.vessel_id == vessel_id,
        Component.tenant_id == current_user.tenant_id,
        Component.is_deleted == False,
    ]
    if component_ids:
        base_where.append(Component.id.in_([uuid.UUID(i) for i in component_ids]))

    result = await db.execute(
        _update(Component)
        .where(*base_where)
        .values(
            pdf_reference=None,
            source_manual_id=None,
            page_reference=None,
            job_pages=None,
            spare_pages=None,
        )
    )
    await db.commit()
    return {"cleared": result.rowcount}


@router.post("/{vessel_id}/components/trigger-extraction", summary="Trigger component extraction")
async def trigger_extraction(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)
    try:
        from app.tasks.extraction import extract_components

        task = extract_components.delay(str(vessel_id))
        return {"task_id": task.id, "status": "queued"}
    except Exception:
        return {"status": "queued", "task_id": "mock"}
