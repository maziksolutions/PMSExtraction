from __future__ import annotations

import io
import time
import uuid
from typing import Annotated, Any, List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, Response, UploadFile, status
from sqlalchemy import exists, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import get_db
from app.deps import get_current_user
from app.models.component import QCStatus
from app.models.feedback import CorrectionType, FeedbackEntry
from app.models.job import FrequencyType, Job
from app.models.standard_jobs import StandardJob
from app.models.user import User
from app.services.feedback_learning import schedule_feedback_learning
from app.services.job_ranks import (
    backfill_manual_job_ranks,
    derive_job_ranks_from_library_context,
    ensure_job_rank,
    infer_rank_from_component,
    normalize_rank_name,
)
from app.models.vessel import VesselProject
from app.schemas.job import JobCreate, JobOut, JobUpdate
from app.models.component import Component
from app.services.review_workflow import (
    broadcast_activity,
    log_activity,
    sync_jobs_to_global_library,
)
from app.services.job_naming import (
    append_source_references_to_description,
    build_canonical_job_name,
    split_reference_entries,
    summarize_reference_entries,
)
from app.services.upload_security import validate_uploaded_file_bytes

router = APIRouter()
_JOB_MAINTENANCE_TTL_SECONDS = settings.HOT_PATH_MAINTENANCE_TTL_SECONDS
_job_maintenance_last_run: dict[str, float] = {}


JOB_SOURCE_KIND_LABELS = {
    "instruction_manual": "Instruction Manual",
    "standard_library": "Standard Library",
    "critical_library": "Critical Jobs Library",
    "cms_file": "CMS File",
}


def _should_run_job_maintenance(vessel_id: uuid.UUID) -> bool:
    now = time.time()
    key = str(vessel_id)
    last_run = _job_maintenance_last_run.get(key, 0.0)
    if now - last_run < _JOB_MAINTENANCE_TTL_SECONDS:
        return False
    _job_maintenance_last_run[key] = now
    return True


def _merge_text_values(*values: Optional[str]) -> Optional[str]:
    seen: set[str] = set()
    merged: list[str] = []
    for value in values:
        for part in (value or "").split("\n\n"):
            cleaned = part.strip()
            if not cleaned:
                continue
            key = " ".join(cleaned.lower().split())
            if key in seen:
                continue
            seen.add(key)
            merged.append(cleaned)
    return "\n\n".join(merged) if merged else None


def _job_source_tags(
    job: Job,
    *,
    has_standard_match: bool = False,
    has_critical_match: bool = False,
) -> list[str]:
    tags: list[str] = []
    if job.source_manual_id:
        tags.append("instruction_manual")
    if has_standard_match:
        tags.append("standard_library")
    if has_critical_match:
        tags.append("critical_library")
    if job.cms_id:
        tags.append("cms_file")
    return tags


def _job_source_summary(source_tags: list[str]) -> str | None:
    labels = [JOB_SOURCE_KIND_LABELS[tag] for tag in source_tags if tag in JOB_SOURCE_KIND_LABELS]
    return " / ".join(labels) if labels else None


def _normalize_text(value: Optional[str]) -> str:
    return " ".join((value or "").strip().lower().split())


def _job_has_library_source(job: Job, standard_job: StandardJob) -> bool:
    library_reference = _normalize_text(standard_job.library_reference)
    source_reference = _normalize_text(job.source_reference)
    if library_reference and source_reference and library_reference in source_reference:
        return True
    if job.source_manual_id is None and _normalize_text(job.job_name) == _normalize_text(standard_job.job_name):
        return True
    return False


def _job_out_payload(job: Job) -> dict[str, Any]:
    return {
        "id": job.id,
        "vessel_id": job.vessel_id,
        "component_id": job.component_id,
        "job_name": job.job_name,
        "job_code": job.job_code,
        "job_description": job.job_description,
        "safety_precaution": job.safety_precaution,
        "tools_required": job.tools_required,
        "performing_rank": job.performing_rank,
        "verifying_rank": job.verifying_rank,
        "frequency": job.frequency,
        "frequency_type": job.frequency_type,
        "initial_due": job.initial_due,
        "initial_frequency_type": job.initial_frequency_type,
        "cms_id": job.cms_id,
        "page_reference": job.page_reference,
        "pdf_reference": job.pdf_reference,
        "source_reference": job.source_reference,
        "is_critical": job.is_critical,
        "qc_status": job.qc_status,
        "is_unmapped": job.is_unmapped,
        "source_manual_id": job.source_manual_id,
        "confidence_score": job.confidence_score,
        "created_at": job.created_at,
        "updated_at": job.updated_at,
    }


def _job_out(job: Job) -> JobOut:
    return JobOut.model_validate(_job_out_payload(job))


async def _component_name_for_job(db: AsyncSession, component_id: uuid.UUID | None) -> str | None:
    if not component_id:
        return None
    result = await db.execute(select(Component).where(Component.id == component_id, Component.is_deleted == False))
    component = result.scalar_one_or_none()
    return component.component_name if component else None


async def _apply_job_name_and_references(db: AsyncSession, job: Job) -> None:
    component_name = await _component_name_for_job(db, job.component_id)
    reference_entries = split_reference_entries(
        pdf_reference=job.pdf_reference,
        page_reference=job.page_reference,
        source_reference=job.source_reference,
    )
    job.job_name = build_canonical_job_name(
        component_name=component_name,
        job_names=[job.job_name],
        job_descriptions=[job.job_description],
    )
    pdf_reference, primary_page, source_reference = summarize_reference_entries(reference_entries)
    job.pdf_reference = pdf_reference
    job.page_reference = primary_page
    job.source_reference = source_reference
    job.job_description = append_source_references_to_description(job.job_description, reference_entries)
    db.add(job)


async def _normalize_vessel_job_names(
    db: AsyncSession,
    *,
    vessel_id: uuid.UUID,
    tenant_id: uuid.UUID,
) -> None:
    result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == tenant_id,
            Job.is_deleted == False,
        )
    )
    jobs = result.scalars().all()
    changed = False
    for job in jobs:
        existing_name = job.job_name
        await _apply_job_name_and_references(db, job)
        if job.job_name != existing_name:
            changed = True
    if changed:
        await db.commit()


async def _run_job_side_effects(
    db: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    vessel_id: uuid.UUID,
    user_id: uuid.UUID,
    activity_payloads: list[dict[str, Any]] | None = None,
    sync_job_ids: list[uuid.UUID] | None = None,
) -> None:
    activities = []
    if activity_payloads:
        try:
            for payload in activity_payloads:
                activities.append(
                    await log_activity(
                        db,
                        tenant_id=tenant_id,
                        vessel_id=vessel_id,
                        user_id=user_id,
                        action_type=payload["action_type"],
                        entity_type="job",
                        entity_id=payload["entity_id"],
                        description=payload["description"],
                        metadata=payload.get("metadata"),
                    )
                )
            await db.commit()
        except Exception:
            await db.rollback()
            activities = []

    for activity in activities:
        try:
            await broadcast_activity(activity)
        except Exception:
            continue

    if sync_job_ids:
        try:
            sync_result = await db.execute(
                select(Job).where(
                    Job.id.in_(sync_job_ids),
                    Job.vessel_id == vessel_id,
                    Job.is_deleted == False,
                )
            )
            sync_jobs = [
                job
                for job in sync_result.scalars().all()
                if job.qc_status == QCStatus.accepted
            ]
            await sync_jobs_to_global_library(
                db,
                tenant_id=tenant_id,
                vessel_id=vessel_id,
                jobs=sync_jobs,
            )
            await db.commit()
        except Exception:
            await db.rollback()


async def _get_vessel_or_404(vessel_id: uuid.UUID, db: AsyncSession) -> VesselProject:
    result = await db.execute(
        select(VesselProject).where(
            VesselProject.id == vessel_id, VesselProject.is_deleted == False
        )
    )
    vessel = result.scalar_one_or_none()
    if vessel is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vessel not found")
    return vessel


async def _rerun_manual_extraction(manual_ids: list[str]) -> None:
    from app.services.extractor import auto_extract_from_manual

    for manual_id in manual_ids:
        try:
            await auto_extract_from_manual(manual_id)
        except Exception:
            continue


async def _restore_soft_deleted_manual_jobs_if_vessel_empty(
    db: AsyncSession,
    *,
    vessel_id: uuid.UUID,
    tenant_id: uuid.UUID,
) -> int:
    """
    Recover extracted manual jobs that were soft-deleted by older buggy reruns.
    This only activates when the vessel currently has no active jobs at all.
    """
    active_jobs_result = await db.execute(
        select(func.count()).select_from(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == tenant_id,
            Job.is_deleted == False,
        )
    )
    if active_jobs_result.scalar_one():
        return 0

    deleted_jobs_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == tenant_id,
            Job.source_manual_id.is_not(None),
            Job.is_deleted == True,
        )
    )
    deleted_jobs = deleted_jobs_result.scalars().all()
    if not deleted_jobs:
        return 0

    for job in deleted_jobs:
        job.is_deleted = False
        db.add(job)

    await db.commit()
    return len(deleted_jobs)


@router.get("/{vessel_id}/jobs/source-files", summary="List unique source file names for jobs")
async def list_job_source_files(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    from app.models.ingestion import Manual
    await _get_vessel_or_404(vessel_id, db)
    result = await db.execute(
        select(Manual.original_filename)
        .join(Job, Job.source_manual_id == Manual.id)
        .where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == current_user.tenant_id,
            Job.is_deleted == False,
            Manual.original_filename.isnot(None),
        )
        .distinct()
        .order_by(Manual.original_filename.asc())
    )
    filenames = [row[0] for row in result.all() if row[0]]
    return {"items": filenames}


@router.get("/{vessel_id}/jobs", summary="List jobs with filters")
async def list_jobs(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    component_id: Optional[uuid.UUID] = Query(None),
    job_ids: Optional[str] = Query(None),
    qc_status: Optional[str] = Query(None),
    is_critical: Optional[bool] = Query(None),
    is_unmapped: Optional[bool] = Query(None),
    frequency_type: Optional[str] = Query(None),
    source_kind: Optional[str] = Query(None, pattern="^(instruction_manual|standard_library|critical_library|cms_file)$"),
    pdf_reference: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: str = Query("job_name"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
) -> dict[str, Any]:
    from sqlalchemy import func as _func
    from app.models.component import Component
    from app.models.ingestion import Manual
    await _get_vessel_or_404(vessel_id, db)
    if _should_run_job_maintenance(vessel_id):
        await _restore_soft_deleted_manual_jobs_if_vessel_empty(
            db,
            vessel_id=vessel_id,
            tenant_id=current_user.tenant_id,
        )
        await _normalize_vessel_job_names(
            db,
            vessel_id=vessel_id,
            tenant_id=current_user.tenant_id,
        )
        await backfill_manual_job_ranks(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
        )
    base_where = [
        Job.vessel_id == vessel_id,
        Job.tenant_id == current_user.tenant_id,
        Job.is_deleted == False,
    ]
    has_job_ids_filter = False
    if job_ids:
        parsed_job_ids: list[uuid.UUID] = []
        for raw_id in job_ids.split(","):
            raw_id = raw_id.strip()
            if not raw_id:
                continue
            try:
                parsed_job_ids.append(uuid.UUID(raw_id))
            except ValueError:
                continue
        if parsed_job_ids:
            base_where.append(Job.id.in_(parsed_job_ids))
            has_job_ids_filter = True
    if component_id:
        base_where.append(Job.component_id == component_id)
    if qc_status:
        try:
            base_where.append(Job.qc_status == QCStatus(qc_status))
        except ValueError:
            pass
    if is_critical is not None:
        base_where.append(Job.is_critical == is_critical)
    if is_unmapped is not None:
        base_where.append(Job.is_unmapped == is_unmapped)
    if frequency_type:
        try:
            base_where.append(Job.frequency_type == FrequencyType(frequency_type))
        except ValueError:
            pass
    if source_kind == "instruction_manual" and not has_job_ids_filter:
        base_where.append(Job.source_manual_id.is_not(None))
    elif source_kind in {"standard_library", "critical_library"} and not has_job_ids_filter:
        library_job_exists = (
            select(StandardJob.id)
            .where(
                StandardJob.tenant_id == current_user.tenant_id,
                StandardJob.is_deleted == False,
                StandardJob.is_critical.is_(source_kind == "critical_library"),
                or_(
                    (
                        StandardJob.library_reference.is_not(None)
                        & Job.source_reference.is_not(None)
                        & Job.source_reference.ilike(_func.concat("%", StandardJob.library_reference, "%"))
                    ),
                    (
                        Job.source_manual_id.is_(None)
                        & (Job.job_name == StandardJob.job_name)
                    ),
                ),
            )
            .limit(1)
        )
        base_where.append(exists(library_job_exists))
    elif source_kind == "cms_file" and not has_job_ids_filter:
        base_where.append(Job.cms_id.is_not(None))
    if pdf_reference:
        from app.models.ingestion import Manual as _Manual
        manual_id_result = await db.execute(
            select(_Manual.id).where(
                _Manual.vessel_id == vessel_id,
                _Manual.original_filename == pdf_reference,
                _Manual.is_deleted == False,
            )
        )
        matched_ids = [row[0] for row in manual_id_result.all()]
        base_where.append(Job.source_manual_id.in_(matched_ids) if matched_ids else (Job.id == None))
    if search:
        base_where.append(
            or_(
                Job.job_name.ilike(f"%{search}%"),
                Job.job_code.ilike(f"%{search}%"),
                Job.source_reference.ilike(f"%{search}%"),
                Job.pdf_reference.ilike(f"%{search}%"),
            )
        )

    total_result = await db.execute(select(_func.count()).select_from(Job).where(*base_where))
    total: int = total_result.scalar_one()

    sort_columns = {
        "job_name": Job.job_name,
        "component": Job.component_id,
        "job_code": Job.job_code,
        "frequency": Job.frequency,
        "frequency_type": Job.frequency_type,
        "criticality": Job.is_critical,
        "qc_status": Job.qc_status,
        "confidence": Job.confidence_score,
        "page_reference": Job.page_reference,
        "source_reference": Job.source_reference,
        "created_at": Job.created_at,
    }
    order_col = sort_columns.get(sort_by, Job.job_name)
    order_expr = order_col.desc() if sort_order == "desc" else order_col.asc()
    query = (
        select(Job)
        .where(*base_where)
        .order_by(order_expr, Job.job_name.asc(), Job.id.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(query)
    jobs = result.scalars().all()

    component_ids = {job.component_id for job in jobs if job.component_id}
    manual_ids = {job.source_manual_id for job in jobs if job.source_manual_id}

    component_lookup: dict[uuid.UUID, Component] = {}
    manual_lookup: dict[uuid.UUID, Manual] = {}

    if component_ids:
        component_result = await db.execute(select(Component).where(Component.id.in_(component_ids)))
        component_lookup = {component.id: component for component in component_result.scalars().all()}
    if manual_ids:
        manual_result = await db.execute(select(Manual).where(Manual.id.in_(manual_ids)))
        manual_lookup = {manual.id: manual for manual in manual_result.scalars().all()}

    job_ids = [job.id for job in jobs]
    standard_jobs_lookup: list[StandardJob] = []
    if job_ids:
        std_jobs_result = await db.execute(
            select(StandardJob).where(
                StandardJob.tenant_id == current_user.tenant_id,
                StandardJob.is_deleted == False,
            )
        )
        standard_jobs_lookup = std_jobs_result.scalars().all()

    items = []
    for job in jobs:
        payload = _job_out(job).model_dump()
        component = component_lookup.get(job.component_id) if job.component_id else None
        manual = manual_lookup.get(job.source_manual_id) if job.source_manual_id else None
        std_tags: set[str] = set()
        matched_standard_job: StandardJob | None = None
        for standard_job in standard_jobs_lookup:
            if not _job_has_library_source(job, standard_job):
                continue
            if matched_standard_job is None:
                matched_standard_job = standard_job
            std_tags.add("critical_library" if bool(standard_job.is_critical) else "standard_library")
        performing_rank, verifying_rank = derive_job_ranks_from_library_context(
            job_name=job.job_name,
            source_reference=job.source_reference,
            existing_performing_rank=job.performing_rank,
            existing_verifying_rank=job.verifying_rank,
            matched_standard_job=matched_standard_job,
        )
        payload["performing_rank"] = performing_rank
        payload["verifying_rank"] = verifying_rank
        source_tags = _job_source_tags(
            job,
            has_standard_match="standard_library" in std_tags,
            has_critical_match="critical_library" in std_tags,
        )
        payload.update(
            {
                "component_name": component.component_name if component else None,
                "component_maker": component.maker if component else None,
                "component_model": component.model if component else None,
                "source_manual_name": manual.original_filename if manual else None,
                "source_kinds": source_tags,
                "source_summary": _job_source_summary(source_tags),
            }
        )
        items.append(payload)

    return {"items": items, "page": page, "page_size": page_size, "total": total}


@router.get("/{vessel_id}/jobs/export", summary="Export jobs review data")
async def export_jobs(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    component_id: Optional[uuid.UUID] = Query(None),
    job_ids: Optional[str] = Query(None),
    qc_status: Optional[str] = Query(None),
    is_critical: Optional[bool] = Query(None),
    is_unmapped: Optional[bool] = Query(None),
    frequency_type: Optional[str] = Query(None),
    source_kind: Optional[str] = Query(None, pattern="^(instruction_manual|standard_library|critical_library|cms_file)$"),
    pdf_reference: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: str = Query("job_name"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$"),
    cms_pending: Optional[bool] = Query(None),
) -> Response:
    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Alignment, Font, PatternFill
    from app.models.ingestion import Manual

    vessel = await _get_vessel_or_404(vessel_id, db)
    if _should_run_job_maintenance(vessel_id):
        await _restore_soft_deleted_manual_jobs_if_vessel_empty(
            db,
            vessel_id=vessel_id,
            tenant_id=current_user.tenant_id,
        )
        await _normalize_vessel_job_names(
            db,
            vessel_id=vessel_id,
            tenant_id=current_user.tenant_id,
        )
        await backfill_manual_job_ranks(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
        )

    base_where = [
        Job.vessel_id == vessel_id,
        Job.tenant_id == current_user.tenant_id,
        Job.is_deleted == False,
    ]
    has_job_ids_filter = False
    if job_ids:
        parsed_job_ids: list[uuid.UUID] = []
        for raw_id in job_ids.split(","):
            raw_id = raw_id.strip()
            if not raw_id:
                continue
            try:
                parsed_job_ids.append(uuid.UUID(raw_id))
            except ValueError:
                continue
        if parsed_job_ids:
            base_where.append(Job.id.in_(parsed_job_ids))
            has_job_ids_filter = True
    if component_id:
        base_where.append(Job.component_id == component_id)
    if qc_status:
        try:
            base_where.append(Job.qc_status == QCStatus(qc_status))
        except ValueError:
            pass
    if is_critical is not None:
        base_where.append(Job.is_critical == is_critical)
    if is_unmapped is not None:
        base_where.append(Job.is_unmapped == is_unmapped)
    if frequency_type:
        try:
            base_where.append(Job.frequency_type == FrequencyType(frequency_type))
        except ValueError:
            pass
    if source_kind == "instruction_manual" and not has_job_ids_filter:
        base_where.append(Job.source_manual_id.is_not(None))
    elif source_kind in {"standard_library", "critical_library"} and not has_job_ids_filter:
        library_job_exists = (
            select(StandardJob.id)
            .where(
                StandardJob.tenant_id == current_user.tenant_id,
                StandardJob.is_deleted == False,
                StandardJob.is_critical.is_(source_kind == "critical_library"),
                or_(
                    (
                        StandardJob.library_reference.is_not(None)
                        & Job.source_reference.is_not(None)
                        & Job.source_reference.ilike(func.concat("%", StandardJob.library_reference, "%"))
                    ),
                    (
                        Job.source_manual_id.is_(None)
                        & (Job.job_name == StandardJob.job_name)
                    ),
                ),
            )
            .limit(1)
        )
        base_where.append(exists(library_job_exists))
    elif source_kind == "cms_file" and not has_job_ids_filter:
        base_where.append(Job.cms_id.is_not(None))
    if pdf_reference:
        manual_id_result = await db.execute(
            select(Manual.id).where(
                Manual.vessel_id == vessel_id,
                Manual.original_filename == pdf_reference,
                Manual.is_deleted == False,
            )
        )
        matched_ids = [row[0] for row in manual_id_result.all()]
        base_where.append(Job.source_manual_id.in_(matched_ids) if matched_ids else (Job.id == None))
    if search:
        base_where.append(
            or_(
                Job.job_name.ilike(f"%{search}%"),
                Job.job_code.ilike(f"%{search}%"),
                Job.source_reference.ilike(f"%{search}%"),
                Job.pdf_reference.ilike(f"%{search}%"),
            )
        )
    if cms_pending:
        base_where.append(or_(Job.cms_id.is_(None), Job.cms_id == ""))

    sort_columns = {
        "job_name": Job.job_name,
        "component": Job.component_id,
        "job_code": Job.job_code,
        "frequency": Job.frequency,
        "frequency_type": Job.frequency_type,
        "criticality": Job.is_critical,
        "qc_status": Job.qc_status,
        "confidence": Job.confidence_score,
        "page_reference": Job.page_reference,
        "source_reference": Job.source_reference,
        "created_at": Job.created_at,
    }
    order_col = sort_columns.get(sort_by, Job.job_name)
    order_expr = order_col.desc() if sort_order == "desc" else order_col.asc()

    result = await db.execute(
        select(Job)
        .where(*base_where)
        .order_by(order_expr, Job.job_name.asc(), Job.id.asc())
    )
    jobs = result.scalars().all()

    component_ids = {job.component_id for job in jobs if job.component_id}
    manual_ids = {job.source_manual_id for job in jobs if job.source_manual_id}

    component_lookup: dict[uuid.UUID, Component] = {}
    manual_lookup: dict[uuid.UUID, Manual] = {}

    if component_ids:
        component_result = await db.execute(select(Component).where(Component.id.in_(component_ids)))
        component_lookup = {component.id: component for component in component_result.scalars().all()}
    if manual_ids:
        manual_result = await db.execute(select(Manual).where(Manual.id.in_(manual_ids)))
        manual_lookup = {manual.id: manual for manual in manual_result.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [
        "ID", "PDF File", "Page", "Component", "Job Name", "Job Code",
        "Frequency", "Frequency Type", "Performing Rank", "Description",
        "Current QC", "Reviewer QC", "Reviewer Notes",
    ]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, job in enumerate(jobs, start=2):
        component = component_lookup.get(job.component_id) if job.component_id else None
        manual = manual_lookup.get(job.source_manual_id) if job.source_manual_id else None
        ws.cell(row=row_idx, column=1, value=str(job.id))
        ws.cell(row=row_idx, column=2, value=manual.original_filename if manual else (job.pdf_reference or ""))
        ws.cell(row=row_idx, column=3, value=job.page_reference or "")
        ws.cell(row=row_idx, column=4, value=component.component_name if component else "")
        ws.cell(row=row_idx, column=5, value=job.job_name or "")
        ws.cell(row=row_idx, column=6, value=job.job_code or "")
        ws.cell(row=row_idx, column=7, value=job.frequency or "")
        ws.cell(row=row_idx, column=8, value=job.frequency_type.value if job.frequency_type else "")
        ws.cell(row=row_idx, column=9, value=job.performing_rank or "")
        ws.cell(row=row_idx, column=10, value=(job.job_description or "")[:300])
        ws.cell(row=row_idx, column=11, value=job.qc_status.value if hasattr(job.qc_status, "value") else str(job.qc_status))
        ws.cell(row=row_idx, column=12, value="")
        ws.cell(row=row_idx, column=13, value="")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 4, 12), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    vessel_name = getattr(vessel, "vessel_name", None) or getattr(vessel, "name", None) or str(vessel_id)
    safe_name = "".join(ch if ch.isalnum() or ch in "-_ " else "_" for ch in vessel_name).strip() or str(vessel_id)
    filename = f"Jobs_Review_{safe_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/{vessel_id}/jobs",
    response_model=JobOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_job(
    vessel_id: uuid.UUID,
    body: JobCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> JobOut:
    await _get_vessel_or_404(vessel_id, db)
    job = Job(
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        **body.model_dump(),
    )
    if not job.performing_rank and job.component_id:
        inferred_rank = await infer_rank_from_component(
            db,
            vessel_id=vessel_id,
            component_id=job.component_id,
        )
        job.performing_rank = normalize_rank_name(inferred_rank)
    db.add(job)
    await db.commit()
    await _apply_job_name_and_references(db, job)
    await db.commit()
    await db.refresh(job)
    await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=job.performing_rank)
    await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=job.verifying_rank)
    await db.commit()
    await _run_job_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "job.created",
                "entity_id": job.id,
                "description": f"Created job '{job.job_name}'.",
                "metadata": {"component_id": str(job.component_id) if job.component_id else None},
            }
        ],
        sync_job_ids=[job.id] if job.qc_status == QCStatus.accepted else None,
    )
    await db.refresh(job)
    return _job_out(job)


@router.patch("/{vessel_id}/jobs/{job_id}", response_model=JobOut)
async def update_job(
    vessel_id: uuid.UUID,
    job_id: uuid.UUID,
    body: JobUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> JobOut:
    await _get_vessel_or_404(vessel_id, db)
    result = await db.execute(
        select(Job).where(
            Job.id == job_id, Job.vessel_id == vessel_id, Job.is_deleted == False
        )
    )
    job = result.scalar_one_or_none()
    if job is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Job not found")

    original = {"job_name": job.job_name, "qc_status": job.qc_status.value if job.qc_status else None}
    original_qc_status = job.qc_status
    update_data = body.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(job, field, value)
    if "performing_rank" in update_data:
        job.performing_rank = normalize_rank_name(job.performing_rank)
    if "verifying_rank" in update_data:
        job.verifying_rank = normalize_rank_name(job.verifying_rank)
    if (not job.performing_rank) and job.component_id and ("component_id" in update_data):
        inferred_rank = await infer_rank_from_component(
            db,
            vessel_id=vessel_id,
            component_id=job.component_id,
        )
        job.performing_rank = normalize_rank_name(inferred_rank)
    db.add(job)

    feedback_id: uuid.UUID | None = None
    if job.source_manual_id and update_data:
        feedback = FeedbackEntry(
            tenant_id=current_user.tenant_id,
            manual_id=job.source_manual_id,
            entity_type="job",
            original_value=original,
            corrected_value=update_data,
            correction_type=CorrectionType.wrong_value,
            page_number=job.page_reference,
            context_span=f"Updated job fields: {', '.join(sorted(update_data.keys()))}",
            created_by=current_user.id,
        )
        db.add(feedback)
        await db.flush()
        feedback_id = feedback.id

    await db.commit()
    if feedback_id:
        await schedule_feedback_learning(feedback_id)
    await _apply_job_name_and_references(db, job)
    await db.commit()
    await db.refresh(job)
    await _run_job_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "job.corrected" if job.source_manual_id else "job.modified",
                "entity_id": job.id,
                "description": f"Updated job '{job.job_name}'.",
                "metadata": {"fields": sorted(update_data.keys())},
            }
        ] if update_data else None,
        sync_job_ids=[job.id]
        if bool(update_data)
        and (original_qc_status == QCStatus.accepted or job.qc_status == QCStatus.accepted)
        else None,
    )
    await db.refresh(job)
    await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=job.performing_rank)
    await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=job.verifying_rank)
    await db.commit()
    return _job_out(job)


@router.delete("/{vessel_id}/jobs/{job_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response, response_model=None)
async def delete_job(
    vessel_id: uuid.UUID,
    job_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    result = await db.execute(
        select(Job).where(Job.id == job_id, Job.vessel_id == vessel_id)
    )
    job = result.scalar_one_or_none()
    if job:
        needs_sync = job.qc_status == QCStatus.accepted
        job.is_deleted = True
        db.add(job)
        await db.commit()
        if needs_sync:
            await sync_jobs_to_global_library(
                db,
                tenant_id=current_user.tenant_id,
                vessel_id=vessel_id,
                jobs=[],
            )
            await db.commit()


@router.post("/{vessel_id}/jobs/bulk-accept")
async def bulk_accept_jobs(
    vessel_id: uuid.UUID,
    body: dict[str, List[str]],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    result = await db.execute(select(Job).where(Job.id.in_(ids), Job.vessel_id == vessel_id))
    jobs = result.scalars().all()
    for job in jobs:
        job.qc_status = QCStatus.accepted
        db.add(job)
    await db.commit()
    await _run_job_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "job.accepted",
                "entity_id": job.id,
                "description": f"Accepted job '{job.job_name}'.",
            }
            for job in jobs
        ],
        sync_job_ids=[job.id for job in jobs],
    )
    return {"accepted": len(jobs)}


@router.post("/{vessel_id}/jobs/bulk-reject")
async def bulk_reject_jobs(
    vessel_id: uuid.UUID,
    body: dict[str, List[str]],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    result = await db.execute(select(Job).where(Job.id.in_(ids), Job.vessel_id == vessel_id))
    jobs = result.scalars().all()
    should_sync = any(job.qc_status == QCStatus.accepted for job in jobs)
    for job in jobs:
        job.qc_status = QCStatus.rejected
        db.add(job)
    await db.commit()
    if should_sync:
        await sync_jobs_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            jobs=[],
        )
        await db.commit()
    await _run_job_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "job.rejected",
                "entity_id": job.id,
                "description": f"Rejected job '{job.job_name}'.",
            }
            for job in jobs
        ],
    )
    return {"rejected": len(jobs)}


@router.post("/{vessel_id}/jobs/bulk-delete")
async def bulk_delete_jobs(
    vessel_id: uuid.UUID,
    body: dict[str, List[str]],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    result = await db.execute(select(Job).where(Job.id.in_(ids), Job.vessel_id == vessel_id, Job.is_deleted == False))
    jobs = result.scalars().all()
    should_sync = any(job.qc_status == QCStatus.accepted for job in jobs)
    for job in jobs:
        job.is_deleted = True
        db.add(job)
    await db.commit()
    if should_sync:
        await sync_jobs_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            jobs=[],
        )
        await db.commit()
    await _run_job_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "job.deleted",
                "entity_id": job.id,
                "description": f"Deleted job '{job.job_name}'.",
            }
            for job in jobs
        ],
    )
    return {"deleted": len(jobs)}


@router.post("/{vessel_id}/jobs/bulk-update")
async def bulk_update_jobs(
    vessel_id: uuid.UUID,
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(value) for value in body.get("ids", []) if value]
    updates = body.get("updates", {}) or {}
    if not ids or not updates:
        return {"updated": 0}

    update_payload = JobUpdate.model_validate(updates).model_dump(exclude_unset=True)
    if not update_payload:
        return {"updated": 0}

    result = await db.execute(
        select(Job).where(Job.id.in_(ids), Job.vessel_id == vessel_id, Job.is_deleted == False)
    )
    jobs = result.scalars().all()
    for job in jobs:
        setattr(job, "_original_qc_status_for_sync", job.qc_status)
        for field, value in update_payload.items():
            setattr(job, field, value)
        if "performing_rank" in update_payload:
            job.performing_rank = normalize_rank_name(job.performing_rank)
        if "verifying_rank" in update_payload:
            job.verifying_rank = normalize_rank_name(job.verifying_rank)
        if (not job.performing_rank) and ("component_id" in update_payload) and job.component_id:
            inferred_rank = await infer_rank_from_component(
                db,
                vessel_id=vessel_id,
                component_id=job.component_id,
            )
            job.performing_rank = normalize_rank_name(inferred_rank)
        db.add(job)
    await db.commit()
    for job in jobs:
        await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=job.performing_rank)
        await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=job.verifying_rank)
    await db.commit()
    await _run_job_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "job.bulk_updated",
                "entity_id": job.id,
                "description": f"Bulk updated job '{job.job_name}'.",
                "metadata": {"fields": sorted(update_payload.keys())},
            }
            for job in jobs
        ],
        sync_job_ids=[
            job.id
            for job in jobs
            if getattr(job, "_original_qc_status_for_sync", None) == QCStatus.accepted
            or job.qc_status == QCStatus.accepted
        ],
    )
    return {"updated": len(jobs)}


@router.post("/{vessel_id}/jobs/{job_id}/link-component")
async def link_component(
    vessel_id: uuid.UUID,
    job_id: uuid.UUID,
    body: dict[str, str],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> JobOut:
    result = await db.execute(
        select(Job).where(Job.id == job_id, Job.vessel_id == vessel_id, Job.is_deleted == False)
    )
    job = result.scalar_one_or_none()
    if job is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Job not found")
    job.component_id = uuid.UUID(body["component_id"])
    job.is_unmapped = False
    if not job.performing_rank:
        inferred_rank = await infer_rank_from_component(
            db,
            vessel_id=vessel_id,
            component_id=job.component_id,
        )
        job.performing_rank = normalize_rank_name(inferred_rank)
    db.add(job)
    await db.commit()
    await _apply_job_name_and_references(db, job)
    await db.commit()
    await db.refresh(job)
    await _run_job_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "job.mapped",
                "entity_id": job.id,
                "description": f"Linked job '{job.job_name}' to a vessel component.",
                "metadata": {"component_id": body["component_id"]},
            }
        ],
    )
    await db.refresh(job)
    return _job_out(job)


@router.post("/{vessel_id}/jobs/merge")
async def merge_jobs(
    vessel_id: uuid.UUID,
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> JobOut:
    ids = [uuid.UUID(value) for value in body.get("ids", []) if value]
    if len(ids) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Select at least two jobs to merge.")

    target_id_raw = body.get("target_id")
    target_id = uuid.UUID(target_id_raw) if target_id_raw else ids[0]

    result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.id.in_(ids),
            Job.is_deleted == False,
        )
    )
    jobs = list(result.scalars().all())
    if len(jobs) < 2:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Jobs not found.")

    target = next((job for job in jobs if job.id == target_id), None)
    if target is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Target job must be one of the selected jobs.")

    # Only allow merge of jobs that belong to the same component/frequency bucket.
    component_ids = {job.component_id for job in jobs}
    frequency_keys = {(job.frequency, job.frequency_type.value if job.frequency_type else None) for job in jobs}
    if len(component_ids) > 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only jobs linked to the same component can be merged.",
        )
    if len(frequency_keys) > 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only jobs with the same frequency and frequency type can be merged.",
        )

    had_accepted_job = any(job.qc_status == QCStatus.accepted for job in jobs)
    merged_ids: list[str] = []
    merged_names: list[str | None] = [target.job_name]
    merged_descriptions: list[str | None] = [target.job_description]
    reference_entries = split_reference_entries(
        pdf_reference=target.pdf_reference,
        page_reference=target.page_reference,
        source_reference=target.source_reference,
    )
    for job in jobs:
        if job.id == target.id:
            continue
        merged_names.append(job.job_name)
        merged_descriptions.append(job.job_description)
        target.job_description = _merge_text_values(target.job_description, job.job_description)
        target.safety_precaution = _merge_text_values(target.safety_precaution, job.safety_precaution)
        target.tools_required = _merge_text_values(target.tools_required, job.tools_required)
        if not target.job_code and job.job_code:
            target.job_code = job.job_code
        if not target.component_id and job.component_id:
            target.component_id = job.component_id
            target.is_unmapped = False
        if not target.performing_rank and job.performing_rank:
            target.performing_rank = job.performing_rank
        if not target.verifying_rank and job.verifying_rank:
            target.verifying_rank = job.verifying_rank
        if not target.cms_id and job.cms_id:
            target.cms_id = job.cms_id
        if target.frequency is None and job.frequency is not None:
            target.frequency = job.frequency
        if target.frequency_type is None and job.frequency_type is not None:
            target.frequency_type = job.frequency_type
        target.is_critical = bool(target.is_critical or job.is_critical)
        if (job.confidence_score or 0) > (target.confidence_score or 0):
            target.confidence_score = job.confidence_score
        reference_entries.extend(
            split_reference_entries(
                pdf_reference=job.pdf_reference,
                page_reference=job.page_reference,
                source_reference=job.source_reference,
            )
        )
        if not target.source_manual_id and job.source_manual_id:
            target.source_manual_id = job.source_manual_id
        job.is_deleted = True
        db.add(job)
        merged_ids.append(str(job.id))

    target.job_name = build_canonical_job_name(
        component_name=await _component_name_for_job(db, target.component_id),
        job_names=merged_names,
        job_descriptions=merged_descriptions,
    )
    pdf_reference, primary_page, source_reference = summarize_reference_entries(reference_entries)
    target.pdf_reference = pdf_reference
    target.page_reference = primary_page
    target.source_reference = source_reference
    target.job_description = append_source_references_to_description(target.job_description, reference_entries)
    target.qc_status = QCStatus.modified if target.qc_status == QCStatus.accepted else target.qc_status
    db.add(target)
    await db.commit()
    await _run_job_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "job.merged",
                "entity_id": target.id,
                "description": f"Merged {len(merged_ids)} jobs into '{target.job_name}'.",
                "metadata": {"merged_job_ids": merged_ids},
            }
        ],
        sync_job_ids=[target.id] if (had_accepted_job or target.qc_status == QCStatus.accepted) else None,
    )
    refreshed_result = await db.execute(
        select(Job).where(
            Job.id == target.id,
            Job.vessel_id == vessel_id,
            Job.is_deleted == False,
        )
    )
    refreshed_target = refreshed_result.scalar_one_or_none()
    if refreshed_target is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Merged job not found after merge.")
    return _job_out(refreshed_target)


@router.post("/{vessel_id}/jobs/trigger-extraction")
async def trigger_job_extraction(
    vessel_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)
    from app.models.ingestion import Manual

    result = await db.execute(
        select(Manual).where(
            Manual.vessel_id == vessel_id,
            Manual.tenant_id == current_user.tenant_id,
            Manual.is_deleted == False,
            Manual.pages_with_jobs.isnot(None),
            Manual.pages_with_jobs != "",
        )
    )
    manuals = result.scalars().all()
    manual_ids = [str(manual.id) for manual in manuals]
    if not manual_ids:
        return {"started": False, "total": 0, "message": "No manuals with job page references found."}

    background_tasks.add_task(_rerun_manual_extraction, manual_ids)
    return {"started": True, "total": len(manual_ids), "message": "Job extraction started."}


@router.post("/{vessel_id}/jobs/upload-cms-mapping")
async def upload_cms_mapping(
    vessel_id: uuid.UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    """Upload a CSV with job_name → cms_id mappings."""
    await _get_vessel_or_404(vessel_id, db)
    content = await file.read()
    validate_uploaded_file_bytes(
        filename=file.filename or "cms_mapping.csv",
        content=content,
        allowed_extensions={"csv"},
        max_size_bytes=5 * 1024 * 1024,
    )
    lines = content.decode("utf-8", errors="replace").splitlines()
    updated = 0
    for line in lines[1:]:  # Skip header
        parts = line.split(",", 1)
        if len(parts) == 2:
            job_name = parts[0].strip()
            cms_id = parts[1].strip()
            result = await db.execute(
                select(Job).where(
                    Job.vessel_id == vessel_id,
                    Job.job_name == job_name,
                    Job.is_deleted == False,
                )
            )
            jobs = result.scalars().all()
            for job in jobs:
                job.cms_id = cms_id
                db.add(job)
                updated += 1
    await db.commit()
    return {"status": "ok", "updated": updated}


@router.get("/{vessel_id}/jobs/qc-export", summary="Export Jobs for QC review (jobs only)")
async def export_jobs_qc(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> Response:
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from app.models.component import Component
    from app.models.vessel import VesselProject

    vessel_result = await db.execute(select(VesselProject).where(VesselProject.id == vessel_id))
    vessel = vessel_result.scalar_one_or_none()
    vessel_name = getattr(vessel, "vessel_name", None) or getattr(vessel, "name", None) or str(vessel_id)

    jobs_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == current_user.tenant_id,
            Job.is_deleted == False,
        ).order_by(Job.pdf_reference, Job.page_reference, Job.job_name)
    )
    jobs = list(jobs_result.scalars().all())

    component_ids = {j.component_id for j in jobs if j.component_id}
    component_lookup: dict = {}
    if component_ids:
        cr = await db.execute(select(Component).where(Component.id.in_(component_ids)))
        component_lookup = {c.id: c for c in cr.scalars().all()}

    HEADERS = [
        "ID", "PDF File", "Page", "Component", "Job Name", "Job Code",
        "Frequency", "Frequency Type", "Performing Rank", "Description",
        "Current QC", "Reviewer QC", "Reviewer Notes",
    ]
    QC_COLORS = {"accepted": "C6EFCE", "rejected": "FFC7CE", "modified": "FFEB9C", "pending": "DDEBF7"}

    HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    EDIT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ALT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"
    dv = DataValidation(type="list", formula1='"accepted,rejected,modified,pending"', showDropDown=False, allow_blank=True)
    ws.add_data_validation(dv)

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    rev_col = get_column_letter(12)
    for r, job in enumerate(jobs, 2):
        comp = component_lookup.get(job.component_id)
        comp_name = comp.component_name if comp else ""
        qc_val = (job.qc_status.value if hasattr(job.qc_status, "value") else str(job.qc_status)).lower()
        row_data = [
            str(job.id), job.pdf_reference or "", job.page_reference or "", comp_name,
            job.job_name or "", job.job_code or "", job.frequency or "",
            job.frequency_type.value if job.frequency_type else "",
            job.performing_rank or "", (job.job_description or "")[:300],
            qc_val, "", "",
        ]
        fill = ALT_FILL if r % 2 == 0 else PatternFill(fill_type=None)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c == len(HEADERS)))
            if c == 11:
                color = QC_COLORS.get(qc_val, "DDEBF7")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            elif c in (12, 13):
                cell.fill = EDIT_FILL
            else:
                cell.fill = fill
        dv.add(ws[f"{rev_col}{r}"])

    for col in ws.columns:
        vals = [str(cell.value or "") for cell in col]
        width = min(max(len(v) for v in vals) + 4, 45)
        ws.column_dimensions[col[0].column_letter].width = max(width, 10)
    ws.column_dimensions[get_column_letter(13)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in vessel_name).strip()
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Jobs_QC_{safe_name}.xlsx"'},
    )
