from __future__ import annotations

import hashlib
import io
import logging
import uuid
from typing import Annotated, Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status

logger = logging.getLogger(__name__)
from fastapi.responses import StreamingResponse
from app.core.config import settings
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.models.component import Component, QCStatus as CompQCStatus
from app.models.ingestion import Manual
from app.models.job import Job
from app.models.spare import Spare
from app.models.user import User
from app.models.vessel import VesselProject
from app.services.deduplication import is_duplicate_component, is_duplicate_job, is_duplicate_spare
from app.services.review_workflow import backfill_global_library_from_accepted_records
from app.services.vessel_library import load_library_components_for_vessel

router = APIRouter()

_DEFAULT_VESSEL_TYPES = [
    ("Oil Tanker", 1),
    ("Oil/Chemical Tanker", 2),
    ("Chemical Tanker", 3),
    ("Gas Carrier", 4),
    ("LPG Carrier", 5),
    ("Bulk Carrier", 6),
    ("Offshore Vessel", 7),
]


async def _bootstrap_schema(db: AsyncSession) -> None:
    """
    Runtime DDL bootstrap — creates vessel_types table and related columns if
    Alembic migrations 0008/0009 have not yet successfully run on this Railway
    instance.  All statements use IF NOT EXISTS so they are fully idempotent.
    """
    try:
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS vessel_types (
                id          UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                tenant_id   UUID NOT NULL,
                name        VARCHAR(200) NOT NULL,
                is_system   BOOLEAN NOT NULL DEFAULT false,
                sort_order  INTEGER NOT NULL DEFAULT 0,
                is_deleted  BOOLEAN NOT NULL DEFAULT false,
                created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
        """))
        await db.execute(text("""
            CREATE UNIQUE INDEX IF NOT EXISTS uq_vessel_types_tenant_name
            ON vessel_types (tenant_id, name) WHERE is_deleted = false
        """))
        await db.execute(text("""
            ALTER TABLE component_structure_library
            ADD COLUMN IF NOT EXISTS vessel_type_id UUID
        """))
        await db.execute(text("""
            DO $$
            BEGIN
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'fk_csl_vessel_type') THEN
                    ALTER TABLE component_structure_library
                    ADD CONSTRAINT fk_csl_vessel_type
                    FOREIGN KEY (vessel_type_id) REFERENCES vessel_types(id) ON DELETE SET NULL;
                END IF;
            END $$
        """))
        await db.execute(text("""
            ALTER TABLE components
            ADD COLUMN IF NOT EXISTS criticality VARCHAR(20) NOT NULL DEFAULT 'non_critical'
        """))
        await db.execute(text("""
            ALTER TABLE component_structure_library
            ADD COLUMN IF NOT EXISTS criticality VARCHAR(20) NOT NULL DEFAULT 'non_critical'
        """))
        await db.commit()
    except Exception as err:
        logger.warning("_bootstrap_schema: %s", err)
        await db.rollback()


async def _ensure_vessel_types(tenant_id: str, db: AsyncSession) -> None:
    """Seed any missing default vessel types for a tenant (checks each one individually)."""
    try:
        inserted = 0
        for name, sort_order in _DEFAULT_VESSEL_TYPES:
            try:
                result = await db.execute(
                    text(
                        "WITH payload AS ("
                        "  SELECT CAST(:id AS UUID) AS id, "
                        "         CAST(:tid AS UUID) AS tenant_id, "
                        "         CAST(:name AS VARCHAR(200)) AS name, "
                        "         CAST(:sort AS INTEGER) AS sort_order"
                        ") "
                        "INSERT INTO vessel_types (id, tenant_id, name, is_system, sort_order, is_deleted, created_at, updated_at) "
                        "SELECT payload.id, payload.tenant_id, payload.name, true, payload.sort_order, false, NOW(), NOW() "
                        "FROM payload "
                        "WHERE NOT EXISTS ("
                        "  SELECT 1 FROM vessel_types "
                        "  WHERE tenant_id = payload.tenant_id AND name = payload.name AND is_deleted = false"
                        ")"
                    ),
                    {"id": str(uuid.uuid4()), "tid": tenant_id, "name": name, "sort": sort_order},
                )
                inserted += result.rowcount
            except Exception as row_err:
                logger.warning("_ensure_vessel_types: skipping %r: %s", name, row_err)
                await db.rollback()
        if inserted > 0:
            await db.commit()
    except Exception as err:
        logger.error("_ensure_vessel_types failed for tenant %s: %s", tenant_id, err)
        await db.rollback()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


async def _get_vessel_or_404(vessel_id: uuid.UUID, db: AsyncSession) -> VesselProject:
    result = await db.execute(
        select(VesselProject).where(
            VesselProject.id == vessel_id,
            VesselProject.is_deleted == False,
        )
    )
    vessel = result.scalar_one_or_none()
    if vessel is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vessel not found")
    return vessel


# ===========================================================================
# F-03: Component Structure Library
# ===========================================================================


# Column aliases for the library import format:
# ShipComponentName | HierarchyComponentCode | ShipComponentCode | ComponentType | Priority | Status | Quantity | Category
_LIBRARY_ALIASES = {
    "shipcomponentname": "component_name",
    "ship component name": "component_name",
    "componentname": "component_name",
    "component name": "component_name",
    "hierarchycomponentcode": "hierarchy_code",
    "hierarchy component code": "hierarchy_code",
    "hierarchy code": "hierarchy_code",
    "hierarchycode": "hierarchy_code",
    "shipcomponentcode": "component_code",
    "ship component code": "component_code",
    "componentcode": "component_code",
    "component code": "component_code",
    "componenttype": "component_type",
    "component type": "component_type",
    "priority": "priority",
    "status": "item_status",
    "quantity": "quantity",
    "category": "category",
    # Legacy internal field names also accepted
    "group1_name": "group1_name",
    "group2_name": "group2_name",
    "machinery_name": "machinery_name",
    "component_name": "component_name",
    "group1_code": "group1_code",
    "group2_code": "group2_code",
    "machinery_code": "machinery_code",
}


def _normalise_library_row(raw: dict) -> dict:
    return {_LIBRARY_ALIASES.get(k.lower().strip(), k.lower().strip()): v for k, v in raw.items()}


# ===========================================================================
# Vessel Types
# ===========================================================================


@router.get("/library/vessel-types", summary="List vessel types for tenant")
async def list_vessel_types(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """Return all vessel types, seeding defaults if first call."""
    tid = str(current_user.tenant_id)
    await _bootstrap_schema(db)
    await _ensure_vessel_types(tid, db)
    try:
        result = await db.execute(
            text(
                "SELECT id, name, is_system, sort_order, "
                "(SELECT COUNT(*) FROM component_structure_library csl "
                " WHERE csl.vessel_type_id = vt.id AND csl.is_deleted = false) AS component_count "
                "FROM vessel_types vt "
                "WHERE tenant_id = :tid AND is_deleted = false "
                "ORDER BY sort_order, name"
            ),
            {"tid": tid},
        )
        rows = result.mappings().all()
        return {"items": [dict(r) for r in rows]}
    except Exception as err:
        logger.error("list_vessel_types SELECT failed for tenant %s: %s", tid, err)
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="vessel_types table not available — migration may still be running. Please refresh in a moment.",
        )


@router.post("/library/vessel-types", status_code=status.HTTP_201_CREATED, summary="Create a vessel type")
async def create_vessel_type(
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="name is required")
    new_id = uuid.uuid4()
    # Check for duplicate first to give a proper error
    dup = await db.execute(
        text("SELECT id FROM vessel_types WHERE tenant_id = :tid AND name = :name AND is_deleted = false"),
        {"tid": str(current_user.tenant_id), "name": name},
    )
    if dup.one_or_none() is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"Vessel type '{name}' already exists")
    try:
        await db.execute(
            text(
                "INSERT INTO vessel_types (id, tenant_id, name, is_system, sort_order, is_deleted, created_at, updated_at) "
                "VALUES (:id, :tid, :name, false, 100, false, NOW(), NOW())"
            ),
            {"id": str(new_id), "tid": str(current_user.tenant_id), "name": name},
        )
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"Vessel type '{name}' already exists")
    return {"id": str(new_id), "name": name}


@router.get(
    "/library/component-structure/template",
    summary="F-03: Download blank Excel template for component structure import",
    response_class=StreamingResponse,
)
async def download_library_template() -> StreamingResponse:
    """Return a pre-formatted .xlsx template for the component structure library import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Component Structure"

    headers = [
        "ShipComponentName",
        "HierarchyComponentCode",
        "ShipComponentCode",
        "ComponentType",
        "Priority",
        "Status",
        "Quantity",
        "Category",
    ]
    sample_rows = [
        ["Main Engine", "1.001.001.001", "ME-001", "Engine", "High", "Active", "1", "Propulsion"],
        ["Main Engine - Cylinder Head", "1.001.001.002", "ME-001-CH", "Sub-Component", "High", "Active", "6", "Propulsion"],
        ["Ballast Pump", "2.001.001.001", "BP-001", "Pump", "Medium", "Active", "2", "Ballast System"],
        ["Anchor Windlass", "3.001.001.001", "AW-001", "Deck Machinery", "Medium", "Active", "1", "Deck Equipment"],
    ]

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(sample_rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 18)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=component_library_template.xlsx"},
    )


@router.post(
    "/library/component-structure/import",
    status_code=status.HTTP_201_CREATED,
    summary="F-03: Import component structure library from Excel/CSV",
)
async def import_component_structure(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    file: UploadFile = File(...),
    vessel_type_id: Optional[str] = Query(None, description="Vessel type UUID to tag these components"),
) -> dict[str, Any]:
    """
    Parse an Excel (.xlsx) or CSV file and bulk-insert component_structure_library rows.
    Expected columns: ShipComponentName | HierarchyComponentCode | ShipComponentCode |
    ComponentType | Priority | Status | Quantity | Category
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []
    try:
        if filename.endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
            rows = [dict(r) for r in reader]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [
                str(c.value or "").strip().lower()
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(
                    {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
                )
            wb.close()
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not parse file: {exc}",
        )

    # Determine current max version for this tenant
    # Use library_version if the 0007 migration hasn't run yet, else version
    try:
        ver_result = await db.execute(
            text(
                "SELECT COALESCE(MAX(version), 0) AS max_ver "
                "FROM component_structure_library "
                "WHERE tenant_id = :tid AND is_deleted = false"
            ),
            {"tid": str(current_user.tenant_id)},
        )
        ver_col = "version"
    except Exception:
        await db.rollback()
        ver_result = await db.execute(
            text(
                "SELECT COALESCE(MAX(library_version), 0) AS max_ver "
                "FROM component_structure_library "
                "WHERE tenant_id = :tid AND is_deleted = false"
            ),
            {"tid": str(current_user.tenant_id)},
        )
        ver_col = "library_version"
    ver_row = ver_result.mappings().one()
    new_version: int = (ver_row["max_ver"] or 0) + 1

    imported = 0
    for raw in rows:
        r = _normalise_library_row(raw)

        comp_name = str(r.get("component_name") or "").strip()
        if not comp_name:
            continue

        # Derive group hierarchy from dot-notation hierarchy_code if present
        # e.g. "1.001.001.002" → group1_code="1", group2_code="1.001", machinery_code="1.001.001"
        hierarchy_code = str(r.get("hierarchy_code") or "").strip()
        parts = [p for p in hierarchy_code.split(".") if p]
        group1_code = r.get("group1_code") or (parts[0] if parts else None)
        group2_code = r.get("group2_code") or (".".join(parts[:2]) if len(parts) >= 2 else group1_code)
        if r.get("machinery_code"):
            machinery_code = r["machinery_code"]
        elif len(parts) >= 4:
            machinery_code = ".".join(parts[:3])
        elif len(parts) >= 3:
            machinery_code = ".".join(parts[:3])
        else:
            machinery_code = hierarchy_code or None

        comp_type = str(r.get("component_type") or "").strip() or None
        category = str(r.get("category") or "").strip() or None

        # Determine criticality: explicit "criticality" column > priority mapping > default
        raw_criticality = str(r.get("criticality") or "").strip().lower()
        priority = str(r.get("priority") or "").strip().lower()
        if raw_criticality in ("critical", "essential", "non_critical"):
            criticality = raw_criticality
        elif priority == "critical":
            criticality = "critical"
        elif priority == "essential":
            criticality = "essential"
        else:
            criticality = "non_critical"

        mapped = {
            "group1_code": group1_code or None,
            "group1_name": r.get("group1_name") or category or "Uncategorised",
            "group2_code": group2_code or None,
            "group2_name": r.get("group2_name") or comp_type or category or "Uncategorised",
            "machinery_code": machinery_code or None,
            "machinery_name": r.get("machinery_name") or comp_type or None,
            "component_code": r.get("component_code") or None,
            "component_name": comp_name,
            "component_type": comp_type,
            "criticality": criticality,
        }

        await db.execute(
            text(
                f"INSERT INTO component_structure_library "
                f"(id, tenant_id, group1_code, group1_name, group2_code, group2_name, "
                f"machinery_code, machinery_name, component_code, component_name, "
                f"component_type, is_critical, criticality, {ver_col}, status, is_deleted, vessel_type_id, created_at, updated_at) "
                f"VALUES (:id, :tid, :g1c, :g1n, :g2c, :g2n, :mc, :mn, :cc, :cn, "
                f":ct, false, :crit, :ver, 'active', false, :vtid, NOW(), NOW())"
            ),
            {
                "id": str(uuid.uuid4()),
                "tid": str(current_user.tenant_id),
                "g1c": mapped["group1_code"],
                "g1n": mapped["group1_name"],
                "g2c": mapped["group2_code"],
                "g2n": mapped["group2_name"],
                "mc": mapped["machinery_code"],
                "mn": mapped["machinery_name"],
                "cc": mapped["component_code"],
                "cn": mapped["component_name"],
                "ct": mapped["component_type"],
                "crit": mapped["criticality"],
                "ver": new_version,
                "vtid": vessel_type_id or None,
            },
        )
        imported += 1

    await db.commit()
    return {"imported": imported, "version": new_version}


@router.get(
    "/library/component-structure",
    summary="F-03: List all component structure library nodes for tenant",
)
async def list_component_structure(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    status_filter: Optional[str] = Query(None, alias="status"),
    vessel_type_id: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: str = Query("group1_name"),
    sort_order: str = Query("asc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
) -> dict[str, Any]:
    """List component structure library nodes, optionally filtered by status and vessel type."""
    where_clauses = ["tenant_id = :tid", "is_deleted = false"]
    params: dict[str, Any] = {"tid": str(current_user.tenant_id)}

    if status_filter:
        where_clauses.append("status = :st")
        params["st"] = status_filter
    if vessel_type_id:
        where_clauses.append("vessel_type_id = :vtid")
        params["vtid"] = vessel_type_id
    if search:
        where_clauses.append(
            "("
            "group1_name ILIKE :search_pat OR "
            "group2_name ILIKE :search_pat OR "
            "machinery_name ILIKE :search_pat OR "
            "component_name ILIKE :search_pat OR "
            "component_code ILIKE :search_pat OR "
            "component_type ILIKE :search_pat"
            ")"
        )
        params["search_pat"] = f"%{search}%"

    where_sql = " AND ".join(where_clauses)
    sort_columns = {
        "group1_name": "group1_name",
        "group2_name": "group2_name",
        "machinery_name": "machinery_name",
        "component_name": "component_name",
        "component_code": "component_code",
        "component_type": "component_type",
        "criticality": "criticality",
        "status": "status",
        "created_at": "created_at",
    }
    sort_column = sort_columns.get(sort_by, "group1_name")
    sort_direction = "DESC" if str(sort_order).lower() == "desc" else "ASC"

    count_result = await db.execute(
        text(f"SELECT COUNT(*) FROM component_structure_library WHERE {where_sql}"),
        params,
    )
    total: int = count_result.scalar_one()

    result = await db.execute(
        text(
            "SELECT id, group1_code, group1_name, group2_code, group2_name, "
            "machinery_code, machinery_name, component_code, component_name, "
            "component_type, is_critical, criticality, status, version, vessel_type_id, created_at "
            "FROM component_structure_library "
            f"WHERE {where_sql} "
            f"ORDER BY {sort_column} {sort_direction}, group1_name ASC, group2_name ASC, machinery_name ASC, component_name ASC "
            f"LIMIT {page_size} OFFSET {(page - 1) * page_size}"
        ),
        params,
    )
    rows = result.mappings().all()
    return {"items": [dict(r) for r in rows], "page": page, "page_size": page_size, "total": total}


@router.post(
    "/library/component-structure/nodes",
    status_code=status.HTTP_201_CREATED,
    summary="F-03: Add a single component structure node (pending approval)",
)
async def add_component_structure_node(
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """Create a single node with status=pending_approval."""
    comp_name = body.get("component_name", "").strip()
    if not comp_name:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="component_name is required",
        )

    new_id = uuid.uuid4()
    await db.execute(
        text(
            "INSERT INTO component_structure_library "
            "(id, tenant_id, group1_code, group1_name, group2_code, group2_name, "
            "machinery_code, machinery_name, component_code, component_name, "
            "component_type, is_critical, criticality, version, status, is_deleted, created_at, updated_at) "
            "VALUES (:id, :tid, :g1c, :g1n, :g2c, :g2n, :mc, :mn, :cc, :cn, "
            ":ct, false, :crit, 1, 'pending_approval', false, NOW(), NOW())"
        ),
        {
            "id": str(new_id),
            "tid": str(current_user.tenant_id),
            "g1c": body.get("group1_code") or None,
            "g1n": body.get("group1_name") or None,
            "g2c": body.get("group2_code") or None,
            "g2n": body.get("group2_name") or None,
            "mc": body.get("machinery_code") or None,
            "mn": body.get("machinery_name") or None,
            "cc": body.get("component_code") or None,
            "cn": comp_name,
            "ct": body.get("component_type") or None,
            "crit": str(body.get("criticality") or "non_critical"),
        },
    )
    await db.commit()
    return {"id": str(new_id), "status": "pending_approval"}


@router.get(
    "/library/component-structure/approval-requests",
    summary="F-03: List pending approval requests for component structure nodes",
)
async def list_approval_requests(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """List all component_structure_library nodes with status=pending_approval."""
    result = await db.execute(
        text(
            "SELECT id, group1_code, group1_name, group2_code, group2_name, "
            "machinery_code, machinery_name, component_code, component_name, "
            "component_type, is_critical, criticality, status, created_at "
            "FROM component_structure_library "
            "WHERE tenant_id = :tid AND status = 'pending_approval' AND is_deleted = false "
            "ORDER BY created_at DESC"
        ),
        {"tid": str(current_user.tenant_id)},
    )
    rows = result.mappings().all()
    return {"items": [dict(r) for r in rows]}


@router.post(
    "/library/component-structure/approval-requests/{request_id}/approve",
    summary="F-03: Approve a component structure node",
)
async def approve_node(
    request_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """Set a pending node's status to active."""
    check = await db.execute(
        text(
            "SELECT id FROM component_structure_library "
            "WHERE id = :id AND tenant_id = :tid AND is_deleted = false"
        ),
        {"id": str(request_id), "tid": str(current_user.tenant_id)},
    )
    if check.one_or_none() is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Node not found")

    await db.execute(
        text(
            "UPDATE component_structure_library "
            "SET status = 'active', updated_at = NOW() "
            "WHERE id = :id AND tenant_id = :tid"
        ),
        {"id": str(request_id), "tid": str(current_user.tenant_id)},
    )
    await db.commit()
    return {"id": str(request_id), "status": "active"}


@router.post(
    "/library/component-structure/push-to-vessel",
    summary="F-03: Push active library nodes to a vessel as accepted components (idempotent)",
)
async def push_library_to_vessel(
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """
    Copy all active component_structure_library nodes for this tenant into the
    vessel's components table with qc_status=accepted.  Already-existing rows
    (same group1/group2/main_machinery/component_name) are skipped so the
    operation is safe to call multiple times.
    """
    vessel_id_str = body.get("vessel_id")
    if not vessel_id_str:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="vessel_id is required",
        )
    try:
        vessel_id = uuid.UUID(vessel_id_str)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Invalid vessel_id UUID",
        )

    await _get_vessel_or_404(vessel_id, db)

    vessel_type_id = body.get("vessel_type_id")
    result = await load_library_components_for_vessel(
        db=db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        vessel_type_id=vessel_type_id,
    )
    await db.commit()
    return {
        "added": result.get("added", 0),
        "skipped": result.get("skipped", 0),
        "message": result.get("message"),
    }


@router.post(
    "/library/component-structure/approval-requests/{request_id}/reject",
    summary="F-03: Reject a component structure node",
)
async def reject_node(
    request_id: uuid.UUID,
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """Set a pending node's status to rejected and store the rejection reason."""
    check = await db.execute(
        text(
            "SELECT id FROM component_structure_library "
            "WHERE id = :id AND tenant_id = :tid AND is_deleted = false"
        ),
        {"id": str(request_id), "tid": str(current_user.tenant_id)},
    )
    if check.one_or_none() is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Node not found")

    reason = body.get("reason", "")
    await db.execute(
        text(
            "UPDATE component_structure_library "
            "SET status = 'rejected', rejection_reason = :reason, updated_at = NOW() "
            "WHERE id = :id AND tenant_id = :tid"
        ),
        {"id": str(request_id), "tid": str(current_user.tenant_id), "reason": reason},
    )
    await db.commit()
    return {"id": str(request_id), "status": "rejected", "reason": reason}


# ===========================================================================
# F-04: Global Libraries
# ===========================================================================

_GLOBAL_TABLE_MAP = {
    "component": "global_component_library",
    "job": "global_job_library",
    "spare": "global_spare_library",
}


@router.get(
    "/library/global/{entity_type}",
    summary="F-04: List global library entries (component/job/spare)",
)
async def list_global_library(
    entity_type: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    search: Optional[str] = Query(None),
    sort_by: str = Query("occurrence_count"),
    sort_order: str = Query("desc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> dict[str, Any]:
    """Return paginated records from the relevant global library table."""
    table = _GLOBAL_TABLE_MAP.get(entity_type)
    if not table:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid entity_type '{entity_type}'. Must be one of: component, job, spare.",
        )

    where_clauses = ["tenant_id = :tid", "is_deleted = false"]
    params: dict[str, Any] = {"tid": str(current_user.tenant_id)}
    if search:
        where_clauses.append("CAST(canonical_data AS TEXT) ILIKE :search_pat")
        params["search_pat"] = f"%{search}%"
    where_sql = " AND ".join(where_clauses)
    sort_columns = {
        "occurrence_count": "occurrence_count",
        "first_seen_at": "first_seen_at",
        "created_at": "created_at",
    }
    sort_column = sort_columns.get(sort_by, "occurrence_count")
    sort_direction = "DESC" if str(sort_order).lower() == "desc" else "ASC"

    if page == 1 and not search:
        try:
            await backfill_global_library_from_accepted_records(
                db,
                tenant_id=current_user.tenant_id,
                entity_type=entity_type,
            )
            await db.commit()
        except Exception as err:
            logger.warning("list_global_library backfill failed for %s: %s", entity_type, err)
            await db.rollback()

    count_result = await db.execute(text(f"SELECT COUNT(*) FROM {table} WHERE {where_sql}"), params)
    total = count_result.scalar_one()

    result = await db.execute(
        text(
            f"SELECT id, tenant_id, canonical_data, occurrence_count, "
            f"source_vessels, first_seen_at, created_at "
            f"FROM {table} "
            f"WHERE {where_sql} "
            f"ORDER BY {sort_column} {sort_direction}, occurrence_count DESC, first_seen_at DESC "
            f"LIMIT {page_size} OFFSET {(page - 1) * page_size}"
        ),
        params,
    )
    rows = result.mappings().all()
    return {
        "entity_type": entity_type,
        "items": [dict(r) for r in rows],
        "page": page,
        "page_size": page_size,
        "total": total,
    }


@router.delete(
    "/library/global/{entity_type}/{entry_id}",
    summary="Delete a global library entry (component/job/spare)",
)
async def delete_global_library_entry(
    entity_type: str,
    entry_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, str]:
    table = _GLOBAL_TABLE_MAP.get(entity_type)
    if not table:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid entity_type '{entity_type}'. Must be one of: component, job, spare.",
        )

    await db.execute(
        text(
            f"UPDATE {table} "
            "SET is_deleted = true, updated_at = NOW() "
            "WHERE id = :id AND tenant_id = :tid"
        ),
        {"id": str(entry_id), "tid": str(current_user.tenant_id)},
    )
    await db.commit()
    return {"status": "deleted"}


@router.post(
    "/library/global/{entity_type}/populate",
    summary="F-04: Populate global library from accepted vessel records",
)
async def populate_global_library(
    entity_type: str,
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """
    Find all accepted records for the given vessel, run dedup against the
    global library, and insert non-duplicates.
    """
    table = _GLOBAL_TABLE_MAP.get(entity_type)
    if not table:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid entity_type '{entity_type}'.",
        )

    vessel_id_str = body.get("vessel_id")
    if not vessel_id_str:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="vessel_id is required in request body",
        )
    try:
        vessel_id = uuid.UUID(vessel_id_str)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Invalid vessel_id UUID",
        )

    # ------------------------------------------------------------------
    # Load accepted entity records for the vessel
    # ------------------------------------------------------------------
    accepted_records: list[dict] = []
    if entity_type == "component":
        result = await db.execute(
            select(Component).where(
                Component.vessel_id == vessel_id,
                Component.tenant_id == current_user.tenant_id,
                Component.qc_status == CompQCStatus.accepted,
                Component.is_deleted == False,
            )
        )
        for c in result.scalars().all():
            if not (
                c.source_manual_id
                or c.pdf_reference
                or c.page_reference is not None
                or c.job_pages
                or c.spare_pages
            ):
                continue
            accepted_records.append(
                {
                    "id": str(c.id),
                    "component_name": c.component_name,
                    "maker": c.maker or "",
                    "model": c.model or "",
                    "group1": c.group1,
                    "group2": c.group2,
                    "main_machinery": c.main_machinery,
                }
            )
    elif entity_type == "job":
        result = await db.execute(
            select(Job).where(
                Job.vessel_id == vessel_id,
                Job.tenant_id == current_user.tenant_id,
                Job.qc_status == CompQCStatus.accepted,
                Job.source_manual_id.is_not(None),
                Job.is_deleted == False,
            )
        )
        for j in result.scalars().all():
            accepted_records.append(
                {
                    "id": str(j.id),
                    "job_name": j.job_name,
                    "frequency": j.frequency,
                    "frequency_type": j.frequency_type.value if j.frequency_type else None,
                }
            )
    elif entity_type == "spare":
        result = await db.execute(
            select(Spare).where(
                Spare.vessel_id == vessel_id,
                Spare.tenant_id == current_user.tenant_id,
                Spare.qc_status == CompQCStatus.accepted,
                Spare.source_manual_id.is_not(None),
                Spare.is_deleted == False,
            )
        )
        for s in result.scalars().all():
            accepted_records.append(
                {
                    "id": str(s.id),
                    "part_name": s.part_name,
                    "part_number": s.part_number or "",
                }
            )

    if not accepted_records:
        return {"added": 0, "duplicates": 0, "message": "No accepted records found for vessel."}

    # ------------------------------------------------------------------
    # Load existing global library entries for dedup comparison
    # ------------------------------------------------------------------
    existing_result = await db.execute(
        text(
            f"SELECT id, canonical_data, source_vessels, occurrence_count "
            f"FROM {table} "
            f"WHERE tenant_id = :tid AND is_deleted = false"
        ),
        {"tid": str(current_user.tenant_id)},
    )
    existing_rows = existing_result.mappings().all()
    existing_canonical: list[dict] = []
    for row in existing_rows:
        import json as _json
        raw_data = row["canonical_data"]
        if isinstance(raw_data, str):
            try:
                raw_data = _json.loads(raw_data)
            except Exception:
                raw_data = {}
        existing_canonical.append(
            {
                "id": str(row["id"]),
                "data": raw_data or {},
                "source_vessels": row["source_vessels"] or [],
                "occurrence_count": row["occurrence_count"] or 1,
            }
        )

    import json as _json
    added = 0
    duplicates = 0

    for record in accepted_records:
        is_dup = False
        for ex in existing_canonical:
            if entity_type == "component" and is_duplicate_component(record, ex["data"]):
                is_dup = True
                break
            elif entity_type == "job" and is_duplicate_job(record, ex["data"]):
                is_dup = True
                break
            elif entity_type == "spare" and is_duplicate_spare(record, ex["data"]):
                is_dup = True
                break

        if is_dup:
            duplicates += 1
        else:
            new_id = uuid.uuid4()
            canonical_json = _json.dumps(record)
            source_vessels_json = _json.dumps([vessel_id_str])
            await db.execute(
                text(
                    f"INSERT INTO {table} "
                    f"(id, tenant_id, canonical_data, occurrence_count, source_vessels, "
                    f"first_seen_at, is_deleted) "
                    f"VALUES (:id, :tid, CAST(:cd AS jsonb), 1, CAST(:sv AS jsonb), NOW(), false)"
                ),
                {
                    "id": str(new_id),
                    "tid": str(current_user.tenant_id),
                    "cd": canonical_json,
                    "sv": source_vessels_json,
                },
            )
            # Also add to local dedup list to prevent intra-batch duplicates
            existing_canonical.append(
                {
                    "id": str(new_id),
                    "data": record,
                    "source_vessels": [vessel_id_str],
                    "occurrence_count": 1,
                }
            )
            added += 1

    await db.commit()
    return {"added": added, "duplicates": duplicates}


# ===========================================================================
# F-08: Manual Matching
# ===========================================================================


@router.post(
    "/{vessel_id}/manuals/find-matches",
    summary="F-08: Find matching manuals in other vessel projects (same tenant)",
)
async def find_manual_matches(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """
    For each manual in the vessel, search for matching manuals across other
    vessel projects in the same tenant using SHA-256 exact match first, then
    filename fuzzy similarity. Results are stored in manual_matches table.
    """
    await _get_vessel_or_404(vessel_id, db)

    # Load this vessel's manuals
    own_result = await db.execute(
        select(Manual).where(
            Manual.vessel_id == vessel_id,
            Manual.tenant_id == current_user.tenant_id,
            Manual.is_deleted == False,
        )
    )
    own_manuals = own_result.scalars().all()
    if not own_manuals:
        return {"matches": [], "total": 0}

    # Load all OTHER vessel manuals for the same tenant
    other_result = await db.execute(
        select(Manual).where(
            Manual.vessel_id != vessel_id,
            Manual.tenant_id == current_user.tenant_id,
            Manual.is_deleted == False,
        )
    )
    other_manuals = other_result.scalars().all()

    # Build a lookup for vessel names
    vessel_result = await db.execute(
        select(VesselProject).where(
            VesselProject.tenant_id == current_user.tenant_id,
            VesselProject.is_deleted == False,
        )
    )
    vessel_name_map: dict[str, str] = {
        str(v.id): v.vessel_name for v in vessel_result.scalars().all()
    }

    from difflib import SequenceMatcher

    def _filename_sim(a: str, b: str) -> float:
        # Strip extension and normalise
        a_stem = a.rsplit(".", 1)[0].lower().replace("_", " ").replace("-", " ")
        b_stem = b.rsplit(".", 1)[0].lower().replace("_", " ").replace("-", " ")
        return SequenceMatcher(None, a_stem, b_stem).ratio()

    matches_created: list[dict] = []

    for own in own_manuals:
        for other in other_manuals:
            # --- SHA-256 exact match ---
            if own.sha256_hash and other.sha256_hash and own.sha256_hash == other.sha256_hash:
                score = 1.0
                confidence = "exact"
            else:
                # --- filename fuzzy match ---
                score = _filename_sim(own.original_filename, other.original_filename)
                if score < 0.60:
                    continue
                confidence = "high" if score >= 0.85 else "medium" if score >= 0.70 else "low"

            match_id = uuid.uuid4()
            matched_vessel_name = vessel_name_map.get(str(other.vessel_id), "Unknown Vessel")

            # Upsert into manual_matches (ignore duplicates)
            try:
                await db.execute(
                    text(
                        "INSERT INTO manual_matches "
                        "(id, tenant_id, source_manual_id, matched_manual_id, "
                        "match_score, match_confidence, is_deleted) "
                        "VALUES (:id, :tid, :smid, :mmid, :score, :conf, false) "
                        "ON CONFLICT DO NOTHING"
                    ),
                    {
                        "id": str(match_id),
                        "tid": str(current_user.tenant_id),
                        "smid": str(own.id),
                        "mmid": str(other.id),
                        "score": round(score, 4),
                        "conf": confidence,
                    },
                )
            except Exception:
                pass  # Table may not exist yet; gracefully skip

            matches_created.append(
                {
                    "source_manual_id": str(own.id),
                    "source_manual_name": own.original_filename,
                    "matched_manual_id": str(other.id),
                    "matched_manual_name": other.original_filename,
                    "matched_vessel_id": str(other.vessel_id),
                    "matched_vessel_name": matched_vessel_name,
                    "match_score": round(score, 4),
                    "match_confidence": confidence,
                }
            )

    await db.commit()
    return {"matches": matches_created, "total": len(matches_created)}


@router.get(
    "/{vessel_id}/manuals/matches",
    summary="F-08: List stored manual matches for a vessel",
)
async def list_manual_matches(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    search: Optional[str] = Query(None),
    sort_by: str = Query("match_score"),
    sort_order: str = Query("desc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> dict[str, Any]:
    """Return stored matches for manuals belonging to this vessel."""
    await _get_vessel_or_404(vessel_id, db)

    where_sql = (
        "sm.vessel_id = :vid "
        "AND mm.tenant_id = :tid "
        "AND mm.is_deleted = false"
    )
    params: dict[str, Any] = {"vid": str(vessel_id), "tid": str(current_user.tenant_id)}
    if search:
        where_sql += (
            " AND ("
            "sm.original_filename ILIKE :search_pat OR "
            "tm.original_filename ILIKE :search_pat OR "
            "COALESCE(vp.vessel_name, '') ILIKE :search_pat"
            ")"
        )
        params["search_pat"] = f"%{search}%"
    sort_columns = {
        "source_manual_name": "sm.original_filename",
        "matched_manual_name": "tm.original_filename",
        "matched_vessel_name": "vp.vessel_name",
        "match_score": "mm.match_score",
        "match_confidence": "mm.match_confidence",
        "created_at": "mm.created_at",
    }
    sort_column = sort_columns.get(sort_by, "mm.match_score")
    sort_direction = "DESC" if str(sort_order).lower() == "desc" else "ASC"

    try:
        count_result = await db.execute(
            text(
                "SELECT COUNT(*) "
                "FROM manual_matches mm "
                "JOIN manuals sm ON sm.id = mm.source_manual_id "
                "JOIN manuals tm ON tm.id = mm.matched_manual_id "
                "LEFT JOIN vessel_projects vp ON vp.id = tm.vessel_id "
                f"WHERE {where_sql}"
            ),
            params,
        )
        total = count_result.scalar_one()
        result = await db.execute(
            text(
                "SELECT mm.id, mm.source_manual_id, mm.matched_manual_id, "
                "mm.match_score, mm.match_confidence, mm.created_at, "
                "sm.original_filename AS source_manual_name, "
                "tm.original_filename AS matched_manual_name, "
                "tm.vessel_id AS matched_vessel_id, "
                "vp.vessel_name AS matched_vessel_name "
                "FROM manual_matches mm "
                "JOIN manuals sm ON sm.id = mm.source_manual_id "
                "JOIN manuals tm ON tm.id = mm.matched_manual_id "
                "LEFT JOIN vessel_projects vp ON vp.id = tm.vessel_id "
                f"WHERE {where_sql} "
                f"ORDER BY {sort_column} {sort_direction}, mm.match_score DESC "
                f"LIMIT {page_size} OFFSET {(page - 1) * page_size}"
            ),
            params,
        )
        rows = result.mappings().all()
    except Exception:
        rows = []
        total = 0

    return {"items": [dict(r) for r in rows], "page": page, "page_size": page_size, "total": total}


@router.post(
    "/{vessel_id}/manuals/matches/{match_id}/copy-all",
    summary="F-08: Copy all accepted records from matched vessel to current vessel",
)
async def copy_matched_vessel_records(
    vessel_id: uuid.UUID,
    match_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """
    Copy all accepted Components, Jobs, and Spares from the source vessel
    referenced in the match into this vessel as pending QC records.
    """
    await _get_vessel_or_404(vessel_id, db)

    # Resolve the match to find the source vessel
    try:
        match_result = await db.execute(
            text(
                "SELECT mm.source_manual_id, mm.matched_manual_id, "
                "sm.vessel_id AS source_vessel_id "
                "FROM manual_matches mm "
                "JOIN manuals sm ON sm.id = mm.source_manual_id "
                "WHERE mm.id = :mid AND mm.tenant_id = :tid AND mm.is_deleted = false"
            ),
            {"mid": str(match_id), "tid": str(current_user.tenant_id)},
        )
        match_row = match_result.mappings().one_or_none()
    except Exception:
        match_row = None

    if match_row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Match not found")

    source_vessel_id = uuid.UUID(str(match_row["source_vessel_id"]))

    copied_components = 0
    copied_jobs = 0
    copied_spares = 0

    # --- Copy Components ---
    comp_result = await db.execute(
        select(Component).where(
            Component.vessel_id == source_vessel_id,
            Component.tenant_id == current_user.tenant_id,
            Component.qc_status == CompQCStatus.accepted,
            Component.is_deleted == False,
        )
    )
    for c in comp_result.scalars().all():
        new_comp = Component(
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            group1=c.group1,
            group2=c.group2,
            main_machinery=c.main_machinery,
            component_name=c.component_name,
            maker=c.maker,
            model=c.model,
            serial_number=c.serial_number,
            specification=c.specification,
            is_critical=c.is_critical,
            job_pages=c.job_pages,
            spare_pages=c.spare_pages,
            confidence_score=c.confidence_score,
            qc_status=CompQCStatus.pending,
            source_manual_id=c.source_manual_id,
        )
        db.add(new_comp)
        copied_components += 1

    # --- Copy Jobs ---
    job_result = await db.execute(
        select(Job).where(
            Job.vessel_id == source_vessel_id,
            Job.tenant_id == current_user.tenant_id,
            Job.qc_status == CompQCStatus.accepted,
            Job.is_deleted == False,
        )
    )
    for j in job_result.scalars().all():
        new_job = Job(
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            job_name=j.job_name,
            job_code=j.job_code,
            job_description=j.job_description,
            safety_precaution=j.safety_precaution,
            frequency=j.frequency,
            frequency_type=j.frequency_type,
            is_critical=j.is_critical,
            confidence_score=j.confidence_score,
            qc_status=CompQCStatus.pending,
            source_manual_id=j.source_manual_id,
        )
        db.add(new_job)
        copied_jobs += 1

    # --- Copy Spares ---
    spare_result = await db.execute(
        select(Spare).where(
            Spare.vessel_id == source_vessel_id,
            Spare.tenant_id == current_user.tenant_id,
            Spare.qc_status == CompQCStatus.accepted,
            Spare.is_deleted == False,
        )
    )
    for s in spare_result.scalars().all():
        new_spare = Spare(
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            part_name=s.part_name,
            part_number=s.part_number,
            specification=s.specification,
            spare_maker=s.spare_maker,
            spare_model=s.spare_model,
            confidence_score=s.confidence_score,
            qc_status=CompQCStatus.pending,
            source_manual_id=s.source_manual_id,
        )
        db.add(new_spare)
        copied_spares += 1

    await db.commit()
    return {
        "copied_components": copied_components,
        "copied_jobs": copied_jobs,
        "copied_spares": copied_spares,
        "total": copied_components + copied_jobs + copied_spares,
    }
