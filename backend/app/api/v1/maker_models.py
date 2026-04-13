"""
Global Maker / Model library.

Stores a tenant-scoped catalogue of equipment makers and their models.
Supports bulk import from Excel/CSV (no duplicates) and lookup for dropdowns.
"""
from __future__ import annotations

import io
import time
import uuid
from typing import Annotated, Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import get_db
from app.deps import get_current_user
from app.models.user import User
from app.services.review_workflow import backfill_maker_models_from_accepted_records

router = APIRouter()
_MAKER_MODEL_BACKFILL_TTL_SECONDS = settings.HOT_PATH_MAINTENANCE_TTL_SECONDS
_maker_model_backfill_last_run: dict[str, float] = {}

# ---------------------------------------------------------------------------
# Runtime bootstrap — create table if migration hasn't run yet
# Run DDL in autocommit mode to avoid "cannot run inside a transaction" errors.
# ---------------------------------------------------------------------------

_bootstrapped: bool = False  # module-level flag — only runs once per process


def _should_backfill_maker_models(tenant_id: uuid.UUID) -> bool:
    now = time.time()
    key = str(tenant_id)
    last_run = _maker_model_backfill_last_run.get(key, 0.0)
    if now - last_run < _MAKER_MODEL_BACKFILL_TTL_SECONDS:
        return False
    _maker_model_backfill_last_run[key] = now
    return True


async def _bootstrap(db: AsyncSession) -> None:
    """
    Ensure maker_models table exists. Uses information_schema to check first
    so we never issue DDL that would put the session in an aborted state.
    """
    global _bootstrapped
    if _bootstrapped:
        return
    try:
        # Non-DDL check — safe inside any transaction
        result = await db.execute(text(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema = 'public' AND table_name = 'maker_models'"
        ))
        if result.scalar_one_or_none() is None:
            # Table genuinely doesn't exist — create it
            await db.execute(text("""
                CREATE TABLE maker_models (
                    id               UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                    tenant_id        UUID NOT NULL,
                    maker            VARCHAR(255) NOT NULL,
                    model            VARCHAR(255),
                    component_category VARCHAR(100),
                    is_system_generated BOOLEAN NOT NULL DEFAULT false,
                    is_deleted       BOOLEAN NOT NULL DEFAULT false,
                    created_at       TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    updated_at       TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """))
            await db.execute(text("""
                CREATE UNIQUE INDEX uq_maker_models_tenant_maker_model
                ON maker_models (tenant_id, maker, COALESCE(model, ''))
                WHERE is_deleted = false
            """))
        await db.execute(text("""
            ALTER TABLE maker_models
            ADD COLUMN IF NOT EXISTS is_system_generated BOOLEAN NOT NULL DEFAULT false
        """))
        await db.commit()
    except Exception:
        # Rollback is CRITICAL here — without it the session stays in an
        # "aborted transaction" state and every subsequent query returns 500.
        try:
            await db.rollback()
        except Exception:
            pass
    finally:
        _bootstrapped = True


# ---------------------------------------------------------------------------
# List makers (distinct, for dropdown)
# ---------------------------------------------------------------------------

@router.get("/maker-models/makers", summary="List distinct maker names")
async def list_makers(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    search: Optional[str] = Query(None),
) -> dict[str, Any]:
    await _bootstrap(db)
    if _should_backfill_maker_models(current_user.tenant_id):
        try:
            await backfill_maker_models_from_accepted_records(db, tenant_id=current_user.tenant_id)
            await db.commit()
        except Exception:
            await db.rollback()
    if search:
        query = text("""
            SELECT DISTINCT maker FROM maker_models
            WHERE tenant_id = :tid AND is_deleted = false
            AND maker ILIKE :search_pat
            ORDER BY maker
            LIMIT 200
        """)
        result = await db.execute(query, {
            "tid": str(current_user.tenant_id),
            "search_pat": f"%{search}%",
        })
    else:
        query = text("""
            SELECT DISTINCT maker FROM maker_models
            WHERE tenant_id = :tid AND is_deleted = false
            ORDER BY maker
            LIMIT 200
        """)
        result = await db.execute(query, {
            "tid": str(current_user.tenant_id),
        })
    makers = [row[0] for row in result.fetchall()]
    return {"items": makers, "total": len(makers)}


# ---------------------------------------------------------------------------
# List models for a given maker
# ---------------------------------------------------------------------------

@router.get("/maker-models/models", summary="List models for a maker")
async def list_models(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    maker: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
) -> dict[str, Any]:
    await _bootstrap(db)
    if _should_backfill_maker_models(current_user.tenant_id):
        try:
            await backfill_maker_models_from_accepted_records(db, tenant_id=current_user.tenant_id)
            await db.commit()
        except Exception:
            await db.rollback()
    # Build query dynamically to avoid None parameter issues
    base_where = "WHERE tenant_id = :tid AND is_deleted = false AND model IS NOT NULL"
    params = {"tid": str(current_user.tenant_id)}
    
    if maker:
        base_where += " AND maker ILIKE :maker_pat"
        params["maker_pat"] = f"%{maker}%"
    
    if search:
        base_where += " AND model ILIKE :search_pat"
        params["search_pat"] = f"%{search}%"
    
    query = text(f"""
        SELECT DISTINCT model FROM maker_models
        {base_where}
        ORDER BY model
        LIMIT 200
    """)
    result = await db.execute(query, params)
    models = [row[0] for row in result.fetchall()]
    return {"items": models, "total": len(models)}


# ---------------------------------------------------------------------------
# List all entries
# ---------------------------------------------------------------------------

@router.get("/maker-models", summary="List maker/model library entries")
async def list_maker_models(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    search: Optional[str] = Query(None),
    sort_by: str = Query("maker"),
    sort_order: str = Query("asc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
) -> dict[str, Any]:
    await _bootstrap(db)
    if _should_backfill_maker_models(current_user.tenant_id):
        try:
            await backfill_maker_models_from_accepted_records(db, tenant_id=current_user.tenant_id)
            await db.commit()
        except Exception:
            await db.rollback()
    params: dict[str, Any] = {"tid": str(current_user.tenant_id)}
    where_sql = "tenant_id = :tid AND is_deleted = false"
    if search:
        where_sql += " AND (maker ILIKE :pat1 OR model ILIKE :pat2)"
        params["pat1"] = f"%{search}%"
        params["pat2"] = f"%{search}%"

    count_q = text(f"""
        SELECT COUNT(*) FROM maker_models
        WHERE {where_sql}
    """)
    total = (await db.execute(count_q, params)).scalar_one()

    sort_columns = {
        "maker": "maker",
        "model": "model",
        "component_category": "component_category",
        "created_at": "created_at",
    }
    sort_column = sort_columns.get(sort_by, "maker")
    sort_direction = "DESC" if str(sort_order).lower() == "desc" else "ASC"

    rows_q = text("""
        SELECT id, maker, model, component_category, created_at FROM maker_models
        WHERE """ + where_sql + """
        ORDER BY """ + sort_column + " " + sort_direction + """, maker ASC, model ASC
        LIMIT :lim OFFSET :off
    """)
    row_params = {
        **params,
        "lim": page_size,
        "off": (page - 1) * page_size,
    }
    rows = (await db.execute(rows_q, row_params)).fetchall()

    items = [{"id": str(r[0]), "maker": r[1], "model": r[2], "component_category": r[3], "created_at": str(r[4])} for r in rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


# ---------------------------------------------------------------------------
# Add single entry
# ---------------------------------------------------------------------------

@router.post("/maker-models", summary="Add a maker/model entry", status_code=201)
async def add_maker_model(
    body: dict,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _bootstrap(db)
    maker = (body.get("maker") or "").strip()
    model = (body.get("model") or "").strip() or None
    category = (body.get("component_category") or "").strip() or None
    if not maker:
        raise HTTPException(status_code=400, detail="maker is required")

    # ON CONFLICT DO NOTHING handles the partial unique index (is_deleted=false)
    # without repeating named params (which breaks asyncpg's positional mapping)
    await db.execute(text("""
        INSERT INTO maker_models (id, tenant_id, maker, model, component_category, is_system_generated)
        VALUES (gen_random_uuid(), :tid, :maker, :model, :cat, false)
        ON CONFLICT DO NOTHING
    """), {"tid": str(current_user.tenant_id), "maker": maker, "model": model, "cat": category})
    await db.commit()
    return {"status": "ok", "maker": maker, "model": model}


# ---------------------------------------------------------------------------
# Bulk import from Excel / CSV
# ---------------------------------------------------------------------------

@router.post("/maker-models/import", summary="Bulk import makers/models from Excel or CSV")
async def import_maker_models(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """
    Accepts .xlsx or .csv with columns: Maker, Model (optional), Category (optional).
    Skips duplicates. Returns counts of imported and skipped rows.
    """
    await _bootstrap(db)
    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []
    try:
        if filename.endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for excel_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(excel_row) if i < len(headers)})
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    # Normalise column names
    def _get(row: dict, *keys: str) -> str:
        for k in keys:
            for rk in row:
                if rk.strip().lower() == k.lower():
                    return (row[rk] or "").strip()
        return ""

    imported = 0
    skipped = 0
    tid = str(current_user.tenant_id)

    for row in rows:
        maker = _get(row, "maker", "manufacturer", "make")
        model = _get(row, "model", "type", "model_no") or None
        category = _get(row, "category", "component_category", "type") or None
        if not maker:
            skipped += 1
            continue

        result = await db.execute(text("""
            INSERT INTO maker_models (id, tenant_id, maker, model, component_category, is_system_generated)
            VALUES (gen_random_uuid(), :tid, :maker, :model, :cat, false)
            ON CONFLICT DO NOTHING
        """), {"tid": tid, "maker": maker, "model": model, "cat": category})
        if result.rowcount > 0:
            imported += 1
        else:
            skipped += 1

    await db.commit()
    return {"imported": imported, "skipped": skipped, "total_rows": len(rows)}


# ---------------------------------------------------------------------------
# Delete entry
# ---------------------------------------------------------------------------

@router.delete("/maker-models/{entry_id}", summary="Delete a maker/model entry")
async def delete_maker_model(
    entry_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    await db.execute(text("""
        UPDATE maker_models SET is_deleted = true, updated_at = NOW()
        WHERE id = :id AND tenant_id = :tid
    """), {"id": str(entry_id), "tid": str(current_user.tenant_id)})
    await db.commit()
    return {"status": "deleted"}
