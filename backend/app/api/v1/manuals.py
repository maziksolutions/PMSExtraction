from __future__ import annotations

import base64
import io
import os
import re
import uuid
from typing import Annotated, Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.models.feedback import CorrectionType, FeedbackEntry
from app.models.ingestion import Manual, ManualStatus
from app.models.user import User
from app.models.vessel import VesselProject
from app.schemas.manual import ManualOut, ManualUpdate
from app.services.feedback_learning import schedule_feedback_learning
from app.services.review_workflow import broadcast_activity, log_activity
from app.services.upload_security import validate_uploaded_file_bytes

router = APIRouter()

VESSEL_TYPE_MANUAL_TEMPLATES = {
    "Bulk Carrier": [
        "Instruction Manual",
        "Machinery Particulars",
        "General Arrangement",
        "Pipeline Diagrams/P&ID",
        "Electrical Diagrams",
    ],
    "Tanker": [
        "Instruction Manual",
        "Machinery Particulars",
        "General Arrangement",
        "Pipeline Diagrams/P&ID",
        "LSA/FFA Plans",
        "Tank Capacity Plan",
        "Electrical Diagrams",
    ],
    "Container Ship": [
        "Instruction Manual",
        "Machinery Particulars",
        "General Arrangement",
        "Electrical Diagrams",
        "Class Certificates/Surveys",
    ],
}


async def _get_vessel_or_404(vessel_id: uuid.UUID, db: AsyncSession) -> VesselProject:
    result = await db.execute(
        select(VesselProject).where(
            VesselProject.id == vessel_id,
            VesselProject.is_deleted == False,
        )
    )
    vessel = result.scalar_one_or_none()
    if vessel is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vessel not found")
    return vessel


def _build_manual_filters(
    *,
    vessel_id: uuid.UUID,
    tenant_id: uuid.UUID,
    category: Optional[str] = None,
    manual_status: Optional[str] = None,
    min_confidence: Optional[int] = None,
    search: Optional[str] = None,
    useful_for_extraction: Optional[str] = None,
) -> list[Any]:
    base_filter: list[Any] = [
        Manual.vessel_id == vessel_id,
        Manual.tenant_id == tenant_id,
        Manual.is_deleted == False,
    ]

    if category:
        base_filter.append(Manual.category == category)
    if manual_status:
        try:
            base_filter.append(Manual.status == ManualStatus(manual_status))
        except ValueError:
            pass
    if min_confidence is not None:
        base_filter.append(Manual.classification_confidence >= min_confidence)
    if search:
        base_filter.append(Manual.original_filename.ilike(f"%{search}%"))
    if useful_for_extraction:
        base_filter.append(Manual.useful_for_extraction == useful_for_extraction)

    return base_filter


def _manual_order_expr(sort_by: str, sort_order: str):
    if sort_by == "filename":
        order_col = Manual.original_filename
    elif sort_by == "created_at":
        order_col = Manual.created_at
    else:
        order_col = Manual.classification_confidence
    return order_col.asc() if sort_order == "asc" else order_col.desc()


def _prepare_manual_update_data(
    manual: Manual,
    update_data: dict[str, Any],
) -> tuple[dict[str, Any], bool]:
    prepared = dict(update_data)
    page_field_groups = [
        ("pages_with_components", "pages_with_components_printed", "pages_with_components_physical"),
        ("pages_with_jobs", "pages_with_jobs_printed", "pages_with_jobs_physical"),
        ("pages_with_spares", "pages_with_spares_printed", "pages_with_spares_physical"),
    ]
    page_fields_changed = False

    for canonical_field, printed_field, physical_field in page_field_groups:
        if printed_field in prepared or physical_field in prepared:
            printed_value = prepared.get(printed_field, getattr(manual, printed_field))
            physical_value = prepared.get(physical_field, getattr(manual, physical_field))
            prepared[canonical_field] = printed_value or physical_value or ""
            page_fields_changed = True
        elif canonical_field in prepared:
            canonical_value = prepared[canonical_field]
            prepared.setdefault(printed_field, canonical_value)

    return prepared, page_fields_changed


def _original_manual_snapshot(manual: Manual) -> dict[str, Any]:
    return {
        "category": manual.category,
        "useful_for_extraction": manual.useful_for_extraction,
        "classification_confidence": manual.classification_confidence,
        "pages_with_components": manual.pages_with_components,
        "pages_with_jobs": manual.pages_with_jobs,
        "pages_with_spares": manual.pages_with_spares,
        "pages_with_components_printed": manual.pages_with_components_printed,
        "pages_with_jobs_printed": manual.pages_with_jobs_printed,
        "pages_with_spares_printed": manual.pages_with_spares_printed,
        "pages_with_components_physical": manual.pages_with_components_physical,
        "pages_with_jobs_physical": manual.pages_with_jobs_physical,
        "pages_with_spares_physical": manual.pages_with_spares_physical,
        "page_explanations": manual.page_explanations,
        "reviewer_comments": manual.reviewer_comments,
        "supply_type": manual.supply_type,
    }


async def _apply_manual_updates(
    *,
    manual: Manual,
    update_data: dict[str, Any],
    current_user: User,
    db: AsyncSession,
) -> tuple[dict[str, Any], uuid.UUID | None]:
    prepared, page_fields_changed = _prepare_manual_update_data(manual, update_data)
    original = _original_manual_snapshot(manual)

    if page_fields_changed and "page_explanations" not in prepared:
        # Imported/manual overrides replace model-selected page refs, so do not leave stale reasons behind.
        prepared["page_explanations"] = None

    for field, value in prepared.items():
        setattr(manual, field, value)

    db.add(manual)

    feedback_id: uuid.UUID | None = None
    corrected = {k: v for k, v in prepared.items()}
    if corrected:
        feedback = FeedbackEntry(
            tenant_id=current_user.tenant_id,
            manual_id=manual.id,
            entity_type="manual_classification",
            original_value=original,
            corrected_value=corrected,
            correction_type=CorrectionType.wrong_value,
            vessel_type=None,
            source_manual_category=prepared.get("category") or manual.category,
            context_span=(
                corrected.get("reviewer_comments")
                if isinstance(corrected.get("reviewer_comments"), str)
                else f"Updated manual screening fields: {', '.join(sorted(corrected.keys()))}"
            ),
            created_by=current_user.id,
        )
        db.add(feedback)
        await db.flush()
        feedback_id = feedback.id

    return prepared, feedback_id


def _normalise_excel_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    cleaned = str(value).strip()
    return cleaned or None


def _parse_optional_int(value: Any, *, field_name: str) -> Optional[int]:
    cleaned = _normalise_excel_value(value)
    if cleaned is None:
        return None
    try:
        return int(float(cleaned))
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a whole number") from exc


def _normalise_import_field_value(field: str, value: Any) -> Any:
    cleaned = _normalise_excel_value(value)
    if cleaned is None:
        return None

    lowered = cleaned.lower()
    if field == "useful_for_extraction":
        allowed = {"yes", "partial", "no"}
        if lowered not in allowed:
            raise ValueError("Useful For Extraction must be yes, partial, or no")
        return lowered
    if field == "supply_type":
        if lowered in {"yard supply", "yard_supply", "yard-supply"}:
            return "yard_supply"
        if lowered == "oem":
            return "OEM"
        raise ValueError("Source must be OEM or yard_supply")

    return cleaned


def _manual_export_headers() -> list[str]:
    return [
        "Manual ID",
        "File Name",
        "Size (MB)",
        "Status",
        "Category",
        "Useful For Extraction",
        "Source",
        "Confidence (%)",
        "Components Printed Page Ref",
        "Components Physical Page Ref",
        "Job Printed Page Ref",
        "Job Physical Page Ref",
        "Spare Printed Page Ref",
        "Spare Physical Page Ref",
        "Reviewer Comments",
    ]


def _normalise_header(value: str) -> str:
    return " ".join(value.lower().replace("_", " ").split())


def _manual_import_aliases() -> dict[str, str]:
    return {
        "manual id": "manual_id",
        "id": "manual_id",
        "file name": "original_filename",
        "filename": "original_filename",
        "category": "category",
        "useful for extraction": "useful_for_extraction",
        "useful": "useful_for_extraction",
        "source": "supply_type",
        "supply type": "supply_type",
        "confidence (%)": "classification_confidence",
        "confidence": "classification_confidence",
        "components printed page ref": "pages_with_components_printed",
        "components physical page ref": "pages_with_components_physical",
        "job printed page ref": "pages_with_jobs_printed",
        "job physical page ref": "pages_with_jobs_physical",
        "spare printed page ref": "pages_with_spares_printed",
        "spare physical page ref": "pages_with_spares_physical",
        "reviewer comments": "reviewer_comments",
    }


_MANUAL_PAGE_BLOCK_RE = re.compile(r"\[PAGE\s+(\d+)\]\s*\n?(.*?)(?=(?:\n\[PAGE\s+\d+\])|\Z)", re.S)


def _parse_preview_pages(value: str) -> list[int]:
    pages: set[int] = set()
    for token in value.split(","):
        cleaned = token.strip()
        if not cleaned:
            continue
        if "-" in cleaned:
            start_raw, end_raw = cleaned.split("-", 1)
            try:
                start = int(start_raw.strip())
                end = int(end_raw.strip())
            except ValueError:
                continue
            for page in range(min(start, end), max(start, end) + 1):
                pages.add(page)
            continue
        try:
            pages.add(int(cleaned))
        except ValueError:
            continue
    return sorted(page for page in pages if page > 0)


def _manual_text_by_page(manual: Manual) -> dict[int, str]:
    text = manual.extracted_text or ""
    if not text:
        return {}
    pages: dict[int, str] = {}
    for match in _MANUAL_PAGE_BLOCK_RE.finditer(text):
        pages[int(match.group(1))] = match.group(2).strip()
    if pages:
        return pages
    return {1: text.strip()}


async def _download_manual_bytes(manual: Manual) -> bytes:
    from app.services.blob_storage import BlobStorageService

    blob_key = manual.blob_storage_key
    if not blob_key:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Manual file is not available.")

    if os.path.exists(blob_key):
        with open(blob_key, "rb") as fh:
            return fh.read()

    blob_service = BlobStorageService()
    return await blob_service.download_bytes(blob_key)


def _encode_png_data_url(image: Any) -> str:
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    return f"data:image/png;base64,{base64.b64encode(buffer.getvalue()).decode('ascii')}"


def _build_manual_workbook(manuals: list[Manual], *, include_sample_rows: bool) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manual Review"

    headers = _manual_export_headers()
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows: list[list[Any]] = []
    if include_sample_rows:
        rows.extend(
            [
                [
                    "",
                    "02. SEWAGE TREATMENT PLANT INSTRUCTION MANUAL.pdf",
                    3.3,
                    "classified",
                    "Instruction Manual",
                    "yes",
                    "OEM",
                    92,
                    "1",
                    "1",
                    "9-14",
                    "10-15",
                    "",
                    "",
                    "Sample: component starts on physical page 1 with no printed footer.",
                ],
                [
                    "",
                    "SEWAGE TREATMENT PLANT.pdf",
                    1.1,
                    "classified",
                    "Yard/Finished Drawings",
                    "partial",
                    "yard_supply",
                    88,
                    "1",
                    "2",
                    "",
                    "",
                    "2-12",
                    "3-13",
                    "Sample: drawing package with printed pages offset by one physical page.",
                ],
            ]
        )
    else:
        for manual in manuals:
            rows.append(
                [
                    str(manual.id),
                    manual.original_filename,
                    round((manual.file_size_bytes or 0) / (1024 * 1024), 2),
                    manual.status.value if hasattr(manual.status, "value") else str(manual.status),
                    manual.category or "",
                    manual.useful_for_extraction or "",
                    manual.supply_type or "",
                    manual.classification_confidence or "",
                    manual.pages_with_components_printed or "",
                    manual.pages_with_components_physical or "",
                    manual.pages_with_jobs_printed or "",
                    manual.pages_with_jobs_physical or "",
                    manual.pages_with_spares_printed or "",
                    manual.pages_with_spares_physical or "",
                    manual.reviewer_comments or "",
                ]
            )

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 4, 16), 40)

    info = wb.create_sheet("Instructions")
    info["A1"] = "Manual Review Import Guidance"
    info["A1"].font = Font(bold=True, size=14)
    guidance_rows = [
        "1. Use the Export Screening file for real vessel updates because it includes Manual ID values.",
        "2. Edit only the review columns: Category, Useful For Extraction, Source, confidence, page refs, and Reviewer Comments.",
        "3. Components / Job / Spare page refs accept comma-separated numbers and ranges like 1, 3-5, 9.",
        "4. Printed refs are the page numbers shown inside the document. Physical refs are the actual PDF page positions.",
        "5. Import overwrites the current screening values for matched manuals.",
        "6. If page refs are changed by import, old model page explanations are cleared so reviewers do not see stale reasons.",
        "7. Source should be OEM or yard_supply.",
        "8. Useful For Extraction should be yes, partial, or no.",
        "9. This template sheet is only for guidance. Start from Export Screening when updating a real vessel.",
    ]
    for idx, text in enumerate(guidance_rows, start=3):
        info.cell(row=idx, column=1, value=text)
    info.column_dimensions["A"].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get(
    "/{vessel_id}/manuals",
    summary="List manuals for a vessel with optional filters",
)
async def list_manuals(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    category: Optional[str] = Query(None),
    manual_status: Optional[str] = Query(None, alias="status"),
    min_confidence: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    useful_for_extraction: Optional[str] = Query(None),
    sort_by: str = Query("filename", regex="^(filename|created_at|confidence)$"),
    sort_order: str = Query("asc", regex="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    base_filter = _build_manual_filters(
        vessel_id=vessel_id,
        tenant_id=current_user.tenant_id,
        category=category,
        manual_status=manual_status,
        min_confidence=min_confidence,
        search=search,
        useful_for_extraction=useful_for_extraction,
    )

    # Count query
    count_result = await db.execute(select(func.count()).select_from(Manual).where(*base_filter))
    total = count_result.scalar_one()

    order_expr = _manual_order_expr(sort_by, sort_order)

    query = (
        select(Manual)
        .where(*base_filter)
        .order_by(order_expr)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(query)
    manuals = result.scalars().all()
    return {
        "items": [ManualOut.model_validate(m) for m in manuals],
        "page": page,
        "page_size": page_size,
        "total": total,
    }


@router.get(
    "/{vessel_id}/manuals/{manual_id}/page-preview",
    summary="Return rendered manual pages and extracted text for review",
)
async def preview_manual_pages(
    vessel_id: uuid.UUID,
    manual_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    pages: str = Query(..., description="Comma-separated physical page numbers, e.g. 1,3-5"),
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    result = await db.execute(
        select(Manual).where(
            Manual.id == manual_id,
            Manual.vessel_id == vessel_id,
            Manual.tenant_id == current_user.tenant_id,
            Manual.is_deleted == False,
        )
    )
    manual = result.scalar_one_or_none()
    if manual is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Manual not found")

    requested_pages = _parse_preview_pages(pages)
    if not requested_pages:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Please provide at least one valid page number.")
    if len(requested_pages) > 12:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Preview up to 12 pages at a time.")

    page_text_lookup = _manual_text_by_page(manual)
    preview_pages: list[dict[str, Any]] = []
    page_count = manual.page_count

    if (manual.file_extension or "").lower() == "pdf":
        file_bytes = await _download_manual_bytes(manual)
        try:
            import pdfplumber

            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                page_count = page_count or len(pdf.pages)
                for page_number in requested_pages:
                    page_payload: dict[str, Any] = {
                        "page_number": page_number,
                        "text_excerpt": (page_text_lookup.get(page_number) or "")[:4000],
                        "image_data_url": None,
                    }
                    if 1 <= page_number <= len(pdf.pages):
                        page = pdf.pages[page_number - 1]
                        try:
                            page_payload["image_data_url"] = _encode_png_data_url(
                                page.to_image(resolution=140).original
                            )
                        except Exception:
                            page_payload["image_data_url"] = None
                        if not page_payload["text_excerpt"]:
                            try:
                                page_payload["text_excerpt"] = (page.extract_text() or "")[:4000]
                            except Exception:
                                page_payload["text_excerpt"] = ""
                    else:
                        page_payload["error"] = "Requested page is outside the PDF page count."
                    preview_pages.append(page_payload)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail=f"Could not render preview pages: {exc}",
            ) from exc
    else:
        for page_number in requested_pages:
            preview_pages.append(
                {
                    "page_number": page_number,
                    "text_excerpt": (page_text_lookup.get(page_number) or "")[:4000],
                    "image_data_url": None,
                }
            )

    return {
        "manual_id": str(manual.id),
        "file_name": manual.original_filename,
        "file_extension": manual.file_extension,
        "page_count": page_count,
        "pages": preview_pages,
    }


@router.patch(
    "/{vessel_id}/manuals/{manual_id}",
    response_model=ManualOut,
    summary="Update manual classification fields",
)
async def update_manual(
    vessel_id: uuid.UUID,
    manual_id: uuid.UUID,
    body: ManualUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ManualOut:
    await _get_vessel_or_404(vessel_id, db)

    result = await db.execute(
        select(Manual).where(
            Manual.id == manual_id,
            Manual.vessel_id == vessel_id,
            Manual.is_deleted == False,
        )
    )
    manual: Manual | None = result.scalar_one_or_none()
    if manual is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Manual not found")

    update_data = body.model_dump(exclude_unset=True)
    prepared, feedback_id = await _apply_manual_updates(
        manual=manual,
        update_data=update_data,
        current_user=current_user,
        db=db,
    )
    activity = None
    if prepared:
        activity = await log_activity(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            user_id=current_user.id,
            action_type="manual.corrected",
            entity_type="manual",
            entity_id=manual.id,
            description=f"Updated screening details for '{manual.original_filename}'.",
            metadata={"fields": sorted(prepared.keys())},
        )
    await db.commit()
    if feedback_id:
        await schedule_feedback_learning(feedback_id)
    await db.refresh(manual)
    if activity:
        await broadcast_activity(activity)
    return ManualOut.model_validate(manual)


@router.get(
    "/{vessel_id}/manuals/export-screening",
    summary="Export manual review screening as Excel",
)
async def export_manual_screening(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    category: Optional[str] = Query(None),
    manual_status: Optional[str] = Query(None, alias="status"),
    min_confidence: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    useful_for_extraction: Optional[str] = Query(None),
    sort_by: str = Query("filename", regex="^(filename|created_at|confidence)$"),
    sort_order: str = Query("asc", regex="^(asc|desc)$"),
) -> StreamingResponse:
    await _get_vessel_or_404(vessel_id, db)

    base_filter = _build_manual_filters(
        vessel_id=vessel_id,
        tenant_id=current_user.tenant_id,
        category=category,
        manual_status=manual_status,
        min_confidence=min_confidence,
        search=search,
        useful_for_extraction=useful_for_extraction,
    )
    result = await db.execute(
        select(Manual)
        .where(*base_filter)
        .order_by(_manual_order_expr(sort_by, sort_order))
    )
    manuals = result.scalars().all()

    response = _build_manual_workbook(manuals, include_sample_rows=False)
    response.headers["Content-Disposition"] = f'attachment; filename="manual_review_export_{vessel_id}.xlsx"'
    return response


@router.get(
    "/{vessel_id}/manuals/screening-template",
    summary="Download manual review import template",
)
async def download_manual_screening_template(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> StreamingResponse:
    await _get_vessel_or_404(vessel_id, db)
    response = _build_manual_workbook([], include_sample_rows=True)
    response.headers["Content-Disposition"] = 'attachment; filename="manual_review_template.xlsx"'
    return response


@router.post(
    "/{vessel_id}/manuals/import-screening",
    summary="Import manual review screening from Excel",
)
async def import_manual_screening(
    vessel_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Please upload an .xlsx file exported from Manual Review.",
        )

    content = await file.read()
    validate_uploaded_file_bytes(
        filename=file.filename or "manual_review_import.xlsx",
        content=content,
        allowed_extensions={"xlsx"},
        max_size_bytes=10 * 1024 * 1024,
    )
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook["Manual Review"] if "Manual Review" in workbook.sheetnames else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Could not parse Excel file: {exc}") from exc

    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The uploaded workbook is empty.")

    raw_headers = [str(cell or "").strip() for cell in rows[0]]
    aliases = _manual_import_aliases()
    mapped_headers = [aliases.get(_normalise_header(header), "") for header in raw_headers]

    if "manual_id" not in mapped_headers and "original_filename" not in mapped_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The workbook must contain 'Manual ID' or 'File Name' columns.",
        )

    result = await db.execute(
        select(Manual).where(
            Manual.vessel_id == vessel_id,
            Manual.tenant_id == current_user.tenant_id,
            Manual.is_deleted == False,
        )
    )
    manuals = result.scalars().all()
    manuals_by_id = {str(manual.id): manual for manual in manuals}
    manuals_by_filename = {manual.original_filename.strip().lower(): manual for manual in manuals}

    editable_fields = {
        "category",
        "useful_for_extraction",
        "supply_type",
        "classification_confidence",
        "pages_with_components_printed",
        "pages_with_components_physical",
        "pages_with_jobs_printed",
        "pages_with_jobs_physical",
        "pages_with_spares_printed",
        "pages_with_spares_physical",
        "reviewer_comments",
    }

    updated = 0
    skipped = 0
    errors: list[str] = []
    activities = []
    feedback_ids: list[uuid.UUID] = []

    for row_number, row in enumerate(rows[1:], start=2):
        row_map = {
            mapped_headers[index]: row[index]
            for index in range(min(len(mapped_headers), len(row)))
            if mapped_headers[index]
        }
        if not row_map or not any(_normalise_excel_value(value) for value in row_map.values()):
            continue

        manual = None
        manual_id_value = _normalise_excel_value(row_map.get("manual_id"))
        if manual_id_value:
            manual = manuals_by_id.get(manual_id_value)
        if manual is None:
            filename_value = _normalise_excel_value(row_map.get("original_filename"))
            if filename_value:
                manual = manuals_by_filename.get(filename_value.lower())

        if manual is None:
            skipped += 1
            errors.append(f"Row {row_number}: manual not found for the provided Manual ID / File Name.")
            continue

        try:
            update_data: dict[str, Any] = {}
            for field in editable_fields:
                if field not in row_map:
                    continue
                if field == "classification_confidence":
                    update_data[field] = _parse_optional_int(row_map[field], field_name="Confidence (%)")
                else:
                    update_data[field] = _normalise_import_field_value(field, row_map[field])

            if not update_data:
                skipped += 1
                continue

            prepared, feedback_id = await _apply_manual_updates(
                manual=manual,
                update_data=update_data,
                current_user=current_user,
                db=db,
            )
            if feedback_id:
                feedback_ids.append(feedback_id)
            activities.append(
                await log_activity(
                    db,
                    tenant_id=current_user.tenant_id,
                    vessel_id=vessel_id,
                    user_id=current_user.id,
                    action_type="manual.imported_screening",
                    entity_type="manual",
                    entity_id=manual.id,
                    description=f"Imported screening updates for '{manual.original_filename}'.",
                    metadata={"fields": sorted(prepared.keys())},
                )
            )
            updated += 1
        except ValueError as exc:
            skipped += 1
            errors.append(f"Row {row_number}: {exc}")

    await db.commit()
    for feedback_id in feedback_ids:
        await schedule_feedback_learning(feedback_id)
    for activity in activities:
        await broadcast_activity(activity)

    return {
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:50],
        "message": f"Imported screening updates for {updated} manual(s). Imported values overwrite existing screening fields.",
    }


@router.post(
    "/{vessel_id}/manuals/{manual_id}/trigger-classification",
    summary="Re-run classification for a manual",
)
async def trigger_classification(
    vessel_id: uuid.UUID,
    manual_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    result = await db.execute(
        select(Manual).where(
            Manual.id == manual_id,
            Manual.vessel_id == vessel_id,
            Manual.is_deleted == False,
        )
    )
    manual: Manual | None = result.scalar_one_or_none()
    if manual is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Manual not found")

    try:
        from app.tasks.classification import classify_manual

        task = classify_manual.delay(str(manual_id))
        return {"task_id": task.id, "status": "queued"}
    except Exception:
        return {"status": "queued", "task_id": "mock"}


@router.get(
    "/{vessel_id}/manuals/missing-report",
    summary="Missing manual gap analysis for vessel type",
)
async def missing_manual_report(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    vessel = await _get_vessel_or_404(vessel_id, db)

    expected = VESSEL_TYPE_MANUAL_TEMPLATES.get(vessel.vessel_type or "", [])

    result = await db.execute(
        select(Manual.category).where(
            Manual.vessel_id == vessel_id,
            Manual.tenant_id == current_user.tenant_id,
            Manual.is_deleted == False,
            Manual.category.isnot(None),
        )
    )
    classified_categories = {row[0] for row in result.all()}

    gaps = []
    for cat in expected:
        if cat not in classified_categories:
            gaps.append(
                {
                    "category": cat,
                    "status": "missing",
                    "message": f"No {cat} found for this vessel.",
                }
            )

    return {
        "vessel_type": vessel.vessel_type,
        "expected_categories": expected,
        "found_categories": list(classified_categories),
        "gaps": gaps,
        "gap_count": len(gaps),
    }
