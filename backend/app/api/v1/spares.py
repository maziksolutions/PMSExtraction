from __future__ import annotations

import re
import uuid
from typing import Annotated, Any, List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, Query, Response, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.models.component import QCStatus
from app.models.feedback import CorrectionType, FeedbackEntry
from app.models.spare import ExtractionMethod, Spare
from app.models.user import User
from app.models.vessel import VesselProject
from app.schemas.spare import SpareCreate, SpareOut, SpareUpdate
from app.services.feedback_learning import schedule_feedback_learning
from app.services.review_workflow import (
    broadcast_activity,
    log_activity,
    sync_spares_to_global_library,
)

router = APIRouter()

CJK_PATTERN = re.compile(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
PARENTHETICAL_CJK_PATTERN = re.compile(r"[（(][^）)]*[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff][^）)]*[）)]")
SPARE_TRANSLATIONS = {
    "ビニル手袋": "Vinyl Gloves",
    "収納箱": "Storage Box",
    "塩素系水処理剤": "Chlorine Water Treatment Chemical",
    "異物取出し用ハサミ": "Foreign Matter Removal Scissors",
    "ハサミ": "Scissors",
    "手袋": "Gloves",
    "収納": "Storage",
    "箱": "Box",
}


def _format_spare_source_reference(
    manual_name: Optional[str],
    page_reference: Optional[int],
) -> Optional[str]:
    if manual_name and page_reference:
        return f"{manual_name} (p.{page_reference})"
    return manual_name or (f"p.{page_reference}" if page_reference else None)


def _fallback_spare_part_number(
    part_number: Optional[str],
    manual_name: Optional[str],
    page_reference: Optional[int],
) -> Optional[str]:
    if part_number:
        return part_number
    return _format_spare_source_reference(manual_name, page_reference)


def _contains_cjk(value: Optional[str]) -> bool:
    return bool(value and CJK_PATTERN.search(value))


def _normalize_spare_text(
    value: Optional[str],
    *,
    fallback: Optional[str] = None,
    compact: bool = False,
) -> Optional[str]:
    text = str(value or "").strip()
    if not text:
        return fallback
    translated = SPARE_TRANSLATIONS.get(text)
    if translated:
        return translated

    cleaned = PARENTHETICAL_CJK_PATTERN.sub("", text)
    for source, target in SPARE_TRANSLATIONS.items():
        cleaned = cleaned.replace(source, f" {target} ")
    cleaned = cleaned.replace("（", "(").replace("）", ")")
    cleaned = CJK_PATTERN.sub(" ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -/,;:")

    if compact:
        cleaned = re.sub(r"\s{2,}", " ", cleaned).strip()

    if cleaned and any(ch.isalpha() or ch.isdigit() for ch in cleaned):
        return cleaned
    return fallback


def _spare_out_payload(spare: Spare) -> dict[str, Any]:
    return {
        "id": spare.id,
        "vessel_id": spare.vessel_id,
        "component_id": spare.component_id,
        "part_name": spare.part_name,
        "part_number": spare.part_number,
        "drawing_number": spare.drawing_number,
        "drawing_position": spare.drawing_position,
        "specification": spare.specification,
        "spare_assembly": spare.spare_assembly or spare.spare_model,
        "assembly_description": spare.assembly_description or spare.spare_assembly or spare.spare_model,
        "spare_maker": spare.spare_maker,
        "spare_model": spare.spare_model,
        "machinery_maker": spare.machinery_maker,
        "machinery_model": spare.machinery_model,
        "source_manual_id": spare.source_manual_id,
        "pdf_reference": None,
        "source_reference": None,
        "page_reference": spare.page_reference,
        "extraction_method": spare.extraction_method,
        "is_critical": spare.is_critical,
        "qc_status": spare.qc_status,
        "confidence_score": spare.confidence_score,
        "is_duplicate": spare.is_duplicate,
        "merged_into_id": spare.merged_into_id,
        "created_at": spare.created_at,
        "updated_at": spare.updated_at,
    }


def _spare_out(spare: Spare) -> SpareOut:
    return SpareOut.model_validate(_spare_out_payload(spare))


async def _run_spare_side_effects(
    db: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    vessel_id: uuid.UUID,
    user_id: uuid.UUID,
    activity_payloads: list[dict[str, Any]] | None = None,
    sync_spare_ids: list[uuid.UUID] | None = None,
) -> None:
    activities = []
    if activity_payloads:
        try:
            for payload in activity_payloads:
                activities.append(
                    await log_activity(
                        db,
                        tenant_id=tenant_id,
                        vessel_id=vessel_id,
                        user_id=user_id,
                        action_type=payload["action_type"],
                        entity_type="spare",
                        entity_id=payload["entity_id"],
                        description=payload["description"],
                        metadata=payload.get("metadata"),
                    )
                )
            await db.commit()
        except Exception:
            await db.rollback()
            activities = []

    for activity in activities:
        try:
            await broadcast_activity(activity)
        except Exception:
            continue

    if sync_spare_ids:
        try:
            sync_result = await db.execute(
                select(Spare).where(
                    Spare.id.in_(sync_spare_ids),
                    Spare.vessel_id == vessel_id,
                    Spare.is_deleted == False,
                )
            )
            sync_spares = [
                spare
                for spare in sync_result.scalars().all()
                if spare.qc_status == QCStatus.accepted
            ]
            await sync_spares_to_global_library(
                db,
                tenant_id=tenant_id,
                vessel_id=vessel_id,
                spares=sync_spares,
            )
            await db.commit()
        except Exception:
            await db.rollback()


async def _get_vessel_or_404(vessel_id: uuid.UUID, db: AsyncSession) -> VesselProject:
    result = await db.execute(
        select(VesselProject).where(
            VesselProject.id == vessel_id, VesselProject.is_deleted == False
        )
    )
    vessel = result.scalar_one_or_none()
    if vessel is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vessel not found")
    return vessel


async def _rerun_manual_extraction(manual_ids: list[str]) -> None:
    from app.services.extractor import auto_extract_from_manual

    for manual_id in manual_ids:
        try:
            await auto_extract_from_manual(manual_id)
        except Exception:
            continue


@router.get("/{vessel_id}/spares/source-files", summary="List unique source file names for spares")
async def list_spare_source_files(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    from app.models.ingestion import Manual
    await _get_vessel_or_404(vessel_id, db)
    result = await db.execute(
        select(Manual.original_filename)
        .join(Spare, Spare.source_manual_id == Manual.id)
        .where(
            Spare.vessel_id == vessel_id,
            Spare.tenant_id == current_user.tenant_id,
            Spare.is_deleted == False,
            Manual.original_filename.isnot(None),
            Manual.is_deleted == False,
        )
        .distinct()
        .order_by(Manual.original_filename.asc())
    )
    filenames = [row[0] for row in result.all() if row[0]]
    return {"items": filenames}


@router.get("/{vessel_id}/spares", summary="List spares with filters")
async def list_spares(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    component_id: Optional[uuid.UUID] = Query(None),
    extraction_method: Optional[str] = Query(None),
    qc_status: Optional[str] = Query(None),
    is_critical: Optional[bool] = Query(None),
    pdf_reference: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: str = Query("part_name"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
) -> dict[str, Any]:
    from sqlalchemy import func as _func, or_
    from app.models.component import Component
    from app.models.ingestion import Manual
    await _get_vessel_or_404(vessel_id, db)
    base_where = [
        Spare.vessel_id == vessel_id,
        Spare.tenant_id == current_user.tenant_id,
        Spare.is_deleted == False,
    ]
    if component_id:
        base_where.append(Spare.component_id == component_id)
    if extraction_method:
        try:
            base_where.append(Spare.extraction_method == ExtractionMethod(extraction_method))
        except ValueError:
            pass
    if qc_status:
        try:
            base_where.append(Spare.qc_status == QCStatus(qc_status))
        except ValueError:
            pass
    if is_critical is not None:
        base_where.append(Spare.is_critical == is_critical)
    if pdf_reference:
        manual_id_result = await db.execute(
            select(Manual.id).where(
                Manual.vessel_id == vessel_id,
                Manual.original_filename == pdf_reference,
                Manual.is_deleted == False,
            )
        )
        matched_ids = [row[0] for row in manual_id_result.all()]
        base_where.append(Spare.source_manual_id.in_(matched_ids) if matched_ids else (Spare.id == None))
    if search:
        base_where.append(
            or_(
                Spare.part_name.ilike(f"%{search}%"),
                Spare.part_number.ilike(f"%{search}%"),
                Spare.spare_maker.ilike(f"%{search}%"),
            )
        )

    total_result = await db.execute(select(_func.count()).select_from(Spare).where(*base_where))
    total: int = total_result.scalar_one()

    from sqlalchemy import cast, Integer, case, func

    sort_columns = {
        "part_name": Spare.part_name,
        "part_number": Spare.part_number,
        "drawing_number": Spare.drawing_number,
        "drawing_position": Spare.drawing_position,
        "spare_maker": Spare.spare_maker,
        "component": Spare.component_id,
        "extraction_method": Spare.extraction_method,
        "criticality": Spare.is_critical,
        "qc_status": Spare.qc_status,
        "page_reference": Spare.page_reference,
        "created_at": Spare.created_at,
    }

    if sort_by == "page_order":
        # Extract leading digits so ranges like '1~58' sort on their start value
        leading_digits = func.nullif(func.substring(Spare.drawing_position, r'^\d+'), '')
        numeric_pos = case(
            (leading_digits.isnot(None), cast(leading_digits, Integer)),
            else_=99999,
        )
        query = (
            select(Spare)
            .where(*base_where)
            .order_by(
                Spare.page_reference.asc().nulls_last(),
                numeric_pos.asc(),
                Spare.id.asc(),
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    else:
        order_col = sort_columns.get(sort_by, Spare.part_name)
        order_expr = order_col.desc() if sort_order == "desc" else order_col.asc()
        query = (
            select(Spare)
            .where(*base_where)
            .order_by(order_expr, Spare.page_reference.asc().nulls_last(), Spare.id.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    result = await db.execute(query)
    spares = result.scalars().all()

    component_ids = {spare.component_id for spare in spares if spare.component_id}
    manual_ids = {spare.source_manual_id for spare in spares if spare.source_manual_id}

    component_lookup: dict[uuid.UUID, Component] = {}
    manual_lookup: dict[uuid.UUID, Manual] = {}

    if component_ids:
        component_result = await db.execute(select(Component).where(Component.id.in_(component_ids)))
        component_lookup = {component.id: component for component in component_result.scalars().all()}
    if manual_ids:
        manual_result = await db.execute(
            select(Manual).where(Manual.id.in_(manual_ids), Manual.is_deleted == False)
        )
        manual_lookup = {manual.id: manual for manual in manual_result.scalars().all()}

    items = []
    for spare in spares:
        payload = _spare_out(spare).model_dump()
        component = component_lookup.get(spare.component_id) if spare.component_id else None
        manual = manual_lookup.get(spare.source_manual_id) if spare.source_manual_id else None
        manual_name = manual.original_filename if manual else None
        source_reference = _format_spare_source_reference(manual_name, spare.page_reference)
        component_name = component.component_name if component else None
        payload.update(
            {
                "component_name": component_name,
                "component_maker": component.maker if component else None,
                "component_model": component.model if component else None,
                "source_manual_name": manual_name,
                "pdf_reference": manual_name,
                "source_reference": source_reference,
                "part_number": _fallback_spare_part_number(
                    payload.get("part_number"),
                    manual_name,
                    spare.page_reference,
                ),
            }
        )
        english_part_number = _normalize_spare_text(
            payload.get("part_number"),
            fallback=_fallback_spare_part_number(
                None,
                manual_name,
                spare.page_reference,
            ),
            compact=True,
        )
        payload["part_number"] = english_part_number
        payload["part_name"] = _normalize_spare_text(
            payload.get("part_name"),
            fallback=(f"Spare Item {english_part_number}" if english_part_number else f"{component_name or 'Spare Item'}"),
        )
        payload["specification"] = _normalize_spare_text(payload.get("specification"))
        payload["drawing_number"] = _normalize_spare_text(payload.get("drawing_number"), compact=True)
        payload["drawing_position"] = _normalize_spare_text(payload.get("drawing_position"), compact=True)
        payload["spare_assembly"] = _normalize_spare_text(payload.get("spare_assembly"))
        payload["assembly_description"] = _normalize_spare_text(payload.get("assembly_description"))
        payload["spare_maker"] = _normalize_spare_text(payload.get("spare_maker"), compact=True)
        payload["spare_model"] = _normalize_spare_text(payload.get("spare_model"), compact=True)
        items.append(payload)

    import math
    total_pages = max(1, math.ceil(total / page_size))
    return {"items": items, "page": page, "page_size": page_size, "total": total, "total_pages": total_pages}


@router.get("/{vessel_id}/spares/export", summary="Export spares review data")
async def export_spares(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    component_id: Optional[uuid.UUID] = Query(None),
    extraction_method: Optional[str] = Query(None),
    qc_status: Optional[str] = Query(None),
    is_critical: Optional[bool] = Query(None),
    pdf_reference: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: str = Query("part_name"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$"),
) -> Response:
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Alignment, Font, PatternFill
    from sqlalchemy import Integer, case, cast, func, or_
    from app.models.component import Component
    from app.models.ingestion import Manual

    vessel = await _get_vessel_or_404(vessel_id, db)

    base_where = [
        Spare.vessel_id == vessel_id,
        Spare.tenant_id == current_user.tenant_id,
        Spare.is_deleted == False,
    ]
    if component_id:
        base_where.append(Spare.component_id == component_id)
    if extraction_method:
        try:
            base_where.append(Spare.extraction_method == ExtractionMethod(extraction_method))
        except ValueError:
            pass
    if qc_status:
        try:
            base_where.append(Spare.qc_status == QCStatus(qc_status))
        except ValueError:
            pass
    if is_critical is not None:
        base_where.append(Spare.is_critical == is_critical)
    if pdf_reference:
        manual_id_result = await db.execute(
            select(Manual.id).where(
                Manual.vessel_id == vessel_id,
                Manual.original_filename == pdf_reference,
                Manual.is_deleted == False,
            )
        )
        matched_ids = [row[0] for row in manual_id_result.all()]
        base_where.append(Spare.source_manual_id.in_(matched_ids) if matched_ids else (Spare.id == None))
    if search:
        base_where.append(
            or_(
                Spare.part_name.ilike(f"%{search}%"),
                Spare.part_number.ilike(f"%{search}%"),
                Spare.spare_maker.ilike(f"%{search}%"),
            )
        )

    sort_columns = {
        "part_name": Spare.part_name,
        "part_number": Spare.part_number,
        "drawing_number": Spare.drawing_number,
        "drawing_position": Spare.drawing_position,
        "spare_maker": Spare.spare_maker,
        "component": Spare.component_id,
        "extraction_method": Spare.extraction_method,
        "criticality": Spare.is_critical,
        "qc_status": Spare.qc_status,
        "page_reference": Spare.page_reference,
        "created_at": Spare.created_at,
    }
    if sort_by == "page_order":
        leading_digits = func.nullif(func.substring(Spare.drawing_position, r'^\d+'), '')
        numeric_pos = case(
            (leading_digits.isnot(None), cast(leading_digits, Integer)),
            else_=99999,
        )
        query = (
            select(Spare)
            .where(*base_where)
            .order_by(
                Spare.page_reference.asc().nulls_last(),
                numeric_pos.asc(),
                Spare.id.asc(),
            )
        )
    else:
        order_col = sort_columns.get(sort_by, Spare.part_name)
        order_expr = order_col.desc() if sort_order == "desc" else order_col.asc()
        query = (
            select(Spare)
            .where(*base_where)
            .order_by(order_expr, Spare.page_reference.asc().nulls_last(), Spare.id.asc())
        )
    result = await db.execute(query)
    spares = result.scalars().all()

    component_ids = {spare.component_id for spare in spares if spare.component_id}
    manual_ids = {spare.source_manual_id for spare in spares if spare.source_manual_id}

    component_lookup: dict[uuid.UUID, Component] = {}
    manual_lookup: dict[uuid.UUID, Manual] = {}

    if component_ids:
        component_result = await db.execute(select(Component).where(Component.id.in_(component_ids)))
        component_lookup = {component.id: component for component in component_result.scalars().all()}
    if manual_ids:
        manual_result = await db.execute(
            select(Manual).where(Manual.id.in_(manual_ids), Manual.is_deleted == False)
        )
        manual_lookup = {manual.id: manual for manual in manual_result.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spares"

    headers = [
        "ID", "PDF File", "Page", "Component", "Part Name", "Part Number",
        "Drawing #", "POS", "Specification", "Maker", "Spare Model",
        "Current QC", "Reviewer QC", "Reviewer Notes",
    ]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, spare in enumerate(spares, start=2):
        component = component_lookup.get(spare.component_id) if spare.component_id else None
        manual = manual_lookup.get(spare.source_manual_id) if spare.source_manual_id else None
        ws.cell(row=row_idx, column=1, value=str(spare.id))
        ws.cell(row=row_idx, column=2, value=manual.original_filename if manual else "")
        ws.cell(row=row_idx, column=3, value=spare.page_reference or "")
        ws.cell(row=row_idx, column=4, value=component.component_name if component else "")
        ws.cell(row=row_idx, column=5, value=spare.part_name or "")
        ws.cell(row=row_idx, column=6, value=spare.part_number or "")
        ws.cell(row=row_idx, column=7, value=spare.drawing_number or "")
        ws.cell(row=row_idx, column=8, value=spare.drawing_position or "")
        ws.cell(row=row_idx, column=9, value=spare.specification or "")
        ws.cell(row=row_idx, column=10, value=spare.spare_maker or "")
        ws.cell(row=row_idx, column=11, value=spare.spare_model or "")
        ws.cell(row=row_idx, column=12, value=spare.qc_status.value if hasattr(spare.qc_status, "value") else str(spare.qc_status))
        ws.cell(row=row_idx, column=13, value="")
        ws.cell(row=row_idx, column=14, value="")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 4, 12), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    vessel_name = getattr(vessel, "vessel_name", None) or getattr(vessel, "name", None) or str(vessel_id)
    safe_name = "".join(ch if ch.isalnum() or ch in "-_ " else "_" for ch in vessel_name).strip() or str(vessel_id)
    filename = f"Spares_Review_{safe_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/{vessel_id}/spares",
    response_model=SpareOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_spare(
    vessel_id: uuid.UUID,
    body: SpareCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> SpareOut:
    await _get_vessel_or_404(vessel_id, db)
    spare = Spare(
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        **body.model_dump(),
    )
    if spare.spare_assembly and not spare.assembly_description:
        spare.assembly_description = spare.spare_assembly
    db.add(spare)
    await db.commit()
    await db.refresh(spare)
    await _run_spare_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "spare.created",
                "entity_id": spare.id,
                "description": f"Created spare '{spare.part_name}'.",
                "metadata": {"component_id": str(spare.component_id) if spare.component_id else None},
            }
        ],
        sync_spare_ids=[spare.id] if spare.qc_status == QCStatus.accepted else None,
    )
    await db.refresh(spare)
    return _spare_out(spare)


@router.patch("/{vessel_id}/spares/{spare_id}", response_model=SpareOut)
async def update_spare(
    vessel_id: uuid.UUID,
    spare_id: uuid.UUID,
    body: SpareUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> SpareOut:
    await _get_vessel_or_404(vessel_id, db)
    result = await db.execute(
        select(Spare).where(
            Spare.id == spare_id, Spare.vessel_id == vessel_id, Spare.is_deleted == False
        )
    )
    spare = result.scalar_one_or_none()
    if spare is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Spare not found")

    original = {"part_name": spare.part_name, "qc_status": spare.qc_status.value}
    original_qc_status = spare.qc_status
    update_data = body.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(spare, field, value)
    if "spare_assembly" in update_data and "assembly_description" not in update_data:
        spare.assembly_description = spare.spare_assembly
    db.add(spare)

    feedback_id: uuid.UUID | None = None
    if spare.source_manual_id and update_data:
        feedback = FeedbackEntry(
            tenant_id=current_user.tenant_id,
            manual_id=spare.source_manual_id,
            entity_type="spare",
            original_value=original,
            corrected_value=body.model_dump(mode="json", exclude_unset=True),
            correction_type=CorrectionType.wrong_value,
            page_number=spare.page_reference,
            context_span=f"Updated spare fields: {', '.join(sorted(update_data.keys()))}",
            created_by=current_user.id,
        )
        db.add(feedback)
        await db.flush()
        feedback_id = feedback.id

    await db.commit()
    if feedback_id:
        await schedule_feedback_learning(feedback_id)
    await db.refresh(spare)
    await _run_spare_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "spare.corrected" if spare.source_manual_id else "spare.modified",
                "entity_id": spare.id,
                "description": f"Updated spare '{spare.part_name}'.",
                "metadata": {"fields": sorted(update_data.keys())},
            }
        ] if update_data else None,
        sync_spare_ids=[spare.id]
        if bool(update_data)
        and (original_qc_status == QCStatus.accepted or spare.qc_status == QCStatus.accepted)
        else None,
    )
    await db.refresh(spare)
    return _spare_out(spare)


@router.delete("/{vessel_id}/spares/{spare_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response, response_model=None)
async def delete_spare(
    vessel_id: uuid.UUID,
    spare_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    result = await db.execute(
        select(Spare).where(Spare.id == spare_id, Spare.vessel_id == vessel_id)
    )
    spare = result.scalar_one_or_none()
    if spare:
        needs_sync = spare.qc_status == QCStatus.accepted
        spare.is_deleted = True
        db.add(spare)
        await db.commit()
        if needs_sync:
            await sync_spares_to_global_library(
                db,
                tenant_id=current_user.tenant_id,
                vessel_id=vessel_id,
                spares=[],
            )
            await db.commit()


@router.post("/{vessel_id}/spares/bulk-accept")
async def bulk_accept_spares(
    vessel_id: uuid.UUID,
    body: dict[str, List[str]],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    result = await db.execute(select(Spare).where(Spare.id.in_(ids), Spare.vessel_id == vessel_id))
    spares = result.scalars().all()
    for spare in spares:
        spare.qc_status = QCStatus.accepted
        db.add(spare)
    await db.commit()
    await _run_spare_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "spare.accepted",
                "entity_id": spare.id,
                "description": f"Accepted spare '{spare.part_name}'.",
            }
            for spare in spares
        ],
        sync_spare_ids=[spare.id for spare in spares],
    )
    return {"accepted": len(spares)}


@router.post("/{vessel_id}/spares/bulk-reject")
async def bulk_reject_spares(
    vessel_id: uuid.UUID,
    body: dict[str, List[str]],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    result = await db.execute(select(Spare).where(Spare.id.in_(ids), Spare.vessel_id == vessel_id))
    spares = result.scalars().all()
    should_sync = any(spare.qc_status == QCStatus.accepted for spare in spares)
    for spare in spares:
        spare.qc_status = QCStatus.rejected
        db.add(spare)
    await db.commit()
    if should_sync:
        await sync_spares_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            spares=[],
        )
        await db.commit()
    await _run_spare_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "spare.rejected",
                "entity_id": spare.id,
                "description": f"Rejected spare '{spare.part_name}'.",
            }
            for spare in spares
        ],
    )
    return {"rejected": len(spares)}


@router.post("/{vessel_id}/spares/bulk-delete")
async def bulk_delete_spares(
    vessel_id: uuid.UUID,
    body: dict[str, List[str]],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(i) for i in body.get("ids", [])]
    result = await db.execute(select(Spare).where(Spare.id.in_(ids), Spare.vessel_id == vessel_id, Spare.is_deleted == False))
    spares = result.scalars().all()
    should_sync = any(spare.qc_status == QCStatus.accepted for spare in spares)
    for spare in spares:
        spare.is_deleted = True
        db.add(spare)
    await db.commit()
    if should_sync:
        await sync_spares_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            spares=[],
        )
        await db.commit()
    await _run_spare_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "spare.deleted",
                "entity_id": spare.id,
                "description": f"Deleted spare '{spare.part_name}'.",
            }
            for spare in spares
        ],
    )
    return {"deleted": len(spares)}


@router.post("/{vessel_id}/spares/bulk-update")
async def bulk_update_spares(
    vessel_id: uuid.UUID,
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    ids = [uuid.UUID(value) for value in body.get("ids", []) if value]
    updates = body.get("updates", {}) or {}
    if not ids or not updates:
        return {"updated": 0}

    update_payload = SpareUpdate.model_validate(updates).model_dump(exclude_unset=True)
    if not update_payload:
        return {"updated": 0}

    result = await db.execute(
        select(Spare).where(Spare.id.in_(ids), Spare.vessel_id == vessel_id, Spare.is_deleted == False)
    )
    spares = result.scalars().all()
    for spare in spares:
        setattr(spare, "_original_qc_status_for_sync", spare.qc_status)
        for field, value in update_payload.items():
            setattr(spare, field, value)
        db.add(spare)
    await db.commit()
    await _run_spare_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "spare.bulk_updated",
                "entity_id": spare.id,
                "description": f"Bulk updated spare '{spare.part_name}'.",
                "metadata": {"fields": sorted(update_payload.keys())},
            }
            for spare in spares
        ],
        sync_spare_ids=[
            spare.id
            for spare in spares
            if getattr(spare, "_original_qc_status_for_sync", None) == QCStatus.accepted
            or spare.qc_status == QCStatus.accepted
        ],
    )
    return {"updated": len(spares)}


@router.post("/{vessel_id}/spares/{spare_id}/merge")
async def merge_spare(
    vessel_id: uuid.UUID,
    spare_id: uuid.UUID,
    body: dict[str, str],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> SpareOut:
    result = await db.execute(
        select(Spare).where(
            Spare.id == spare_id, Spare.vessel_id == vessel_id, Spare.is_deleted == False
        )
    )
    spare = result.scalar_one_or_none()
    if spare is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Spare not found")

    target_id = uuid.UUID(body["target_spare_id"])
    needs_sync = spare.qc_status == QCStatus.accepted
    spare.merged_into_id = target_id
    spare.is_duplicate = True
    spare.is_deleted = True
    db.add(spare)
    await db.commit()
    if needs_sync:
        await sync_spares_to_global_library(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            spares=[],
        )
        await db.commit()
    await db.refresh(spare)
    return _spare_out(spare)


@router.post("/{vessel_id}/spares/{spare_id}/reassign-component")
async def reassign_component(
    vessel_id: uuid.UUID,
    spare_id: uuid.UUID,
    body: dict[str, str],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> SpareOut:
    result = await db.execute(
        select(Spare).where(
            Spare.id == spare_id, Spare.vessel_id == vessel_id, Spare.is_deleted == False
        )
    )
    spare = result.scalar_one_or_none()
    if spare is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Spare not found")

    spare.component_id = uuid.UUID(body["component_id"])
    db.add(spare)
    await db.commit()
    await db.refresh(spare)
    await _run_spare_side_effects(
        db,
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        user_id=current_user.id,
        activity_payloads=[
            {
                "action_type": "spare.mapped",
                "entity_id": spare.id,
                "description": f"Linked spare '{spare.part_name}' to a vessel component.",
                "metadata": {"component_id": body["component_id"]},
            }
        ],
    )
    await db.refresh(spare)
    return _spare_out(spare)


@router.post("/{vessel_id}/spares/trigger-extraction")
async def trigger_spare_extraction(
    vessel_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)
    from app.models.ingestion import Manual

    result = await db.execute(
        select(Manual).where(
            Manual.vessel_id == vessel_id,
            Manual.tenant_id == current_user.tenant_id,
            Manual.is_deleted == False,
            Manual.pages_with_spares.isnot(None),
            Manual.pages_with_spares != "",
        )
    )
    manuals = result.scalars().all()
    manual_ids = [str(manual.id) for manual in manuals]
    if not manual_ids:
        return {"started": False, "total": 0, "message": "No manuals with spare page references found."}

    background_tasks.add_task(_rerun_manual_extraction, manual_ids)
    return {"started": True, "total": len(manual_ids), "message": "Spare extraction started."}


@router.get("/{vessel_id}/spares/{spare_id}/page-image")
async def get_spare_page_image(
    vessel_id: uuid.UUID,
    spare_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    result = await db.execute(
        select(Spare).where(
            Spare.id == spare_id, Spare.vessel_id == vessel_id, Spare.is_deleted == False
        )
    )
    spare = result.scalar_one_or_none()
    if spare is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Spare not found")

    if spare.source_manual_id and spare.page_reference:
        blob_key = f"page-images/{spare.source_manual_id}/page_{spare.page_reference}.png"
        try:
            from app.services.blob_storage import BlobStorageService

            blob_svc = BlobStorageService()
            url = blob_svc.get_presigned_url(blob_key)
            return {"image_url": url}
        except Exception:
            pass

    return {"image_url": None, "message": "Page image not available"}


@router.post("/{vessel_id}/spares/snip-extract", summary="Extract spare parts from an uploaded image (snip tool)")
async def snip_extract_spares(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    image: UploadFile = File(...),
    page_number: Optional[int] = Form(None),
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)
    image_bytes = await image.read()
    if not image_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Image file is empty")

    from app.services.extractor import _extract_entities_from_page_image

    # Claude vision primary, OpenAI fallback. Enhancement applied inside the vision function.
    records = await _extract_entities_from_page_image(
        image_bytes=image_bytes,
        filename=image.filename or "snipped_region.png",
        page_no=page_number or 0,
        extraction_type="spare",
        context_note=(
            "This is a manually snipped/cropped image of a spare parts table. "
            "STEP 1: Count the number of data rows visible (not headers). "
            "STEP 2: Output EXACTLY that many JSON records — do not stop early or skip rows. "
            "Typical columns: REF.NO | CODE NO | PC.NO | DESCRIPTION | QTY | REMARKS. "
            "ALL output fields MUST be in English — translate ALL Japanese or non-English text to English. "
            "Never output Japanese/non-English characters in any field."
        ),
    )
    return {"records": records, "count": len(records)}


@router.post("/{vessel_id}/spares/snip-save", summary="Save spare records extracted via the snip tool")
async def snip_save_spares(
    vessel_id: uuid.UUID,
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)
    records: list[dict[str, Any]] = body.get("records", [])
    source_manual_id_str: Optional[str] = body.get("source_manual_id")
    page_number: Optional[int] = body.get("page_number")

    source_manual_id: Optional[uuid.UUID] = None
    if source_manual_id_str:
        try:
            source_manual_id = uuid.UUID(source_manual_id_str)
        except ValueError:
            pass

    saved_spares: list[Spare] = []
    for record in records:
        part_name = str(record.get("part_name") or "").strip()
        if not part_name:
            continue
        spare = Spare(
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            part_name=part_name,
            part_number=record.get("part_number") or None,
            drawing_number=record.get("drawing_number") or None,
            drawing_position=record.get("drawing_position") or None,
            specification=record.get("specification") or None,
            spare_assembly=record.get("spare_model") or None,
            assembly_description=record.get("spare_model") or None,
            spare_maker=record.get("spare_maker") or None,
            spare_model=record.get("spare_model") or None,
            source_manual_id=source_manual_id,
            page_reference=page_number,
            extraction_method=ExtractionMethod.manual,
            is_critical=False,
            qc_status=QCStatus.pending,
            confidence_score=int(record.get("confidence_score") or 75),
        )
        db.add(spare)
        saved_spares.append(spare)

    if saved_spares:
        await db.commit()
        await _run_spare_side_effects(
            db,
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            user_id=current_user.id,
            activity_payloads=[
                {
                    "action_type": "spare.created",
                    "entity_id": spare.id,
                    "description": f"Added spare '{spare.part_name}' via snip extraction.",
                }
                for spare in saved_spares
            ],
        )

    return {"saved": len(saved_spares)}


# ---------------------------------------------------------------------------
# QC Review Export / Import
# ---------------------------------------------------------------------------

_QC_VALUES = ["accepted", "rejected", "modified", "pending"]

_COMP_HEADERS = [
    "ID", "PDF File", "Page", "Component Name", "Main Machinery",
    "Group 1", "Group 2", "Maker", "Model", "Location",
    "Current QC", "Reviewer QC", "Reviewer Notes",
]
_JOB_HEADERS = [
    "ID", "PDF File", "Page", "Component", "Job Name", "Job Code",
    "Frequency", "Frequency Type", "Performing Rank", "Description",
    "Current QC", "Reviewer QC", "Reviewer Notes",
]
_SPARE_HEADERS = [
    "ID", "PDF File", "Page", "Component", "Part Name", "Part Number",
    "Drawing #", "POS", "Specification", "Maker", "Spare Model",
    "Current QC", "Reviewer QC", "Reviewer Notes",
]

_QC_COLORS = {
    "accepted": "C6EFCE",
    "rejected": "FFC7CE",
    "modified": "FFEB9C",
    "pending": "DDEBF7",
}


def _qc_review_workbook(
    vessel_name: str,
    components: "list | None",
    jobs: "list | None",
    spares: "list | None",
    component_lookup: dict,
    manual_lookup: dict,
) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    EDIT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ALT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    qc_dv = DataValidation(
        type="list",
        formula1='"accepted,rejected,modified,pending"',
        showDropDown=False,
        allow_blank=True,
    )

    wb = openpyxl.Workbook()

    # ── Instructions sheet ──────────────────────────────────────────────────
    ins = wb.active
    ins.title = "Instructions"
    ins["A1"] = f"QC Review Export — {vessel_name}"
    ins["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ins["A3"] = "HOW TO USE THIS FILE"
    ins["A3"].font = Font(bold=True, size=11)
    for i, line in enumerate([
        "1. Open the Components, Jobs, or Spares sheet.",
        "2. Locate the row using the PDF File and Page columns to find the item in the original manual.",
        "3. In the 'Reviewer QC' column, select: accepted / rejected / modified  (leave blank to skip).",
        "4. Optionally add a comment in the 'Reviewer Notes' column.",
        "5. Save the file and upload it back via the 'Import QC Feedback' button on the Export page.",
        "",
        "IMPORTANT: Do not change the ID column or add/remove rows.",
        "Only the 'Reviewer QC' and 'Reviewer Notes' columns are read during import.",
    ], start=5):
        ins[f"A{i}"] = line
    ins.column_dimensions["A"].width = 90

    def _write_sheet(ws, headers, rows, reviewer_col_idx):
        ws.add_data_validation(qc_dv)
        # Header row
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        rev_col_letter = get_column_letter(reviewer_col_idx)
        note_col_letter = get_column_letter(reviewer_col_idx + 1)

        for r, row_data in enumerate(rows, 2):
            qc_val = str(row_data[10] or "pending").lower()  # Current QC always at index 10
            fill = ALT_FILL if r % 2 == 0 else PatternFill(fill_type=None)
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=(c == len(headers)))
                # Current QC colour
                if c == reviewer_col_idx - 1:
                    color = _QC_COLORS.get(qc_val, "DDEBF7")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                elif c == reviewer_col_idx:
                    cell.fill = EDIT_FILL
                elif c == reviewer_col_idx + 1:
                    cell.fill = EDIT_FILL
                else:
                    cell.fill = fill
            # Attach dropdown validation to reviewer QC cell
            qc_dv.add(ws[f"{rev_col_letter}{r}"])

        # Auto-width (capped)
        for col in ws.columns:
            vals = [str(cell.value or "") for cell in col]
            width = min(max(len(v) for v in vals) + 4, 45)
            ws.column_dimensions[col[0].column_letter].width = max(width, 10)
        # Notes column wider
        ws.column_dimensions[note_col_letter].width = 30

    # ── Components sheet ────────────────────────────────────────────────────
    if components is not None:
        ws_comp = wb.create_sheet("Components")
        comp_rows = []
        for comp in components:
            manual_name = manual_lookup.get(comp.source_manual_id, "")
            comp_rows.append([
                str(comp.id),
                manual_name,
                comp.page_reference or "",
                comp.component_name or "",
                comp.main_machinery or "",
                comp.group1 or "",
                comp.group2 or "",
                comp.maker or "",
                comp.model or "",
                comp.location or "",
                comp.qc_status.value if hasattr(comp.qc_status, "value") else str(comp.qc_status),
                "",  # Reviewer QC
                "",  # Reviewer Notes
            ])
        _write_sheet(ws_comp, _COMP_HEADERS, comp_rows, reviewer_col_idx=12)

    # ── Jobs sheet ──────────────────────────────────────────────────────────
    if jobs is not None:
        ws_jobs = wb.create_sheet("Jobs")
        job_rows = []
        for job in jobs:
            comp = component_lookup.get(job.component_id)
            comp_name = comp.component_name if comp else ""
            job_rows.append([
                str(job.id),
                job.pdf_reference or "",
                job.page_reference or "",
                comp_name,
                job.job_name or "",
                job.job_code or "",
                job.frequency or "",
                job.frequency_type.value if job.frequency_type else "",
                job.performing_rank or "",
                (job.job_description or "")[:300],
                job.qc_status.value if hasattr(job.qc_status, "value") else str(job.qc_status),
                "",  # Reviewer QC
                "",  # Reviewer Notes
            ])
        _write_sheet(ws_jobs, _JOB_HEADERS, job_rows, reviewer_col_idx=12)

    # ── Spares sheet ────────────────────────────────────────────────────────
    if spares is not None:
        ws_spares = wb.create_sheet("Spares")
        spare_rows = []
        for spare in spares:
            comp = component_lookup.get(spare.component_id)
            comp_name = comp.component_name if comp else ""
            manual_name = manual_lookup.get(spare.source_manual_id, "")
            spare_rows.append([
                str(spare.id),
                manual_name,
                spare.page_reference or "",
                comp_name,
                spare.part_name or "",
                spare.part_number or "",
                spare.drawing_number or "",
                spare.drawing_position or "",
                spare.specification or "",
                spare.spare_maker or "",
                spare.spare_model or "",
                spare.qc_status.value if hasattr(spare.qc_status, "value") else str(spare.qc_status),
                "",  # Reviewer QC
                "",  # Reviewer Notes
            ])
        _write_sheet(ws_spares, _SPARE_HEADERS, spare_rows, reviewer_col_idx=13)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/{vessel_id}/spares/qc-export", summary="Export Spares for QC review (spares only)")
async def export_spares_qc(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> Response:
    from fastapi.responses import StreamingResponse
    import io
    from app.models.component import Component
    from app.models.ingestion import Manual

    vessel = await _get_vessel_or_404(vessel_id, db)
    vessel_name = getattr(vessel, "vessel_name", None) or getattr(vessel, "name", None) or str(vessel_id)

    spares_result = await db.execute(
        select(Spare).where(
            Spare.vessel_id == vessel_id,
            Spare.tenant_id == current_user.tenant_id,
            Spare.is_deleted == False,
        ).order_by(Spare.source_manual_id, Spare.page_reference, Spare.drawing_position)
    )
    spares = list(spares_result.scalars().all())

    component_ids = {s.component_id for s in spares if s.component_id}
    manual_ids = {s.source_manual_id for s in spares if s.source_manual_id}

    component_lookup: dict = {}
    if component_ids:
        cr = await db.execute(select(Component).where(Component.id.in_(component_ids)))
        component_lookup = {c.id: c for c in cr.scalars().all()}

    manual_lookup: dict = {}
    if manual_ids:
        mr = await db.execute(select(Manual).where(Manual.id.in_(manual_ids), Manual.is_deleted == False))
        manual_lookup = {m.id: m.original_filename for m in mr.scalars().all()}

    xlsx_bytes = _qc_review_workbook(
        vessel_name=vessel_name,
        components=None,
        jobs=None,
        spares=spares,
        component_lookup=component_lookup,
        manual_lookup=manual_lookup,
    )

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in vessel_name).strip()
    filename = f"Spares_QC_{safe_name}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{vessel_id}/spares/qc-export-all", summary="Export Components, Jobs and Spares for QC review")
async def export_qc_review(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> Response:
    from fastapi.responses import StreamingResponse
    import io
    from app.models.component import Component
    from app.models.ingestion import Manual
    from app.models.job import Job

    vessel = await _get_vessel_or_404(vessel_id, db)
    vessel_name = getattr(vessel, "vessel_name", None) or getattr(vessel, "name", None) or str(vessel_id)

    comps_result = await db.execute(
        select(Component).where(
            Component.vessel_id == vessel_id,
            Component.tenant_id == current_user.tenant_id,
            Component.is_deleted == False,
        ).order_by(Component.main_machinery, Component.component_name)
    )
    components = list(comps_result.scalars().all())

    jobs_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == current_user.tenant_id,
            Job.is_deleted == False,
        ).order_by(Job.pdf_reference, Job.page_reference, Job.job_name)
    )
    jobs = list(jobs_result.scalars().all())

    spares_result = await db.execute(
        select(Spare).where(
            Spare.vessel_id == vessel_id,
            Spare.tenant_id == current_user.tenant_id,
            Spare.is_deleted == False,
        ).order_by(Spare.source_manual_id, Spare.page_reference, Spare.drawing_position)
    )
    spares = list(spares_result.scalars().all())

    # Build lookup maps
    all_component_ids = (
        {c.id for c in components}
        | {j.component_id for j in jobs if j.component_id}
        | {s.component_id for s in spares if s.component_id}
    )
    all_manual_ids = (
        {c.source_manual_id for c in components if c.source_manual_id}
        | {s.source_manual_id for s in spares if s.source_manual_id}
    )

    component_lookup: dict = {}
    if all_component_ids:
        cr = await db.execute(select(Component).where(Component.id.in_(all_component_ids)))
        component_lookup = {c.id: c for c in cr.scalars().all()}

    manual_lookup: dict = {}
    if all_manual_ids:
        mr = await db.execute(
            select(Manual).where(Manual.id.in_(all_manual_ids), Manual.is_deleted == False)
        )
        manual_lookup = {m.id: m.original_filename for m in mr.scalars().all()}

    xlsx_bytes = _qc_review_workbook(
        vessel_name=vessel_name,
        components=components,
        jobs=jobs,
        spares=spares,
        component_lookup=component_lookup,
        manual_lookup=manual_lookup,
    )

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in vessel_name).strip()
    filename = f"QC_Review_{safe_name}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{vessel_id}/spares/qc-import", summary="Import QC feedback from review Excel")
async def import_qc_review(
    vessel_id: uuid.UUID,
    file: UploadFile,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    from app.models.component import Component
    from app.models.job import Job

    await _get_vessel_or_404(vessel_id, db)

    content = await file.read()
    try:
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file.")

    valid_qc = set(_QC_VALUES)
    updated = {"components": 0, "jobs": 0, "spares": 0}

    def _parse_sheet(ws, id_col=0, qc_col=11, note_col=12):
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header
            if not row or not row[id_col]:
                continue
            raw_id = str(row[id_col]).strip()
            raw_qc = str(row[qc_col] or "").strip().lower()
            raw_note = str(row[note_col] or "").strip() if len(row) > note_col else ""
            if not raw_id:
                continue
            rows.append((raw_id, raw_qc, raw_note))
        return rows

    # ── Components sheet ──
    if "Components" in wb.sheetnames:
        rows = _parse_sheet(wb["Components"], id_col=0, qc_col=11, note_col=12)
        for raw_id, raw_qc, _ in rows:
            if raw_qc not in valid_qc:
                continue
            try:
                entity_id = uuid.UUID(raw_id)
            except ValueError:
                continue
            result = await db.execute(
                select(Component).where(
                    Component.id == entity_id,
                    Component.vessel_id == vessel_id,
                    Component.tenant_id == current_user.tenant_id,
                    Component.is_deleted == False,
                )
            )
            obj = result.scalar_one_or_none()
            if obj and obj.qc_status.value != raw_qc:
                obj.qc_status = QCStatus(raw_qc)
                db.add(obj)
                updated["components"] += 1

    # ── Jobs sheet ──
    if "Jobs" in wb.sheetnames:
        rows = _parse_sheet(wb["Jobs"], id_col=0, qc_col=11, note_col=12)
        for raw_id, raw_qc, _ in rows:
            if raw_qc not in valid_qc:
                continue
            try:
                entity_id = uuid.UUID(raw_id)
            except ValueError:
                continue
            result = await db.execute(
                select(Job).where(
                    Job.id == entity_id,
                    Job.vessel_id == vessel_id,
                    Job.tenant_id == current_user.tenant_id,
                    Job.is_deleted == False,
                )
            )
            obj = result.scalar_one_or_none()
            if obj and obj.qc_status.value != raw_qc:
                obj.qc_status = QCStatus(raw_qc)
                db.add(obj)
                updated["jobs"] += 1

    # ── Spares sheet ──
    if "Spares" in wb.sheetnames:
        rows = _parse_sheet(wb["Spares"], id_col=0, qc_col=12, note_col=13)
        for raw_id, raw_qc, _ in rows:
            if raw_qc not in valid_qc:
                continue
            try:
                entity_id = uuid.UUID(raw_id)
            except ValueError:
                continue
            result = await db.execute(
                select(Spare).where(
                    Spare.id == entity_id,
                    Spare.vessel_id == vessel_id,
                    Spare.tenant_id == current_user.tenant_id,
                    Spare.is_deleted == False,
                )
            )
            obj = result.scalar_one_or_none()
            if obj and obj.qc_status.value != raw_qc:
                obj.qc_status = QCStatus(raw_qc)
                db.add(obj)
                updated["spares"] += 1

    total = sum(updated.values())
    if total:
        await db.commit()

    return {
        "updated": total,
        "components": updated["components"],
        "jobs": updated["jobs"],
        "spares": updated["spares"],
    }
