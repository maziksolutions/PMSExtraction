from __future__ import annotations

import csv
import io
import re
import uuid
from difflib import SequenceMatcher
from typing import Annotated, Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import String, asc, desc, func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import engine, get_db
from app.deps import get_current_user
from app.models.component import Component, QCStatus
from app.models.job import Job
from app.models.missing_manual import MissingManualGap
from app.models.standard_jobs import (
    ClassSociety,
    MatchStatus,
    StandardJob,
    StandardJobMatch,
    VesselTypeTemplate,
)
from app.models.user import User
from app.models.vessel import VesselProject
from app.services.job_ranks import (
    derive_standard_job_ranks,
    ensure_job_rank,
    infer_rank_from_component,
    normalize_rank_name,
)
from app.services.upload_security import validate_uploaded_file_bytes

router = APIRouter()


async def _get_vessel_or_404(vessel_id: uuid.UUID, db: AsyncSession) -> VesselProject:
    result = await db.execute(
        select(VesselProject).where(
            VesselProject.id == vessel_id, VesselProject.is_deleted == False
        )
    )
    vessel = result.scalar_one_or_none()
    if vessel is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vessel not found")
    return vessel


def _norm_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _tokenize_text(value: Any) -> set[str]:
    return {
        token
        for token in re.split(r"[^a-z0-9]+", _norm_text(value))
        if token
    }


def _header_key(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _clean_text(value: Any) -> Optional[str]:
    text = str(value or "").strip()
    return text or None


def _first_text(row: dict[str, Any], *keys: str) -> Optional[str]:
    for key in keys:
        value = _clean_text(row.get(key))
        if value:
            return value
    return None


def _enum_value(value: Any) -> str | None:
    if value is None:
        return None
    return value.value if hasattr(value, "value") else str(value)


def _class_society_value(value: Any) -> str:
    raw = (_enum_value(value) or "General").strip()
    if not raw:
        return "General"
    normalized = _norm_text(raw)
    for member in ClassSociety:
        member_tokens = {
            _norm_text(member.value),
            _norm_text(member.name),
            _norm_text(f"{member.__class__.__name__}.{member.name}"),
        }
        if normalized in member_tokens:
            return member.value
    return raw


def _frequency_type_value(value: Any) -> str | None:
    raw = _enum_value(value)
    if not isinstance(raw, str) or not raw.strip():
        return None
    cleaned = raw.strip()
    normalized = _norm_text(cleaned).replace(" ", "_")
    mapped = FREQ_ALIASES.get(normalized.replace("_", " ")) or FREQ_ALIASES.get(normalized.replace("_", ""))
    if mapped:
        return mapped
    for member in FREQ_MAP.values():
        member_tokens = {
            _norm_text(member.value).replace(" ", "_"),
            _norm_text(member.name).replace(" ", "_"),
            _norm_text(f"{member.__class__.__name__}.{member.name}").replace(" ", "_"),
        }
        if normalized in member_tokens:
            return member.value
    return cleaned


def _job_type_for_standard_job(job: StandardJob) -> str:
    if bool(job.is_critical):
        return "critical"
    normalized_society = _norm_text(_class_society_value(job.class_society))
    return "class" if normalized_society in CLASS_LIBRARY_VALUES else "standard"


def _serialize_standard_job(job: StandardJob) -> dict[str, Any]:
    performing_rank, verifying_rank = derive_standard_job_ranks(
        job_name=job.job_name,
        machinery_type=job.machinery_type,
        library_reference=job.library_reference,
        performing_rank=job.performing_rank,
        verifying_rank=job.verifying_rank,
    )
    class_society = _class_society_value(job.class_society)
    frequency_type = _frequency_type_value(job.frequency_type)
    if job.is_critical:
        job_type = "critical"
    else:
        job_type = "class" if _norm_text(class_society) in CLASS_LIBRARY_VALUES else "standard"
    return {
        "id": str(job.id),
        "class_society": class_society,
        "job_type": job_type,
        "machinery_type": job.machinery_type,
        "job_name": job.job_name,
        "job_description": job.job_description,
        "performing_rank": performing_rank,
        "verifying_rank": verifying_rank,
        "frequency": job.frequency,
        "frequency_type": frequency_type,
        "is_critical": job.is_critical,
        "library_reference": job.library_reference,
    }


def _most_common_rank(values: list[Optional[str]]) -> Optional[str]:
    counts: dict[str, int] = {}
    for value in values:
        cleaned = normalize_rank_name(value)
        if not cleaned:
            continue
        counts[cleaned] = counts.get(cleaned, 0) + 1
    if not counts:
        return None
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]


async def _hydrate_standard_job_ranks(
    db: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    jobs: list[StandardJob],
) -> None:
    target_jobs = [job for job in jobs if not job.performing_rank or not job.verifying_rank]
    if not target_jobs:
        return

    result = await db.execute(
        select(
            StandardJobMatch.standard_job_id,
            Job.performing_rank,
            Job.verifying_rank,
        )
        .join(Job, Job.id == StandardJobMatch.matched_job_id)
        .where(
            StandardJobMatch.tenant_id == tenant_id,
            StandardJobMatch.is_deleted == False,
            StandardJobMatch.standard_job_id.in_([job.id for job in target_jobs]),
            Job.tenant_id == tenant_id,
            Job.is_deleted == False,
        )
    )

    performing_candidates: dict[uuid.UUID, list[Optional[str]]] = {}
    verifying_candidates: dict[uuid.UUID, list[Optional[str]]] = {}
    for standard_job_id, performing_rank, verifying_rank in result.all():
        performing_candidates.setdefault(standard_job_id, []).append(performing_rank)
        verifying_candidates.setdefault(standard_job_id, []).append(verifying_rank)

    job_names = list({job.job_name for job in target_jobs if job.job_name})
    if job_names:
        name_match_result = await db.execute(
            select(
                Job.job_name,
                Job.performing_rank,
                Job.verifying_rank,
            )
            .where(
                Job.tenant_id == tenant_id,
                Job.is_deleted == False,
                Job.job_name.in_(job_names),
            )
        )
        performing_by_name: dict[str, list[Optional[str]]] = {}
        verifying_by_name: dict[str, list[Optional[str]]] = {}
        for job_name, performing_rank, verifying_rank in name_match_result.all():
            performing_by_name.setdefault(job_name, []).append(performing_rank)
            verifying_by_name.setdefault(job_name, []).append(verifying_rank)
    else:
        performing_by_name = {}
        verifying_by_name = {}

    changed = False
    for job in target_jobs:
        job_changed = False
        next_performing_rank = job.performing_rank or _most_common_rank(
            performing_candidates.get(job.id, [])
        ) or _most_common_rank(
            performing_by_name.get(job.job_name, [])
        )
        next_verifying_rank = job.verifying_rank or _most_common_rank(
            verifying_candidates.get(job.id, [])
        ) or _most_common_rank(
            verifying_by_name.get(job.job_name, [])
        )

        if next_performing_rank and next_performing_rank != job.performing_rank:
            job.performing_rank = next_performing_rank
            changed = True
            job_changed = True
        if next_verifying_rank and next_verifying_rank != job.verifying_rank:
            job.verifying_rank = next_verifying_rank
            changed = True
            job_changed = True

        if job_changed:
            db.add(job)

    if not changed:
        return

    for job in target_jobs:
        await ensure_job_rank(db, tenant_id=tenant_id, rank_name=job.performing_rank)
        await ensure_job_rank(db, tenant_id=tenant_id, rank_name=job.verifying_rank)
    await db.commit()


def _coerce_bool(value: Any) -> bool:
    return _norm_text(value) in {"yes", "true", "1", "y"}


def _coerce_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text or _norm_text(text) in {"na", "n/a", "maker instruction", "makers instruction"}:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


FREQ_MAP = {member.value: member for member in __import__("app.models.job", fromlist=["FrequencyType"]).FrequencyType}
FREQ_ALIASES = {
    # daily
    "daily": "daily",
    # weekly
    "weekly": "weekly",
    "fortnightly": "weekly",
    "biweekly": "weekly",
    # monthly
    "monthly": "monthly",
    "quarterly": "monthly",
    "half yearly": "monthly",
    "halfyearly": "monthly",
    "half_yearly": "monthly",
    "sixmonthly": "monthly",
    "6monthly": "monthly",
    # yearly
    "yearly": "yearly",
    "annual": "yearly",
    "biannual": "yearly",
    "biennial": "yearly",
    # hourly (running hours)
    "hourly": "hourly",
    "hours": "hourly",
    "runhours": "hourly",
    "runninghours": "hourly",
    "running_hours": "hourly",
    "runninghour": "hourly",
}
CS_MAP = {_norm_text(member.value): member for member in ClassSociety}
CS_MAP.update(
    {
        "dnv": ClassSociety.dnv_gl,
        "dnvgl": ClassSociety.dnv_gl,
        "lr": ClassSociety.lr,
        "lloydsregister": ClassSociety.lr,
        "bv": ClassSociety.bv,
        "bureauveritas": ClassSociety.bv,
        "nk": ClassSociety.classnk,
        "classnk": ClassSociety.classnk,
        "kr": ClassSociety.kr,
        "koreanregister": ClassSociety.kr,
        "irs": ClassSociety.irs,
        "irclass": ClassSociety.irs,
        "indianregisterofshipping": ClassSociety.irs,
    }
)
CLASS_LIBRARY_VALUES = {
    _norm_text(member.value)
    for member in ClassSociety
    if member != ClassSociety.general
}
CLASS_LIBRARY_VALUES.update(
    {
        _norm_text(f"{member.__class__.__name__}.{member.name}")
        for member in ClassSociety
        if member != ClassSociety.general
    }
)
CLASS_WORKBOOK_SKIP_CATEGORIES = {"class surveys", "statutory surveys"}
CLASS_WORKBOOK_REQUIRED_HEADERS = {
    "standardjobname",
    "jobaction",
    "surveycategory",
    "componentnamepms",
    "frequencyinterval",
}


async def _ensure_class_society_enum_members() -> None:
    """
    Keep class-society imports deploy-safe.

    Railway container startup should stay fast, so we avoid schema work in the
    global boot command. Instead, when class-job flows need KR / IRS values, we
    extend the PostgreSQL enum just-in-time in autocommit mode. The block is
    idempotent and safely no-ops if the enum type is already absent or already
    contains the values.
    """
    sql = text(
        """
        DO $$
        BEGIN
            IF EXISTS (SELECT 1 FROM pg_type WHERE typname = 'class_society') THEN
                ALTER TYPE class_society ADD VALUE IF NOT EXISTS 'KR';
                ALTER TYPE class_society ADD VALUE IF NOT EXISTS 'IRS';
            END IF;
        END
        $$;
        """
    )
    async with engine.connect() as connection:
        autocommit_connection = await connection.execution_options(
            isolation_level="AUTOCOMMIT"
        )
        await autocommit_connection.execute(sql)


def _split_class_interval_candidates(value: str) -> list[str]:
    return [
        part.strip()
        for part in re.split(r"/|,|;|\band\b", value, flags=re.IGNORECASE)
        if part and part.strip()
    ]


def _parse_class_frequency(value: Any) -> tuple[Optional[int], Optional[Any], Optional[str]]:
    raw = _clean_text(value)
    if not raw:
        return None, None, None

    normalized = _norm_text(raw).replace("~", "").replace("approximately", "").strip()
    if not normalized:
        return None, None, raw

    for candidate in _split_class_interval_candidates(normalized):
        mapped_type = _map_frequency_type(candidate)
        if mapped_type is not None:
            if candidate in {"daily", "weekly", "monthly", "annual", "yearly", "hourly"}:
                base_frequency = 1
            else:
                number_match = re.search(r"(\d+(?:\.\d+)?)", candidate)
                base_frequency = None
                if number_match:
                    try:
                        numeric_value = float(number_match.group(1))
                    except ValueError:
                        numeric_value = None
                    if numeric_value is not None:
                        if mapped_type == FREQ_MAP["yearly"]:
                            if numeric_value.is_integer():
                                base_frequency = int(numeric_value)
                            else:
                                monthly_value = numeric_value * 12
                                if monthly_value.is_integer():
                                    return int(monthly_value), FREQ_MAP["monthly"], raw
                        elif numeric_value.is_integer():
                            base_frequency = int(numeric_value)

            if base_frequency is not None:
                return base_frequency, mapped_type, raw

    return None, None, raw


def _strip_class_location_tokens(value: Optional[str]) -> Optional[str]:
    text = _clean_text(value)
    if not text:
        return None

    cleaned = re.sub(
        r"\s*\((?=[^)]*\bno\.?\s*\d+)[^)]*\)",
        "",
        text,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\bno\.?\s*\d+[a-z/-]*\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\(\s*\)", "", cleaned)
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    cleaned = re.sub(r"\s+([,;/)])", r"\1", cleaned)
    cleaned = re.sub(r"([(])\s+", r"\1", cleaned)
    cleaned = cleaned.strip(" -—,;/")
    return cleaned or text


def _match_class_society(value: Any) -> Optional[ClassSociety]:
    normalized = _norm_text(value)
    if not normalized:
        return None
    mapped = CS_MAP.get(normalized)
    if mapped is not None:
        return mapped
    for alias, member in sorted(CS_MAP.items(), key=lambda item: len(item[0]), reverse=True):
        if alias and alias in normalized:
            return member
    return None


def _class_library_reference(class_society: ClassSociety, survey_category: Optional[str]) -> Optional[str]:
    if survey_category:
        return f"{class_society.value} / {survey_category}"[:200]
    return class_society.value[:200]


def _class_job_description(
    *,
    survey_category: Optional[str],
    job_action: Optional[str],
    raw_interval: Optional[str],
) -> Optional[str]:
    lines = ["Imported from class workbook"]
    if survey_category:
        lines.append(f"Survey category: {survey_category}")
    if job_action:
        lines.append(f"Job action: {job_action}")
    if raw_interval:
        lines.append(f"Imported interval: {raw_interval}")
    return "\n".join(lines)


def _sheet_looks_like_class_job_library(headers: list[str]) -> bool:
    header_set = {header for header in headers if header}
    return CLASS_WORKBOOK_REQUIRED_HEADERS.issubset(header_set)


def _extract_class_workbook_rows(wb: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx: Optional[int] = None
        headers: list[str] = []
        class_society: Optional[ClassSociety] = None

        preview_rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
        title_text = " ".join(str(cell or "") for row in preview_rows[:2] for cell in row if cell)
        class_society = _match_class_society(f"{sheet_name} {title_text}")

        for row_idx, row in enumerate(preview_rows, start=1):
            candidate_headers = [_header_key(cell) for cell in row]
            if _sheet_looks_like_class_job_library(candidate_headers):
                header_row_idx = row_idx
                headers = candidate_headers
                break

        if header_row_idx is None or class_society is None:
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            mapped = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            survey_category = _clean_text(mapped.get("surveycategory"))
            if not survey_category:
                continue
            if _norm_text(survey_category) in CLASS_WORKBOOK_SKIP_CATEGORIES:
                continue

            job_name = _strip_class_location_tokens(mapped.get("standardjobname"))
            machinery_type = _strip_class_location_tokens(mapped.get("componentnamepms"))
            if not job_name or not machinery_type:
                continue

            frequency, frequency_type, raw_interval = _parse_class_frequency(mapped.get("frequencyinterval"))
            rows.append(
                {
                    "class_society": class_society,
                    "machinery_type": machinery_type,
                    "job_name": job_name,
                    "job_description": _class_job_description(
                        survey_category=survey_category,
                        job_action=_clean_text(mapped.get("jobaction")),
                        raw_interval=raw_interval,
                    ),
                    "performing_rank": None,
                    "verifying_rank": None,
                    "frequency": frequency,
                    "frequency_type": frequency_type,
                    "is_critical": False,
                    "library_reference": _class_library_reference(class_society, survey_category),
                    "is_system": False,
                }
            )
    return rows


def _build_standard_jobs_query(
    *,
    tenant_id: uuid.UUID,
    class_society: Optional[str] = None,
    machinery_type: Optional[str] = None,
    is_critical: Optional[bool] = None,
    job_type: Optional[str] = None,
    search: Optional[str] = None,
) -> tuple[Any, Any]:
    class_society_expr = func.lower(func.trim(StandardJob.class_society.cast(String)))
    query = select(StandardJob).where(
        StandardJob.tenant_id == tenant_id,
        StandardJob.is_deleted == False,
    )

    if job_type == "critical":
        query = query.where(StandardJob.is_critical.is_(True))
    elif job_type == "standard":
        query = query.where(
            StandardJob.is_critical.is_(False),
            or_(
                StandardJob.class_society.is_(None),
                ~class_society_expr.in_(sorted(CLASS_LIBRARY_VALUES)),
            ),
        )
    elif job_type == "class":
        query = query.where(
            StandardJob.is_critical.is_(False),
            class_society_expr.in_(sorted(CLASS_LIBRARY_VALUES)),
        )

    if class_society:
        normalized_society = _norm_text(class_society)
        mapped_society = CS_MAP.get(normalized_society)
        if mapped_society is not None:
            allowed_values = {
                mapped_society.value.lower(),
                f"{mapped_society.__class__.__name__}.{mapped_society.name}".lower(),
            }
            query = query.where(class_society_expr.in_(sorted(allowed_values)))
        else:
            query = query.where(class_society_expr == normalized_society)

    if machinery_type:
        query = query.where(StandardJob.machinery_type.ilike(f"%{machinery_type.strip()}%"))

    if is_critical is not None:
        query = query.where(StandardJob.is_critical.is_(is_critical))

    if search:
        term = f"%{search.strip()}%"
        query = query.where(
            or_(
                StandardJob.job_name.ilike(term),
                StandardJob.machinery_type.ilike(term),
                StandardJob.job_description.ilike(term),
                StandardJob.library_reference.ilike(term),
                StandardJob.class_society.cast(String).ilike(term),
                StandardJob.frequency.cast(String).ilike(term),
                StandardJob.frequency_type.cast(String).ilike(term),
            )
        )

    return query, class_society_expr


def _find_jobs_review_job_for_standard_job(
    std_job: StandardJob,
    jobs_review_jobs: list[Job],
) -> Optional[Job]:
    for candidate in jobs_review_jobs:
        if _job_looks_library_imported(candidate, std_job):
            return candidate
    return None


def _serialize_standard_job_match(
    match: StandardJobMatch,
    *,
    job_lookup: dict[uuid.UUID, Job],
    standard_job_lookup: dict[uuid.UUID, StandardJob],
    jobs_review_jobs: list[Job],
) -> dict[str, Any]:
    matched_job = job_lookup.get(match.matched_job_id) if match.matched_job_id else None
    standard_job = standard_job_lookup.get(match.standard_job_id)
    review_job = _find_jobs_review_job_for_standard_job(standard_job, jobs_review_jobs) if standard_job else None
    return {
        "id": str(match.id),
        "standard_job_id": str(match.standard_job_id),
        "matched_job_id": str(match.matched_job_id) if match.matched_job_id else None,
        "match_status": match.match_status.value,
        "match_score": match.match_score,
        "not_applicable_reason": match.not_applicable_reason,
        "matched_job_name": matched_job.job_name if matched_job else None,
        "matched_job_code": matched_job.job_code if matched_job else None,
        "matched_job_description": matched_job.job_description if matched_job else None,
        "matched_job_qc_status": matched_job.qc_status.value if matched_job and matched_job.qc_status else None,
        "matched_job_origin": (
            "manual"
            if matched_job and matched_job.source_manual_id is not None
            else "review"
            if matched_job
            else None
        ),
        "jobs_review_job_id": str(review_job.id) if review_job else None,
        "jobs_review_job_name": review_job.job_name if review_job else None,
        "jobs_review_job_qc_status": review_job.qc_status.value if review_job and review_job.qc_status else None,
    }
WORKBOOK_SHEETS = {"annex1pmsjobs", "auditstandardjobs", "annexjobtitle", "criticaljobs"}
IMPORTABLE_JOB_HEADERS = {"jobname", "jobtitle", "jobdescription", "frequencytype", "frequency", "iscritical"}


def _map_frequency_type(value: Any) -> Optional[Any]:
    normalized = _norm_text(value).replace("-", " ")
    compact = normalized.replace(" ", "")
    if not normalized or normalized in {"na", "n/a", "maker instruction", "makers instruction"}:
        return None
    mapped = FREQ_ALIASES.get(normalized) or FREQ_ALIASES.get(compact)
    return FREQ_MAP.get(mapped) if mapped else None


def _choose_frequency(row: dict[str, Any]) -> tuple[Optional[int], Optional[Any]]:
    primary_type = _map_frequency_type(row.get("frequencytype"))
    primary_freq = _coerce_int(row.get("frequency"))
    alt_type = _map_frequency_type(row.get("alternatefrequencytype"))
    alt_freq = _coerce_int(row.get("alternatefrequency"))
    if primary_type is not None:
        return primary_freq, primary_type
    if alt_type is not None:
        return alt_freq, alt_type
    return primary_freq, None


def _derive_machinery_type(raw_name: Optional[str], fallback_name: Optional[str], group: Optional[str]) -> str:
    source = raw_name or fallback_name or group or "General"
    source = re.sub(r"\s+", " ", source).strip()
    parts = [part.strip() for part in re.split(r"\s*-\s*", source) if part.strip()]
    if len(parts) >= 2:
        return " - ".join(parts[:2])
    return parts[0] if parts else source


def _build_library_reference(sheet_name: str, row: dict[str, Any]) -> Optional[str]:
    for key in ("procedurereferencecode", "operationalprocedurecode", "safetyprocedurecode", "documentreferencecode"):
        value = _clean_text(row.get(key))
        if value:
            return value
    remarks = _clean_text(row.get("remarks")) or _clean_text(row.get(""))
    if remarks:
        return f"{sheet_name}: {remarks}"[:200]
    return sheet_name[:200]


def _sheet_looks_like_job_library(sheet_key: str, headers: list[str]) -> bool:
    if sheet_key in WORKBOOK_SHEETS:
        return True
    header_set = {header for header in headers if header}
    return len(header_set.intersection(IMPORTABLE_JOB_HEADERS)) >= 3 and (
        "jobname" in header_set or "jobtitle" in header_set
    )


def _canonical_standard_row(sheet_name: str, row: dict[str, Any], *, default_cs: ClassSociety) -> Optional[dict[str, Any]]:
    if _coerce_bool(row.get("isinactive")):
        return None

    job_name = (
        _clean_text(row.get("jobtitle"))
        or _clean_text(row.get("job title"))
        or _clean_text(row.get("jobname"))
    )
    raw_name = _clean_text(row.get("jobname")) or job_name
    if not job_name:
        return None

    frequency, frequency_type = _choose_frequency(row)
    is_critical = _coerce_bool(row.get("iscritical")) or _coerce_bool(row.get("is_critical"))
    class_society = default_cs
    if row.get("classsociety"):
        class_society = CS_MAP.get(_norm_text(row.get("classsociety")), default_cs)

    responsibility = _first_text(
        row,
        "performingrank",
        "performing_rank",
        "performerank",
        "rank",
        "rankname",
        "responsibility",
        "responsible",
        "performedby",
    )
    verifying_rank = _first_text(
        row,
        "verifyingrank",
        "verifying_rank",
        "verifierrank",
        "verifier",
        "verifiedby",
        "checkedby",
        "approvedby",
    )
    return {
        "class_society": class_society,
        "machinery_type": _derive_machinery_type(raw_name, job_name, _clean_text(row.get("jobgroup"))),
        "job_name": job_name,
        "job_description": _clean_text(row.get("jobdescription")),
        "performing_rank": responsibility,
        "verifying_rank": verifying_rank,
        "frequency": frequency,
        "frequency_type": frequency_type,
        "is_critical": is_critical,
        "library_reference": _build_library_reference(sheet_name, row),
        "is_system": False,
    }


def _standard_job_key(
    *,
    class_society: ClassSociety,
    machinery_type: str,
    job_name: str,
    is_critical: bool,
) -> tuple[str, str, str, bool]:
    return (
        class_society.value.lower(),
        _norm_text(machinery_type),
        _norm_text(job_name),
        is_critical,
    )


def _extract_structured_workbook_rows(wb: Any, *, job_type: str) -> list[dict[str, Any]]:
    if job_type not in {"standard", "critical"}:
        return []

    rows: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_key = _header_key(sheet_name)
        headers = [_header_key(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not _sheet_looks_like_job_library(sheet_key, headers):
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            mapped = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            canonical = _canonical_standard_row(sheet_name, mapped, default_cs=ClassSociety.general)
            if canonical is not None:
                is_critical = bool(canonical.get("is_critical"))
                if job_type == "critical":
                    if not is_critical:
                        continue
                    canonical["is_critical"] = True
                elif is_critical:
                    continue
                rows.append(canonical)
    return rows


def _extract_rows_from_upload(content: bytes, filename: str, *, job_type: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = [{_header_key(k): v for k, v in dict(r).items()} for r in reader]
    else:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        if job_type == "class":
            class_rows = _extract_class_workbook_rows(wb)
            if class_rows:
                return class_rows
        structured_rows = _extract_structured_workbook_rows(wb, job_type=job_type)
        if structured_rows:
            return structured_rows
        ws = wb.active
        headers = [_header_key(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return rows


def _score_standard_job(std_job: StandardJob, vessel_job: Job, component_lookup: dict[uuid.UUID, Component]) -> int:
    std_name = _norm_text(std_job.job_name)
    vessel_name = _norm_text(vessel_job.job_name)
    if not std_name or not vessel_name:
        return 0

    ratio = SequenceMatcher(None, std_name, vessel_name).ratio()
    score = int(ratio * 100)
    if std_name == vessel_name:
        score = max(score, 100)
    if std_job.frequency and vessel_job.frequency and std_job.frequency == vessel_job.frequency:
        score += 5
    if std_job.frequency_type and vessel_job.frequency_type and std_job.frequency_type == vessel_job.frequency_type:
        score += 5

    linked_component = component_lookup.get(vessel_job.component_id) if vessel_job.component_id else None
    machinery_context = _norm_text(linked_component.main_machinery if linked_component else "")
    std_machinery = _norm_text(std_job.machinery_type)
    if std_machinery and machinery_context and (
        std_machinery in machinery_context or machinery_context in std_machinery
    ):
        score += 10

    std_tokens = _tokenize_text(std_job.job_name)
    vessel_tokens = _tokenize_text(vessel_job.job_name)
    if std_tokens and vessel_tokens:
        overlap = len(std_tokens & vessel_tokens)
        if overlap == 0:
            score -= 35
        elif overlap == 1:
            score -= 15
        else:
            score += min(15, overlap * 4)

        directional_groups = [
            {"inside", "outside"},
            {"open", "close"},
            {"port", "starboard"},
            {"left", "right"},
            {"forward", "aft"},
            {"upper", "lower"},
        ]
        for group in directional_groups:
            std_dir = std_tokens & group
            vessel_dir = vessel_tokens & group
            if std_dir and vessel_dir and std_dir != vessel_dir:
                score -= 30

    return min(score, 100)


def _merge_source_reference(existing: Optional[str], new_value: Optional[str]) -> Optional[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for raw in (existing, new_value):
        for part in (raw or "").split(" | "):
            cleaned = part.strip()
            if not cleaned:
                continue
            key = cleaned.lower()
            if key in seen:
                continue
            seen.add(key)
            merged.append(cleaned)
    return " | ".join(merged) if merged else None


def _remove_source_reference(existing: Optional[str], remove_value: Optional[str]) -> Optional[str]:
    if not existing:
        return None
    remove_key = _norm_text(remove_value)
    kept: list[str] = []
    for part in existing.split(" | "):
        cleaned = part.strip()
        if not cleaned:
            continue
        if remove_key and _norm_text(cleaned) == remove_key:
            continue
        kept.append(cleaned)
    return " | ".join(kept) if kept else None


async def _resolve_component_mapping(
    *,
    vessel_id: uuid.UUID,
    component_id: Optional[str],
    db: AsyncSession,
) -> Optional[uuid.UUID]:
    if not component_id:
        return None
    try:
        component_uuid = uuid.UUID(component_id)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Invalid component mapping selected") from exc

    result = await db.execute(
        select(Component).where(
            Component.id == component_uuid,
            Component.vessel_id == vessel_id,
            Component.is_deleted == False,
        )
    )
    component = result.scalar_one_or_none()
    if component is None:
        raise HTTPException(status_code=404, detail="Mapped component not found on this vessel")
    return component_uuid


async def _upsert_standard_job_match(
    *,
    vessel_id: uuid.UUID,
    standard_job_id: uuid.UUID,
    matched_job_id: Optional[uuid.UUID],
    current_user: User,
    db: AsyncSession,
    match_status: MatchStatus = MatchStatus.matched,
    match_score: int = 100,
) -> None:
    result = await db.execute(
        select(StandardJobMatch).where(
            StandardJobMatch.vessel_id == vessel_id,
            StandardJobMatch.standard_job_id == standard_job_id,
            StandardJobMatch.is_deleted == False,
        )
    )
    match = result.scalar_one_or_none()
    if match is None:
        match = StandardJobMatch(
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            standard_job_id=standard_job_id,
        )
    match.matched_job_id = matched_job_id
    match.match_status = match_status
    match.match_score = match_score if matched_job_id else 0
    db.add(match)


async def _import_standard_job_to_vessel(
    *,
    vessel_id: uuid.UUID,
    std_job: StandardJob,
    component_id: Optional[uuid.UUID],
    cms_id: Optional[str],
    current_user: User,
    db: AsyncSession,
) -> tuple[str, Job]:
    cms_id = _clean_text(cms_id)
    inferred_rank = None
    if not std_job.performing_rank and component_id:
        inferred_rank = await infer_rank_from_component(
            db,
            vessel_id=vessel_id,
            component_id=component_id,
        )
        inferred_rank = normalize_rank_name(inferred_rank)
        if inferred_rank and inferred_rank != std_job.performing_rank:
            std_job.performing_rank = inferred_rank
            db.add(std_job)
            await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=std_job.performing_rank)
    existing_match_result = await db.execute(
        select(StandardJobMatch).where(
            StandardJobMatch.vessel_id == vessel_id,
            StandardJobMatch.standard_job_id == std_job.id,
            StandardJobMatch.is_deleted == False,
        )
    )
    existing_match = existing_match_result.scalar_one_or_none()
    if existing_match and existing_match.matched_job_id:
        matched_job_result = await db.execute(
            select(Job).where(
                Job.id == existing_match.matched_job_id,
                Job.vessel_id == vessel_id,
                Job.is_deleted == False,
            )
        )
        matched_job = matched_job_result.scalar_one_or_none()
        if matched_job is not None:
            if matched_job.job_name != std_job.job_name:
                matched_job.job_name = std_job.job_name
            matched_job.is_critical = bool(matched_job.is_critical or std_job.is_critical)
            matched_job.source_reference = _merge_source_reference(
                matched_job.source_reference,
                std_job.library_reference,
            )
            if component_id is not None:
                matched_job.component_id = component_id
                matched_job.is_unmapped = False
            if cms_id and cms_id != matched_job.cms_id:
                matched_job.cms_id = cms_id
            if not matched_job.performing_rank and (std_job.performing_rank or inferred_rank):
                matched_job.performing_rank = std_job.performing_rank or inferred_rank
            if not matched_job.verifying_rank and std_job.verifying_rank:
                matched_job.verifying_rank = std_job.verifying_rank
            if not matched_job.job_description and std_job.job_description:
                matched_job.job_description = std_job.job_description
            db.add(matched_job)
            await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=matched_job.performing_rank)
            await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=matched_job.verifying_rank)
            await _upsert_standard_job_match(
                vessel_id=vessel_id,
                standard_job_id=std_job.id,
                matched_job_id=matched_job.id,
                current_user=current_user,
                db=db,
            )
            return "merged", matched_job

    existing_job_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.job_name == std_job.job_name,
            Job.is_deleted == False,
        )
    )
    existing_job = existing_job_result.scalar_one_or_none()
    if existing_job is not None:
        if existing_job.job_name != std_job.job_name:
            existing_job.job_name = std_job.job_name
        existing_job.is_critical = bool(existing_job.is_critical or std_job.is_critical)
        existing_job.source_reference = _merge_source_reference(
            existing_job.source_reference,
            std_job.library_reference,
        )
        if component_id is not None:
            existing_job.component_id = component_id
            existing_job.is_unmapped = False
        if cms_id and cms_id != existing_job.cms_id:
            existing_job.cms_id = cms_id
        if not existing_job.performing_rank and (std_job.performing_rank or inferred_rank):
            existing_job.performing_rank = std_job.performing_rank or inferred_rank
        if not existing_job.verifying_rank and std_job.verifying_rank:
            existing_job.verifying_rank = std_job.verifying_rank
        if not existing_job.job_description and std_job.job_description:
            existing_job.job_description = std_job.job_description
        db.add(existing_job)
        await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=existing_job.performing_rank)
        await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=existing_job.verifying_rank)
        await _upsert_standard_job_match(
            vessel_id=vessel_id,
            standard_job_id=std_job.id,
            matched_job_id=existing_job.id,
            current_user=current_user,
            db=db,
        )
        return "merged", existing_job

    new_job = Job(
        tenant_id=current_user.tenant_id,
        vessel_id=vessel_id,
        job_name=std_job.job_name,
        job_description=std_job.job_description,
        performing_rank=std_job.performing_rank or inferred_rank,
        verifying_rank=std_job.verifying_rank,
        frequency=std_job.frequency,
        frequency_type=std_job.frequency_type,
        is_critical=std_job.is_critical,
        source_reference=std_job.library_reference,
        cms_id=cms_id,
        qc_status=QCStatus.pending,
        component_id=component_id,
        is_unmapped=component_id is None,
    )
    db.add(new_job)
    await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=new_job.performing_rank)
    await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=new_job.verifying_rank)
    await db.flush()
    await _upsert_standard_job_match(
        vessel_id=vessel_id,
        standard_job_id=std_job.id,
        matched_job_id=new_job.id,
        current_user=current_user,
        db=db,
    )
    return "imported", new_job


def _job_looks_library_imported(job: Job, std_job: StandardJob) -> bool:
    source_ref = _norm_text(job.source_reference)
    std_ref = _norm_text(std_job.library_reference)
    if job.source_manual_id is None and _norm_text(job.job_name) == _norm_text(std_job.job_name):
        return True
    if source_ref and std_ref and std_ref in source_ref:
        return True
    return False


def _job_matches_critical_reference(job: Job, library_reference: Optional[str]) -> bool:
    source_ref = _norm_text(job.source_reference)
    std_ref = _norm_text(library_reference)
    return bool(source_ref and std_ref and std_ref in source_ref)


def _job_is_library_added_critical(job: Job, std_job: Optional[StandardJob] = None) -> bool:
    if job.source_manual_id is not None:
        return False
    if std_job is not None and _norm_text(job.job_name) == _norm_text(std_job.job_name):
        return True
    if std_job is not None and _job_matches_critical_reference(job, std_job.library_reference):
        return True
    return bool(job.source_manual_id is None and job.source_reference and "critical" in _norm_text(job.source_reference))


@router.get("/standard-jobs", summary="List standard jobs library")
async def list_standard_jobs(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    class_society: Optional[str] = Query(None),
    machinery_type: Optional[str] = Query(None),
    is_critical: Optional[bool] = Query(None),
    job_type: Optional[str] = Query(None, description="'standard', 'class', or 'critical'"),
    search: Optional[str] = Query(None),
    sort_by: str = Query("job_name"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> dict[str, Any]:
    query, class_society_expr = _build_standard_jobs_query(
        tenant_id=current_user.tenant_id,
        class_society=class_society,
        machinery_type=machinery_type,
        is_critical=is_critical,
        job_type=job_type,
        search=search,
    )

    sort_columns = {
        "job_name": StandardJob.job_name,
        "machinery_type": StandardJob.machinery_type,
        "class_society": StandardJob.class_society.cast(String),
        "frequency": StandardJob.frequency,
        "frequency_type": StandardJob.frequency_type.cast(String),
        "critical": StandardJob.is_critical,
        "reference": StandardJob.library_reference,
        "job_type": class_society_expr,
    }
    order_col = sort_columns.get(sort_by, StandardJob.job_name)
    order_expr = desc(order_col) if sort_order == "desc" else asc(order_col)

    total = (
        await db.execute(select(func.count()).select_from(query.order_by(None).subquery()))
    ).scalar_one()
    total_pages = max(1, (total + page_size - 1) // page_size)
    jobs_result = await db.execute(
        query.order_by(order_expr).offset((page - 1) * page_size).limit(page_size)
    )
    page_jobs = jobs_result.scalars().all()
    jobs = [_serialize_standard_job(job) for job in page_jobs]
    return {
        "items": jobs,
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages,
    }


@router.post("/standard-jobs/bulk-import", summary="Bulk import standard or class jobs from Excel/CSV")
async def bulk_import_standard_jobs(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    file: UploadFile = File(...),
    job_type: str = Query("standard", description="'standard' or 'class'"),
) -> dict[str, Any]:
    """Import Standard Jobs or Class Society Jobs from Excel/CSV."""
    content = await file.read()
    filename = (file.filename or "").lower()
    validate_uploaded_file_bytes(
        filename=file.filename or "standard_jobs_import.xlsx",
        content=content,
        allowed_extensions={"csv", "xlsx"},
        max_size_bytes=25 * 1024 * 1024,
    )
    await _ensure_class_society_enum_members()
    if job_type not in {"standard", "class", "critical"}:
        raise HTTPException(status_code=400, detail="job_type must be 'standard', 'class', or 'critical'")
    rows = _extract_rows_from_upload(content, filename, job_type=job_type)
    parsed_rows = len(rows)
    default_cs = ClassSociety.general if job_type == "standard" else None
    if job_type == "critical":
        default_cs = ClassSociety.general

    existing_result = await db.execute(
        select(StandardJob).where(
            StandardJob.tenant_id == current_user.tenant_id,
            StandardJob.is_deleted == False,
        )
    )
    existing_jobs = {
        _standard_job_key(
            class_society=job.class_society,
            machinery_type=job.machinery_type,
            job_name=job.job_name,
            is_critical=job.is_critical,
        ): job
        for job in existing_result.scalars().all()
    }

    if not rows:
        raise HTTPException(
            status_code=400,
            detail=(
                "No importable rows found. Use Audit standard jobs / Annex Job Title for Standard Jobs, "
                "Critical Jobs for the Critical Jobs tab, or the class survey workbook vessel sheets for Class Jobs."
            ),
        )

    imported = 0
    updated = 0
    unchanged = 0
    skipped = 0
    seen_upload_keys: set[tuple[str, str, str, bool]] = set()
    for row in rows:
        if "class_society" in row:
            canonical = row
        else:
            job_name = _clean_text(row.get("jobname") or row.get("job_name"))
            machinery_type = _clean_text(row.get("machinerytype") or row.get("machinery_type"))
            if not job_name or not machinery_type:
                skipped += 1
                continue
            responsibility = _first_text(
                row,
                "performingrank",
                "performing_rank",
                "performerank",
                "rank",
                "rankname",
                "responsibility",
                "responsible",
                "performedby",
            )
            verifying_rank = _first_text(
                row,
                "verifyingrank",
                "verifying_rank",
                "verifierrank",
                "verifier",
                "verifiedby",
                "checkedby",
                "approvedby",
            )
            cs_raw = _norm_text(row.get("classsociety") or row.get("class_society"))
            class_society = CS_MAP.get(cs_raw, default_cs)
            if class_society is None:
                skipped += 1
                continue
            canonical = {
                "class_society": class_society,
                "machinery_type": machinery_type,
                "job_name": job_name,
                "job_description": _clean_text(row.get("jobdescription") or row.get("job_description")),
                "performing_rank": responsibility,
                "verifying_rank": verifying_rank,
                "frequency": _coerce_int(row.get("frequency")),
                "frequency_type": _map_frequency_type(row.get("frequencytype") or row.get("frequency_type")),
                "is_critical": _coerce_bool(row.get("iscritical") or row.get("is_critical")),
                "library_reference": _clean_text(row.get("libraryreference") or row.get("library_reference")),
                "is_system": False,
            }

        canonical["performing_rank"] = normalize_rank_name(canonical.get("performing_rank"))
        canonical["verifying_rank"] = normalize_rank_name(canonical.get("verifying_rank"))

        key = _standard_job_key(
            class_society=canonical["class_society"],
            machinery_type=canonical["machinery_type"],
            job_name=canonical["job_name"],
            is_critical=canonical["is_critical"],
        )
        if key in seen_upload_keys:
            skipped += 1
            continue
        seen_upload_keys.add(key)

        existing = existing_jobs.get(key)
        if existing is None and canonical["is_critical"]:
            existing = existing_jobs.get(
                _standard_job_key(
                    class_society=canonical["class_society"],
                    machinery_type=canonical["machinery_type"],
                    job_name=canonical["job_name"],
                    is_critical=False,
                )
            )
        if existing is None:
            db.add(
                StandardJob(
                    tenant_id=current_user.tenant_id,
                    class_society=canonical["class_society"],
                    machinery_type=canonical["machinery_type"],
                    job_name=canonical["job_name"],
                    job_description=canonical["job_description"],
                    performing_rank=canonical.get("performing_rank"),
                    verifying_rank=canonical.get("verifying_rank"),
                    frequency=canonical["frequency"],
                    frequency_type=canonical["frequency_type"],
                    is_critical=canonical["is_critical"],
                    library_reference=canonical["library_reference"],
                    is_system=False,
                )
            )
            await ensure_job_rank(
                db,
                tenant_id=current_user.tenant_id,
                rank_name=canonical.get("performing_rank"),
            )
            await ensure_job_rank(
                db,
                tenant_id=current_user.tenant_id,
                rank_name=canonical.get("verifying_rank"),
            )
            imported += 1
            continue

        changed = False
        if canonical["job_description"] and canonical["job_description"] != existing.job_description:
            existing.job_description = canonical["job_description"]
            changed = True
        if canonical["frequency"] is not None and canonical["frequency"] != existing.frequency:
            existing.frequency = canonical["frequency"]
            changed = True
        if canonical["frequency_type"] is not None and canonical["frequency_type"] != existing.frequency_type:
            existing.frequency_type = canonical["frequency_type"]
            changed = True
        if canonical["library_reference"] and canonical["library_reference"] != existing.library_reference:
            existing.library_reference = canonical["library_reference"]
            changed = True
        if canonical.get("performing_rank") and canonical.get("performing_rank") != existing.performing_rank:
            existing.performing_rank = canonical.get("performing_rank")
            changed = True
        if canonical.get("verifying_rank") and canonical.get("verifying_rank") != existing.verifying_rank:
            existing.verifying_rank = canonical.get("verifying_rank")
            changed = True
        if canonical["is_critical"] and not existing.is_critical:
            existing.is_critical = True
            changed = True
            existing_jobs[key] = existing
        if changed:
            db.add(existing)
            await ensure_job_rank(
                db,
                tenant_id=current_user.tenant_id,
                rank_name=existing.performing_rank,
            )
            await ensure_job_rank(
                db,
                tenant_id=current_user.tenant_id,
                rank_name=existing.verifying_rank,
            )
            updated += 1
        else:
            unchanged += 1

    await db.commit()
    if imported == 0 and updated == 0 and unchanged == 0 and skipped == parsed_rows:
        raise HTTPException(
            status_code=400,
            detail="No valid rows were imported. Please check that the selected tab matches the workbook content and required job columns are present.",
        )
    return {
        "parsed_rows": parsed_rows,
        "imported": imported,
        "updated": updated,
        "unchanged": unchanged,
        "skipped": skipped,
        "job_type": job_type,
    }


@router.post("/standard-jobs", summary="Create a standard or class job")
async def create_standard_job(
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    job_name = _clean_text(body.get("job_name"))
    machinery_type = _clean_text(body.get("machinery_type"))
    if not job_name or not machinery_type:
        raise HTTPException(status_code=400, detail="job_name and machinery_type are required")

    await _ensure_class_society_enum_members()
    class_society_raw = _clean_text(body.get("class_society")) or "General"
    class_society = CS_MAP.get(_norm_text(class_society_raw))
    if class_society is None:
        raise HTTPException(status_code=400, detail="Invalid class_society")

    std_job = StandardJob(
        tenant_id=current_user.tenant_id,
        class_society=class_society,
        machinery_type=machinery_type,
        job_name=job_name,
        job_description=_clean_text(body.get("job_description")),
        performing_rank=normalize_rank_name(body.get("performing_rank")),
        verifying_rank=normalize_rank_name(body.get("verifying_rank")),
        frequency=_coerce_int(body.get("frequency")),
        frequency_type=_map_frequency_type(body.get("frequency_type")),
        is_critical=_coerce_bool(body.get("is_critical")),
        library_reference=_clean_text(body.get("library_reference")),
        is_system=False,
    )
    db.add(std_job)
    await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=std_job.performing_rank)
    await ensure_job_rank(db, tenant_id=current_user.tenant_id, rank_name=std_job.verifying_rank)
    await db.commit()
    await db.refresh(std_job)
    return _serialize_standard_job(std_job)


@router.delete("/standard-jobs/{standard_job_id}", summary="Delete a standard job")
async def delete_standard_job(
    standard_job_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    result = await db.execute(
        select(StandardJob).where(StandardJob.id == standard_job_id, StandardJob.is_deleted == False)
    )
    job = result.scalar_one_or_none()
    if job is None:
        raise HTTPException(status_code=404, detail="Standard job not found")
    job.is_deleted = True
    db.add(job)
    await db.commit()
    return {"deleted": True}


@router.get("/standard-jobs/vessel-types", summary="List vessel type templates")
async def list_vessel_types(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    result = await db.execute(
        select(VesselTypeTemplate).where(VesselTypeTemplate.is_deleted == False)
    )
    templates = result.scalars().all()
    return {
        "items": [
            {
                "id": str(t.id),
                "vessel_type": t.vessel_type,
                "machinery_group": t.machinery_group,
                "machinery_name": t.machinery_name,
                "is_mandatory": t.is_mandatory,
                "extraction_types": t.extraction_types,
            }
            for t in templates
        ]
    }


@router.post("/vessels/{vessel_id}/standard-jobs/run-comparison", summary="Run standard job matching")
async def run_comparison(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    body: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    vessel = await _get_vessel_or_404(vessel_id, db)

    body = body or {}

    query, _ = _build_standard_jobs_query(
        tenant_id=current_user.tenant_id,
        class_society=_clean_text(body.get("class_society")),
        machinery_type=_clean_text(body.get("machinery_type")),
        is_critical=False,
        job_type=_clean_text(body.get("job_type")),
        search=None,
    )
    std_jobs_result = await db.execute(query.order_by(None))
    std_jobs = std_jobs_result.scalars().all()

    # Get vessel jobs
    vessel_jobs_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.qc_status != QCStatus.rejected,
            Job.is_deleted == False,
            Job.source_manual_id.is_not(None),
        )
    )
    vessel_jobs = vessel_jobs_result.scalars().all()
    component_ids = {job.component_id for job in vessel_jobs if job.component_id}
    component_lookup: dict[uuid.UUID, Component] = {}
    if component_ids:
        component_result = await db.execute(select(Component).where(Component.id.in_(component_ids)))
        component_lookup = {component.id: component for component in component_result.scalars().all()}

    existing_result = await db.execute(
        select(StandardJobMatch).where(
            StandardJobMatch.vessel_id == vessel_id,
            StandardJobMatch.is_deleted == False,
        )
    )
    existing_matches = {match.standard_job_id: match for match in existing_result.scalars().all()}

    matches_created = 0
    best_scores: dict[uuid.UUID, tuple[Optional[uuid.UUID], int]] = {}
    candidates: list[tuple[int, uuid.UUID, uuid.UUID]] = []
    for std_job in std_jobs:
        top_job_id: Optional[uuid.UUID] = None
        top_score = 0
        for vessel_job in vessel_jobs:
            score = _score_standard_job(std_job, vessel_job, component_lookup)
            if score > top_score:
                top_score = score
                top_job_id = vessel_job.id
            if score >= 70:
                candidates.append((score, std_job.id, vessel_job.id))
        best_scores[std_job.id] = (top_job_id, top_score)

    assigned_std_jobs: set[uuid.UUID] = set()
    assigned_vessel_jobs: set[uuid.UUID] = set()
    final_matches: dict[uuid.UUID, tuple[Optional[uuid.UUID], int]] = {}
    for score, std_job_id, vessel_job_id in sorted(
        candidates,
        key=lambda item: (item[0], str(item[1]), str(item[2])),
        reverse=True,
    ):
        if std_job_id in assigned_std_jobs or vessel_job_id in assigned_vessel_jobs:
            continue
        assigned_std_jobs.add(std_job_id)
        assigned_vessel_jobs.add(vessel_job_id)
        final_matches[std_job_id] = (vessel_job_id, score)

    for std_job in std_jobs:
        matched_job_id, matched_score = final_matches.get(std_job.id, (None, 0))
        if matched_job_id is not None and matched_score >= 92:
            match_status = MatchStatus.matched
        elif matched_job_id is not None and matched_score >= 70:
            match_status = MatchStatus.partial
        else:
            fallback_job_id, fallback_score = best_scores.get(std_job.id, (None, 0))
            matched_job_id = None
            matched_score = 0
            match_status = MatchStatus.not_found
            if fallback_job_id is not None and fallback_score >= 70:
                match_status = MatchStatus.not_found

        match = existing_matches.get(std_job.id)
        if match is None:
            match = StandardJobMatch(
                tenant_id=current_user.tenant_id,
                vessel_id=vessel_id,
                standard_job_id=std_job.id,
            )
            matches_created += 1

        match.matched_job_id = matched_job_id
        match.match_status = match_status
        match.match_score = matched_score
        db.add(match)

    await db.commit()
    return {"status": "completed", "matches_created": matches_created}


@router.get("/vessels/{vessel_id}/standard-jobs/matches", summary="Get match results")
async def get_matches(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    match_status: Optional[str] = Query(None),
    standard_job_ids: Optional[str] = Query(None, description="Comma separated standard job ids"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)
    conditions = [
        StandardJobMatch.vessel_id == vessel_id,
        StandardJobMatch.is_deleted == False,
    ]
    if standard_job_ids:
        ids: list[uuid.UUID] = []
        for raw in standard_job_ids.split(","):
            value = raw.strip()
            if not value:
                continue
            try:
                ids.append(uuid.UUID(value))
            except ValueError:
                continue
        if ids:
            conditions.append(StandardJobMatch.standard_job_id.in_(ids))
    if match_status:
        try:
            conditions.append(StandardJobMatch.match_status == MatchStatus(match_status))
        except ValueError:
            pass
    total = await db.scalar(select(func.count()).select_from(StandardJobMatch).where(*conditions)) or 0
    total_pages = max(1, (total + page_size - 1) // page_size)
    query = (
        select(StandardJobMatch)
        .where(*conditions)
        .order_by(StandardJobMatch.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(query)
    matches = result.scalars().all()
    matched_job_ids = [match.matched_job_id for match in matches if match.matched_job_id]
    job_lookup: dict[uuid.UUID, Job] = {}
    if matched_job_ids:
        jobs_result = await db.execute(select(Job).where(Job.id.in_(matched_job_ids), Job.is_deleted == False))
        job_lookup = {job.id: job for job in jobs_result.scalars().all()}
    standard_job_lookup: dict[uuid.UUID, StandardJob] = {}
    standard_job_ids_for_matches = [match.standard_job_id for match in matches]
    if standard_job_ids_for_matches:
        std_jobs_result = await db.execute(
            select(StandardJob).where(StandardJob.id.in_(standard_job_ids_for_matches), StandardJob.is_deleted == False)
        )
        standard_job_lookup = {job.id: job for job in std_jobs_result.scalars().all()}
    jobs_review_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == current_user.tenant_id,
            Job.is_deleted == False,
            Job.source_manual_id.is_(None),
        )
    )
    jobs_review_jobs = jobs_review_result.scalars().all()
    return {
        "items": [
            _serialize_standard_job_match(
                m,
                job_lookup=job_lookup,
                standard_job_lookup=standard_job_lookup,
                jobs_review_jobs=jobs_review_jobs,
            )
            for m in matches
        ],
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages,
    }


@router.get("/vessels/{vessel_id}/standard-jobs/summary", summary="Get standard jobs comparison summary")
async def get_standard_jobs_summary(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    class_society: Optional[str] = Query(None),
    machinery_type: Optional[str] = Query(None),
    is_critical: Optional[bool] = Query(None),
    job_type: Optional[str] = Query(None, description="'standard', 'class', or 'critical'"),
    search: Optional[str] = Query(None),
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)
    query, _ = _build_standard_jobs_query(
        tenant_id=current_user.tenant_id,
        class_society=class_society,
        machinery_type=machinery_type,
        is_critical=is_critical,
        job_type=job_type,
        search=search,
    )
    std_jobs_result = await db.execute(query.order_by(None))
    std_jobs = std_jobs_result.scalars().all()
    if not std_jobs:
        return {
            "library_total": 0,
            "added_to_review_total": 0,
            "not_applicable_total": 0,
            "manual_linked_total": 0,
        }

    std_job_ids = [job.id for job in std_jobs]
    matches_result = await db.execute(
        select(StandardJobMatch).where(
            StandardJobMatch.vessel_id == vessel_id,
            StandardJobMatch.standard_job_id.in_(std_job_ids),
            StandardJobMatch.is_deleted == False,
        )
    )
    matches = matches_result.scalars().all()

    matched_job_ids = [match.matched_job_id for match in matches if match.matched_job_id]
    matched_jobs_lookup: dict[uuid.UUID, Job] = {}
    if matched_job_ids:
        matched_jobs_result = await db.execute(
            select(Job).where(Job.id.in_(matched_job_ids), Job.is_deleted == False)
        )
        matched_jobs_lookup = {job.id: job for job in matched_jobs_result.scalars().all()}

    jobs_review_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == current_user.tenant_id,
            Job.is_deleted == False,
            Job.source_manual_id.is_(None),
        )
    )
    jobs_review_jobs = jobs_review_result.scalars().all()

    added_to_review_total = sum(
        1 for std_job in std_jobs if _find_jobs_review_job_for_standard_job(std_job, jobs_review_jobs) is not None
    )
    not_applicable_total = sum(1 for match in matches if match.match_status == MatchStatus.not_applicable)
    manual_linked_total = sum(
        1
        for match in matches
        if match.matched_job_id in matched_jobs_lookup
        and matched_jobs_lookup[match.matched_job_id].source_manual_id is not None
    )
    return {
        "library_total": len(std_jobs),
        "added_to_review_total": added_to_review_total,
        "not_applicable_total": not_applicable_total,
        "manual_linked_total": manual_linked_total,
    }


@router.patch("/vessels/{vessel_id}/standard-jobs/matches/{match_id}", summary="Update match status")
async def update_match(
    vessel_id: uuid.UUID,
    match_id: uuid.UUID,
    body: dict[str, str],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    result = await db.execute(
        select(StandardJobMatch).where(
            StandardJobMatch.id == match_id,
            StandardJobMatch.vessel_id == vessel_id,
            StandardJobMatch.is_deleted == False,
        )
    )
    match = result.scalar_one_or_none()
    if match is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Match not found")

    if "match_status" in body:
        try:
            match.match_status = MatchStatus(body["match_status"])
        except ValueError:
            pass
    if "not_applicable_reason" in body:
        match.not_applicable_reason = body["not_applicable_reason"]

    db.add(match)
    await db.commit()
    return {"id": str(match.id), "match_status": match.match_status.value}


@router.post("/vessels/{vessel_id}/standard-jobs/import/{standard_job_id}", summary="Import standard job")
async def import_standard_job(
    vessel_id: uuid.UUID,
    standard_job_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    component_id: Optional[str] = Query(None),
    cms_id: Optional[str] = Query(None),
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    std_result = await db.execute(
        select(StandardJob).where(
            StandardJob.id == standard_job_id, StandardJob.is_deleted == False
        )
    )
    std_job = std_result.scalar_one_or_none()
    if std_job is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Standard job not found")

    mapped_component_id = await _resolve_component_mapping(
        vessel_id=vessel_id,
        component_id=component_id,
        db=db,
    )

    import_status, job = await _import_standard_job_to_vessel(
        vessel_id=vessel_id,
        std_job=std_job,
        component_id=mapped_component_id,
        cms_id=cms_id,
        current_user=current_user,
        db=db,
    )
    await db.commit()
    return {"status": import_status, "job_id": str(job.id)}


@router.post("/vessels/{vessel_id}/standard-jobs/import-batch", summary="Import selected or filtered standard jobs")
async def import_standard_jobs_batch(
    vessel_id: uuid.UUID,
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    selected_ids = [uuid.UUID(v) for v in body.get("standard_job_ids", []) if v]
    component_map_raw = body.get("component_map") or {}
    cms_id_map_raw = body.get("cms_id_map") or {}
    component_map: dict[str, uuid.UUID | None] = {}
    cms_id_map: dict[str, str] = {}
    if isinstance(component_map_raw, dict):
        for standard_job_id, component_id in component_map_raw.items():
            if not isinstance(standard_job_id, str):
                continue
            component_map[standard_job_id] = await _resolve_component_mapping(
                vessel_id=vessel_id,
                component_id=component_id,
                db=db,
            ) if component_id else None
    if isinstance(cms_id_map_raw, dict):
        for standard_job_id, cms_id in cms_id_map_raw.items():
            if not isinstance(standard_job_id, str):
                continue
            cleaned_cms_id = _clean_text(cms_id) if isinstance(cms_id, str) else None
            if cleaned_cms_id:
                cms_id_map[standard_job_id] = cleaned_cms_id
    query = select(StandardJob).where(
        StandardJob.tenant_id == current_user.tenant_id,
        StandardJob.is_deleted == False,
    )
    if selected_ids:
        query = query.where(StandardJob.id.in_(selected_ids))
    else:
        if body.get("job_type") == "standard":
            query = query.where(StandardJob.class_society == ClassSociety.general)
        elif body.get("job_type") == "class":
            query = query.where(StandardJob.class_society != ClassSociety.general)
        if body.get("class_society"):
            cs = CS_MAP.get(_norm_text(body["class_society"]))
            if cs:
                query = query.where(StandardJob.class_society == cs)
        if body.get("machinery_type"):
            query = query.where(StandardJob.machinery_type == body["machinery_type"])
        if body.get("include_critical") is False:
            query = query.where(StandardJob.is_critical == False)
    result = await db.execute(query)
    std_jobs = result.scalars().all()
    if not std_jobs:
        raise HTTPException(status_code=404, detail="No standard jobs found for import")

    imported = 0
    merged = 0
    imported_job_ids: list[str] = []
    merged_job_ids: list[str] = []
    for std_job in std_jobs:
        mapped_component_id = component_map.get(str(std_job.id))
        status_value, job = await _import_standard_job_to_vessel(
            vessel_id=vessel_id,
            std_job=std_job,
            component_id=mapped_component_id,
            cms_id=cms_id_map.get(str(std_job.id)),
            current_user=current_user,
            db=db,
        )
        if status_value == "imported":
            imported += 1
            imported_job_ids.append(str(job.id))
        else:
            merged += 1
            merged_job_ids.append(str(job.id))
    await db.commit()
    imported_job_ids = list(dict.fromkeys(imported_job_ids))
    merged_job_ids = list(dict.fromkeys(merged_job_ids))
    return {
        "imported": imported,
        "merged": merged,
        "total": len(std_jobs),
        "imported_job_ids": imported_job_ids,
        "merged_job_ids": merged_job_ids,
    }


@router.post("/vessels/{vessel_id}/standard-jobs/remove-batch", summary="Remove selected or filtered imported standard jobs from vessel")
async def remove_standard_jobs_batch(
    vessel_id: uuid.UUID,
    body: dict[str, Any],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    selected_ids = [uuid.UUID(v) for v in body.get("standard_job_ids", []) if v]
    query = select(StandardJob).where(
        StandardJob.tenant_id == current_user.tenant_id,
        StandardJob.is_deleted == False,
    )
    if selected_ids:
        query = query.where(StandardJob.id.in_(selected_ids))
    else:
        if body.get("job_type") == "standard":
            query = query.where(StandardJob.class_society == ClassSociety.general)
        elif body.get("job_type") == "class":
            query = query.where(StandardJob.class_society != ClassSociety.general)
        if body.get("class_society"):
            cs = CS_MAP.get(_norm_text(body["class_society"]))
            if cs:
                query = query.where(StandardJob.class_society == cs)
        if body.get("machinery_type"):
            query = query.where(StandardJob.machinery_type == body["machinery_type"])
        if body.get("include_critical") is False:
            query = query.where(StandardJob.is_critical == False)
    result = await db.execute(query)
    std_jobs = result.scalars().all()
    if not std_jobs:
        raise HTTPException(status_code=404, detail="No standard jobs found for removal")

    removed = 0
    skipped = 0
    for std_job in std_jobs:
        match_result = await db.execute(
            select(StandardJobMatch).where(
                StandardJobMatch.vessel_id == vessel_id,
                StandardJobMatch.standard_job_id == std_job.id,
                StandardJobMatch.is_deleted == False,
            )
        )
        match = match_result.scalar_one_or_none()
        if match is None or match.matched_job_id is None:
            skipped += 1
            continue
        job_result = await db.execute(
            select(Job).where(
                Job.id == match.matched_job_id,
                Job.vessel_id == vessel_id,
                Job.is_deleted == False,
            )
        )
        job = job_result.scalar_one_or_none()
        if job is None or not _job_looks_library_imported(job, std_job):
            skipped += 1
            continue
        job.is_deleted = True
        match.matched_job_id = None
        match.match_status = MatchStatus.not_found
        match.match_score = 0
        db.add(job)
        db.add(match)
        removed += 1

    await db.commit()
    return {"removed": removed, "skipped": skipped, "total": len(std_jobs)}


@router.post("/vessels/{vessel_id}/standard-jobs/add-critical-jobs", summary="Add missing critical standard jobs")
async def add_critical_jobs(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    critical_result = await db.execute(
        select(StandardJob).where(
            StandardJob.tenant_id == current_user.tenant_id,
            StandardJob.is_deleted == False,
            StandardJob.is_critical == True,
        )
    )
    critical_jobs = critical_result.scalars().all()

    vessel_jobs_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.tenant_id == current_user.tenant_id,
            Job.is_deleted == False,
            Job.qc_status != QCStatus.rejected,
        )
    )
    vessel_jobs = vessel_jobs_result.scalars().all()
    component_ids = {job.component_id for job in vessel_jobs if job.component_id}
    component_lookup: dict[uuid.UUID, Component] = {}
    if component_ids:
        component_result = await db.execute(select(Component).where(Component.id.in_(component_ids)))
        component_lookup = {component.id: component for component in component_result.scalars().all()}

    added = 0
    updated = 0
    skipped = 0
    for std_job in critical_jobs:
        best_match: Job | None = None
        best_score = 0
        for vessel_job in vessel_jobs:
            score = _score_standard_job(std_job, vessel_job, component_lookup)
            if score > best_score:
                best_score = score
                best_match = vessel_job

        if best_match is not None and best_score >= 92:
            changed = False
            if best_match.job_name != std_job.job_name:
                best_match.job_name = std_job.job_name
                changed = True
            if not best_match.is_critical:
                best_match.is_critical = True
                changed = True
            merged_reference = _merge_source_reference(best_match.source_reference, std_job.library_reference)
            if merged_reference != best_match.source_reference:
                best_match.source_reference = merged_reference
                changed = True
            if changed:
                db.add(best_match)
                updated += 1
            else:
                skipped += 1
            await _upsert_standard_job_match(
                vessel_id=vessel_id,
                standard_job_id=std_job.id,
                matched_job_id=best_match.id,
                current_user=current_user,
                db=db,
                match_status=MatchStatus.matched,
                match_score=best_score,
            )
            continue

        existing_same_name = await db.execute(
            select(Job).where(
                Job.vessel_id == vessel_id,
                Job.job_name == std_job.job_name,
                Job.is_deleted == False,
            )
        )
        if existing_same_name.scalar_one_or_none() is not None:
            skipped += 1
            continue

        new_job = Job(
            tenant_id=current_user.tenant_id,
            vessel_id=vessel_id,
            job_name=std_job.job_name,
            job_description=std_job.job_description,
            frequency=std_job.frequency,
            frequency_type=std_job.frequency_type,
            is_critical=True,
            source_reference=std_job.library_reference,
            qc_status=QCStatus.accepted,
            is_unmapped=True,
        )
        db.add(new_job)
        await db.flush()
        await _upsert_standard_job_match(
            vessel_id=vessel_id,
            standard_job_id=std_job.id,
            matched_job_id=new_job.id,
            current_user=current_user,
            db=db,
            match_status=MatchStatus.not_found,
            match_score=0,
        )
        added += 1

    await db.commit()
    return {"added": added, "updated": updated, "skipped": skipped}


@router.post("/vessels/{vessel_id}/standard-jobs/remove-critical-jobs", summary="Undo critical standard jobs")
async def remove_critical_jobs(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)

    critical_result = await db.execute(
        select(StandardJob).where(
            StandardJob.tenant_id == current_user.tenant_id,
            StandardJob.is_deleted == False,
            StandardJob.is_critical == True,
        )
    )
    critical_jobs = critical_result.scalars().all()
    if not critical_jobs:
        return {"removed": 0, "unmarked": 0, "skipped": 0, "total": 0}

    critical_job_ids = [job.id for job in critical_jobs]
    critical_ref_map = {
        job.id: _norm_text(job.library_reference) for job in critical_jobs if job.library_reference
    }

    match_result = await db.execute(
        select(StandardJobMatch).where(
            StandardJobMatch.vessel_id == vessel_id,
            StandardJobMatch.standard_job_id.in_(critical_job_ids),
            StandardJobMatch.is_deleted == False,
        )
    )
    matches = match_result.scalars().all()
    matches_by_standard_job_id = {match.standard_job_id: match for match in matches}
    matched_job_ids = {match.matched_job_id for match in matches if match.matched_job_id}

    all_jobs_result = await db.execute(
        select(Job).where(
            Job.vessel_id == vessel_id,
            Job.is_deleted == False,
        )
    )
    all_jobs = all_jobs_result.scalars().all()
    jobs_by_id = {job.id: job for job in all_jobs}

    removed = 0
    unmarked = 0
    skipped = 0
    removed_job_ids: set[uuid.UUID] = set()

    for std_job in critical_jobs:
        match = matches_by_standard_job_id.get(std_job.id)

        candidate_job: Job | None = None
        if match is not None and match.matched_job_id is not None:
            candidate_job = jobs_by_id.get(match.matched_job_id)

        if candidate_job is None:
            fallback_jobs = [
                job
                for job in all_jobs
                if not job.is_deleted
                and (
                    _job_looks_library_imported(job, std_job)
                    or _job_matches_critical_reference(job, std_job.library_reference)
                )
            ]
            candidate_job = next((job for job in fallback_jobs if _job_is_library_added_critical(job, std_job)), None)
            if candidate_job is None:
                candidate_job = next(
                    (
                        job
                        for job in fallback_jobs
                        if bool(job.is_critical)
                        and _job_matches_critical_reference(job, std_job.library_reference)
                    ),
                    None,
                )

        if candidate_job is None:
            if match is not None:
                match.matched_job_id = None
                match.match_status = MatchStatus.not_found
                match.match_score = 0
                db.add(match)
            skipped += 1
            continue

        if _job_is_library_added_critical(candidate_job, std_job):
            candidate_job.is_deleted = True
            db.add(candidate_job)
            removed_job_ids.add(candidate_job.id)
            removed += 1
        else:
            changed = False
            if candidate_job.is_critical:
                candidate_job.is_critical = False
                changed = True
            cleaned_source_reference = _remove_source_reference(candidate_job.source_reference, std_job.library_reference)
            if cleaned_source_reference != candidate_job.source_reference:
                candidate_job.source_reference = cleaned_source_reference
                changed = True
            if changed:
                db.add(candidate_job)
                unmarked += 1
            else:
                skipped += 1

        if match is not None:
            match.matched_job_id = None
            match.match_status = MatchStatus.not_found
            match.match_score = 0
            db.add(match)

    remaining_critical_refs = set(critical_ref_map.values())
    for job in all_jobs:
        if job.id in removed_job_ids or job.is_deleted:
            continue
        if not _job_is_library_added_critical(job):
            continue
        source_ref = _norm_text(job.source_reference)
        if not any(ref and ref in source_ref for ref in remaining_critical_refs):
            continue
        job.is_deleted = True
        db.add(job)
        removed_job_ids.add(job.id)
        removed += 1

    await db.commit()
    return {"removed": removed, "unmarked": unmarked, "skipped": skipped, "total": len(critical_jobs)}


@router.get("/vessels/{vessel_id}/missing-manuals", summary="Missing manual gaps")
async def get_missing_manuals(
    vessel_id: uuid.UUID,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    await _get_vessel_or_404(vessel_id, db)
    result = await db.execute(
        select(MissingManualGap).where(
            MissingManualGap.vessel_id == vessel_id,
            MissingManualGap.is_deleted == False,
        )
    )
    gaps = result.scalars().all()
    return {
        "items": [
            {
                "id": str(g.id),
                "machinery_group": g.machinery_group,
                "machinery_name": g.machinery_name,
                "is_mandatory": g.is_mandatory,
                "gap_status": g.gap_status,
                "notes": g.notes,
            }
            for g in gaps
        ]
    }


@router.patch("/vessels/{vessel_id}/missing-manuals/{gap_id}", summary="Update missing manual gap status")
async def update_missing_manual_gap(
    vessel_id: uuid.UUID,
    gap_id: uuid.UUID,
    body: dict[str, str],
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    result = await db.execute(
        select(MissingManualGap).where(
            MissingManualGap.id == gap_id,
            MissingManualGap.vessel_id == vessel_id,
            MissingManualGap.is_deleted == False,
        )
    )
    gap = result.scalar_one_or_none()
    if gap is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Gap not found")

    if "gap_status" in body:
        gap.gap_status = body["gap_status"]
    if "notes" in body:
        gap.notes = body["notes"]

    db.add(gap)
    await db.commit()
    return {"id": str(gap.id), "gap_status": gap.gap_status}
