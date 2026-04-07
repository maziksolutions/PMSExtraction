from __future__ import annotations

import io
import logging
import re
import uuid
from typing import Any

from sqlalchemy import select, text

from app.models.component import Component, QCStatus
from app.models.ingestion import Manual
from app.models.job import Job
from app.models.spare import Spare
from app.models.vessel import VesselProject

logger = logging.getLogger(__name__)

FIELD_MAPPINGS = {
    "group1": "Group 1",
    "group2": "Group 2",
    "main_machinery": "Main Machinery",
    "component_name": "Component Name",
    "maker": "Maker",
    "model": "Model",
    "specification": "Specification / Particulars",
    "serial_number": "Serial Number",
    "is_critical": "Critical",
    "qc_status": "QC Status",
    "job_name": "Job Name",
    "job_code": "Job Code",
    "job_description": "Description",
    "safety_precaution": "Safety Precautions",
    "tools_required": "Tools Required",
    "performing_rank": "Performing Rank",
    "verifying_rank": "Verifying Rank",
    "frequency": "Frequency",
    "frequency_type": "Frequency Type",
    "cms_id": "CMS ID",
    "component_linked": "Component Linked",
    "part_name": "Part Name",
    "part_number": "Part Number",
    "drawing_number": "Drawing Number",
    "drawing_position": "Drawing Position",
    "spare_assembly": "Spare Assembly",
    "assembly_description": "Assembly Description",
    "spare_maker": "Spare Maker",
    "spare_model": "Spare Model",
    "machinery_maker": "Machinery Maker",
    "machinery_model": "Machinery Model",
    "source_reference": "Source Reference",
    "extraction_method": "Extraction Method",
    "reason": "Reason",
    "record_type": "Record Type",
}

COMPONENT_MASTER_HEADERS = [
    "vesselName",
    "ShipComponentName",
    "HierarchyComponentCode",
    "ShipComponentCode",
    "ComponentType",
    "Priority",
    "Status",
    "Quantity",
    "Category",
    "Capacity",
    "Department",
    "AccountCode",
    "RunningAverage",
    "LifeTimeHrs",
    "InstallationDate",
    "WorkLocation",
    "StorageLocation",
    "Critical",
    "UnitBased",
    "Rotate",
    "Spare",
    "Instrument",
    "Inactive",
    "IsOwnerEssential",
    "Power",
    "Poles",
    "RPM",
    "Volts",
    "AMPs",
    "Hz",
    "FiringOrder",
    "FrameType",
    "Specifications",
    "Maker",
    "DateofManufacture",
]

JOB_MASTER_HEADERS = [
    "vesselName",
    "JobName",
    "jobCategory",
    "jobGroup",
    "isInactive",
    "isCritical",
    "jobDescription",
    "responsibility",
    "frequencyType",
    "frequency",
    "alternateFrequencyType",
    "alternateFrequency",
    "jobType",
    "SafetyProcedureCode",
    "OperationalProcedureCode",
    "DocumentReferenceCode",
    "ProcedureReferenceCode",
    "windowStart",
    "windowEnd",
    "AsRequired",
    "IsConditionBased",
    "IsUnitBased",
    "RiskAssesment",
    "SafetyChecklist",
    "CompanyApproval",
    "TechForm",
    "MOC",
    "IsPhotoMandatory",
]

JOB_COMPONENT_LINK_HEADERS = [
    "vesselName",
    "ShipComponentCode",
    "JobName",
    "IsInactive",
    "IsCritical",
    "JobDescription",
    "Responsibility",
    "FrequencyType",
    "Frequency",
    "AlternateFrequencyType",
    "AlternateFrequency",
    "LastDoneDate",
    "lastCounterValue",
    "WindowStart",
    "WindowEnd",
    "AsRequired",
    "CompanyApproval",
    "TechForm",
]

JOB_SPARES_LINK_HEADERS = ["vesselName", "JobName", "SpareCode", "ReqQuantity"]
JOB_ATTACHMENT_HEADERS = ["vesselName", "JobName", "attachmentType", "filename", "filePath", "description"]
MAINT_PROCEDURE_HEADERS = ["vesselName", "procedure", "procedureCode", "description"]
MAINT_REFERENCE_HEADERS = ["vesselName", "reference", "referenceCode", "description"]
SPARE_MASTER_HEADERS = [
    "vesselName",
    "inventoryCode",
    "inventoryName",
    "minimumLevel",
    "priority",
    "partNo",
    "makerReference",
    "capacity",
    "category",
    "critical",
    "spareAssembly",
    "ROB",
    "reOrderQty",
    "Location",
    "inactive",
    "IsOwnerEssential",
    "instrument",
    "autoReOrder",
    "partpostionno",
    "Specification / Particulars",
    "PDF Reference",
    "Page Reference",
    "Source Reference",
]
SPARE_ATTACHMENT_HEADERS = ["vesselName", "inventoryName", "attachmentType", "filename", "filePath", "description"]
SPARE_ASSEMBLY_HEADERS = ["vesselName", "spareAssembly", "description"]

EXPORT_FORMATS = {
    "bundle": [
        ("Ship PMS Master", "ship_pms_master", COMPONENT_MASTER_HEADERS),
        ("Ship Maintenance Master", "ship_maintenance_master", JOB_MASTER_HEADERS),
        ("Job Component Link", "job_component_link", JOB_COMPONENT_LINK_HEADERS),
        ("Job Spares Link", "job_spares_link", JOB_SPARES_LINK_HEADERS),
        ("CSM Maintenance Link", "csm_maintenance_link", ["JobName", "surveycode"]),
        ("Attachment", "component_attachments", JOB_ATTACHMENT_HEADERS),
        ("Maintenance Procedure Load Shee", "maintenance_procedure", MAINT_PROCEDURE_HEADERS),
        ("Maintenance Reference Load Shee", "maintenance_reference", MAINT_REFERENCE_HEADERS),
        ("Ship Spare Parts", "ship_spare_parts", SPARE_MASTER_HEADERS),
        ("Spare Attachment", "spare_attachments", SPARE_ATTACHMENT_HEADERS),
        ("Ship Spare Assembly Load Sheet", "ship_spare_assemblies", SPARE_ASSEMBLY_HEADERS),
        ("Excluded Records", "excluded", []),
    ],
    "component_master": [("Ship PMS Master", "ship_pms_master", COMPONENT_MASTER_HEADERS)],
    "job_master": [
        ("Ship Maintenance Master", "ship_maintenance_master", JOB_MASTER_HEADERS),
        ("Job Component Link", "job_component_link", JOB_COMPONENT_LINK_HEADERS),
        ("Job Spares Link", "job_spares_link", JOB_SPARES_LINK_HEADERS),
        ("CSM Maintenance Link", "csm_maintenance_link", ["JobName", "surveycode"]),
        ("Attachment", "job_attachments", JOB_ATTACHMENT_HEADERS),
    ],
    "maintenance_procedure": [("Maintenance Procedure Load Shee", "maintenance_procedure", MAINT_PROCEDURE_HEADERS)],
    "maintenance_reference": [("Maintenance Reference Load Shee", "maintenance_reference", MAINT_REFERENCE_HEADERS)],
    "spare_master": [
        ("Ship Spare Parts", "ship_spare_parts", SPARE_MASTER_HEADERS),
        ("Attachment", "spare_attachments", SPARE_ATTACHMENT_HEADERS),
    ],
    "spare_assembly": [("Ship Spare Assembly Load Sheet", "ship_spare_assemblies", SPARE_ASSEMBLY_HEADERS)],
}


class ExportService:
    def parse_template(self, file_bytes: bytes, filename: str) -> dict[str, Any]:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            sheets: dict[str, list[dict[str, Any]]] = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = []
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
                for col_idx, cell_value in enumerate(first_row, start=1):
                    if cell_value:
                        system_field = self._auto_map(str(cell_value))
                        headers.append(
                            {
                                "column_index": col_idx,
                                "column_header": str(cell_value),
                                "field_name": system_field,
                                "auto_mapped": system_field is not None,
                            }
                        )
                sheets[sheet_name] = headers
            return {"sheet_mappings": sheets}
        except Exception as exc:
            logger.warning("Could not parse Excel template %s: %s", filename, exc)
            return {"sheet_mappings": {}}

    async def serialize_async(
        self,
        db: Any,
        vessel_id: uuid.UUID | str,
        export_schema: dict | None = None,
    ) -> dict[str, Any]:
        vessel_uuid = uuid.UUID(str(vessel_id))
        vessel = await self._load_vessel(db, vessel_uuid)
        components = await self._load_components(db, vessel_uuid)
        jobs = await self._load_jobs(db, vessel_uuid)
        spares = await self._load_spares(db, vessel_uuid)
        manual_lookup = await self._load_manual_lookup(db, components, jobs, spares)
        code_lookup = await self._load_component_code_lookup(db, vessel)

        generic = self._serialize_generic_sections(
            vessel=vessel,
            components=components,
            jobs=jobs,
            spares=spares,
            manual_lookup=manual_lookup,
        )
        derived = self._build_sample_exports(
            vessel=vessel,
            components=components,
            jobs=jobs,
            spares=spares,
            manual_lookup=manual_lookup,
            code_lookup=code_lookup,
        )
        return {**generic, **derived}

    def to_excel(
        self,
        serialized_data: dict[str, Any],
        template_schema: dict | None = None,
        export_format: str = "bundle",
    ) -> bytes:
        if template_schema and export_format == "bundle":
            return self._to_template_excel(serialized_data, template_schema)
        return self._to_built_in_excel(serialized_data, export_format)

    def _auto_map(self, header: str) -> str | None:
        header_lower = header.lower().replace(" ", "_")
        for field, display in FIELD_MAPPINGS.items():
            if field == header_lower or display.lower().replace(" ", "_") == header_lower:
                return field
        return None

    async def _load_vessel(self, db: Any, vessel_id: uuid.UUID) -> VesselProject:
        result = await db.execute(
            select(VesselProject).where(
                VesselProject.id == vessel_id,
                VesselProject.is_deleted == False,
            )
        )
        vessel = result.scalar_one_or_none()
        if vessel is None:
            raise ValueError("Vessel not found for export")
        return vessel

    async def _load_components(self, db: Any, vessel_id: uuid.UUID) -> list[Component]:
        result = await db.execute(
            select(Component).where(
                Component.vessel_id == vessel_id,
                Component.is_deleted == False,
            )
        )
        return result.scalars().all()

    async def _load_jobs(self, db: Any, vessel_id: uuid.UUID) -> list[Job]:
        result = await db.execute(
            select(Job).where(
                Job.vessel_id == vessel_id,
                Job.is_deleted == False,
            )
        )
        return result.scalars().all()

    async def _load_spares(self, db: Any, vessel_id: uuid.UUID) -> list[Spare]:
        result = await db.execute(
            select(Spare).where(
                Spare.vessel_id == vessel_id,
                Spare.is_deleted == False,
                Spare.is_duplicate == False,
            )
        )
        return result.scalars().all()

    async def _load_manual_lookup(
        self,
        db: Any,
        components: list[Component],
        jobs: list[Job],
        spares: list[Spare],
    ) -> dict[uuid.UUID, Manual]:
        manual_ids = {
            manual_id
            for manual_id in (
                [c.source_manual_id for c in components]
                + [j.source_manual_id for j in jobs]
                + [s.source_manual_id for s in spares]
            )
            if manual_id
        }
        if not manual_ids:
            return {}
        result = await db.execute(select(Manual).where(Manual.id.in_(manual_ids)))
        return {manual.id: manual for manual in result.scalars().all()}

    async def _load_component_code_lookup(
        self,
        db: Any,
        vessel: VesselProject,
    ) -> list[dict[str, Any]]:
        params = {"tid": str(vessel.tenant_id), "vessel_type": vessel.vessel_type}
        query = text(
            "SELECT csl.group1_code, csl.group1_name, csl.group2_code, csl.group2_name, "
            "csl.machinery_code, csl.machinery_name, csl.component_code, csl.component_name, "
            "csl.component_type "
            "FROM component_structure_library csl "
            "LEFT JOIN vessel_types vt ON vt.id = csl.vessel_type_id "
            "WHERE csl.tenant_id = :tid AND csl.is_deleted = false "
            "AND (csl.status = 'active' OR csl.status IS NULL) "
            "AND (vt.name = :vessel_type OR csl.vessel_type_id IS NULL) "
            "ORDER BY csl.group1_name, csl.group2_name, csl.machinery_name, csl.component_name"
        )
        result = await db.execute(query, params)
        rows = [dict(row) for row in result.mappings().all()]
        if rows:
            return rows
        fallback = await db.execute(
            text(
                "SELECT group1_code, group1_name, group2_code, group2_name, machinery_code, "
                "machinery_name, component_code, component_name, component_type "
                "FROM component_structure_library "
                "WHERE tenant_id = :tid AND is_deleted = false "
                "AND (status = 'active' OR status IS NULL)"
            ),
            {"tid": str(vessel.tenant_id)},
        )
        return [dict(row) for row in fallback.mappings().all()]

    def _serialize_generic_sections(
        self,
        *,
        vessel: VesselProject,
        components: list[Component],
        jobs: list[Job],
        spares: list[Spare],
        manual_lookup: dict[uuid.UUID, Manual],
    ) -> dict[str, Any]:
        accepted_components: list[dict[str, Any]] = []
        accepted_jobs: list[dict[str, Any]] = []
        accepted_spares: list[dict[str, Any]] = []
        excluded: list[dict[str, Any]] = []

        for component in components:
            row = {
                "group1": component.group1,
                "group2": component.group2,
                "main_machinery": component.main_machinery,
                "component_name": component.component_name,
                "maker": component.maker or "",
                "model": component.model or "",
                "specification": component.specification or "",
                "serial_number": component.serial_number or "",
                "location": component.location or "",
                "machinery_particulars": component.machinery_particulars or "",
                "job_pages": component.job_pages or "",
                "spare_pages": component.spare_pages or "",
                "pdf_reference": component.pdf_reference or "",
                "page_reference": component.page_reference or "",
                "is_critical": "Yes" if component.is_critical else "No",
                "qc_status": component.qc_status.value,
            }
            if component.qc_status in (QCStatus.accepted, QCStatus.modified):
                accepted_components.append(row)
            else:
                excluded.append({**row, "record_type": "component", "reason": f"QC Status: {component.qc_status.value}"})

        for job in jobs:
            row = {
                "job_name": job.job_name,
                "job_code": job.job_code or "",
                "job_description": job.job_description or "",
                "safety_precaution": job.safety_precaution or "",
                "tools_required": job.tools_required or "",
                "performing_rank": job.performing_rank or "",
                "verifying_rank": job.verifying_rank or "",
                "frequency": job.frequency or "",
                "frequency_type": job.frequency_type.value if job.frequency_type else "",
                "cms_id": job.cms_id or "",
                "pdf_reference": job.pdf_reference or "",
                "page_reference": job.page_reference or "",
                "source_reference": job.source_reference or "",
                "is_critical": "Yes" if job.is_critical else "No",
                "qc_status": job.qc_status.value,
                "component_linked": str(job.component_id) if job.component_id else "",
            }
            if job.qc_status in (QCStatus.accepted, QCStatus.modified):
                accepted_jobs.append(row)
            else:
                excluded.append({**row, "record_type": "job", "reason": f"QC Status: {job.qc_status.value}"})

        for spare in spares:
            manual = manual_lookup.get(spare.source_manual_id) if spare.source_manual_id else None
            pdf_reference = manual.original_filename if manual else ""
            source_reference = self._source_reference(pdf_reference, spare.page_reference, None)
            part_number = spare.part_number or source_reference or pdf_reference or ""
            row = {
                "part_name": spare.part_name,
                "part_number": part_number,
                "drawing_number": spare.drawing_number or "",
                "drawing_position": spare.drawing_position or "",
                "specification": spare.specification or "",
                "spare_assembly": spare.spare_assembly or spare.spare_model or "",
                "assembly_description": spare.assembly_description or spare.spare_assembly or spare.spare_model or "",
                "spare_maker": spare.spare_maker or "",
                "spare_model": spare.spare_model or "",
                "machinery_maker": spare.machinery_maker or "",
                "machinery_model": spare.machinery_model or "",
                "pdf_reference": pdf_reference,
                "page_reference": spare.page_reference or "",
                "source_reference": source_reference,
                "extraction_method": spare.extraction_method.value,
                "is_critical": "Yes" if spare.is_critical else "No",
                "qc_status": spare.qc_status.value,
                "component_linked": str(spare.component_id) if spare.component_id else "",
            }
            if spare.qc_status in (QCStatus.accepted, QCStatus.modified):
                accepted_spares.append(row)
            else:
                excluded.append({**row, "record_type": "spare", "reason": f"QC Status: {spare.qc_status.value}"})

        return {
            "components": accepted_components,
            "jobs": accepted_jobs,
            "spares": accepted_spares,
            "excluded": excluded,
        }

    def _build_sample_exports(
        self,
        *,
        vessel: VesselProject,
        components: list[Component],
        jobs: list[Job],
        spares: list[Spare],
        manual_lookup: dict[uuid.UUID, Manual],
        code_lookup: list[dict[str, Any]],
    ) -> dict[str, list[dict[str, Any]]]:
        component_rows = [component for component in components if component.qc_status in (QCStatus.accepted, QCStatus.modified)]
        job_rows = [job for job in jobs if job.qc_status in (QCStatus.accepted, QCStatus.modified)]
        spare_rows = [spare for spare in spares if spare.qc_status in (QCStatus.accepted, QCStatus.modified)]

        code_map = {
            component.id: self._match_component_code_row(component, code_lookup)
            for component in component_rows
        }

        return {
            "ship_pms_master": self._build_component_loadsheet(vessel, component_rows, code_map),
            "ship_maintenance_master": self._build_job_master_loadsheet(vessel, job_rows, component_rows, code_map),
            "job_component_link": self._build_job_component_links(vessel, job_rows, component_rows, code_map),
            "job_spares_link": self._build_job_spares_links(vessel, job_rows, spare_rows, component_rows, code_map),
            "csm_maintenance_link": [],
            "job_attachments": [],
            "component_attachments": [],
            "maintenance_procedure": self._build_maintenance_procedures(vessel, job_rows),
            "maintenance_reference": self._build_maintenance_references(vessel, job_rows),
            "ship_spare_parts": self._build_spare_master_loadsheet(vessel, spare_rows, component_rows, code_map, manual_lookup),
            "spare_attachments": [],
            "ship_spare_assemblies": self._build_spare_assembly_loadsheet(vessel, spare_rows),
        }

    def _to_built_in_excel(self, data: dict[str, Any], export_format: str) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        spec = EXPORT_FORMATS.get(export_format, EXPORT_FORMATS["bundle"])
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

        for index, (sheet_name, data_key, headers) in enumerate(spec):
            ws = wb.active if index == 0 else wb.create_sheet(sheet_name)
            ws.title = sheet_name
            rows = list(data.get(data_key, []))
            if not headers:
                headers = self._headers_from_rows(rows)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            for row in rows:
                ws.append([row.get(header, "") for header in headers])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _to_template_excel(self, data: dict[str, Any], schema: dict | None) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        sheets = self._resolve_schema_sheets(data, schema)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

        for index, (sheet_name, rows) in enumerate(sheets.items()):
            ws = wb.active if index == 0 else wb.create_sheet(sheet_name)
            ws.title = sheet_name
            headers = self._resolve_headers(rows, schema, sheet_name)
            labels = self._resolve_header_labels(headers, schema, sheet_name)
            ws.append(labels)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            for row in rows:
                ws.append([row.get(header, "") for header in headers])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _resolve_schema_sheets(self, data: dict[str, Any], schema: dict | None) -> dict[str, list[dict[str, Any]]]:
        default = {
            "Components": data.get("components", []),
            "Jobs": data.get("jobs", []),
            "Spares": data.get("spares", []),
            "Excluded Records": data.get("excluded", []),
        }
        if not schema:
            return default
        resolved: dict[str, list[dict[str, Any]]] = {}
        schema_map = schema.get("sheet_mappings", schema) if isinstance(schema, dict) else {}
        for sheet_name in schema_map.keys():
            normalized = sheet_name.strip().lower()
            if "component" in normalized:
                resolved[sheet_name] = data.get("components", [])
            elif "job" in normalized:
                resolved[sheet_name] = data.get("jobs", [])
            elif "spare" in normalized:
                resolved[sheet_name] = data.get("spares", [])
            elif "exclude" in normalized or "reject" in normalized:
                resolved[sheet_name] = data.get("excluded", [])
        return resolved or default

    def _resolve_headers(self, rows: list[dict[str, Any]], schema: dict | None, sheet_name: str) -> list[str]:
        if not rows:
            return []
        schema_map = schema.get("sheet_mappings", schema) if isinstance(schema, dict) else {}
        columns = schema_map.get(sheet_name) or []
        mapped_fields = [col.get("field_name") for col in columns if col.get("field_name")]
        return mapped_fields or list(rows[0].keys())

    def _resolve_header_labels(self, headers: list[str], schema: dict | None, sheet_name: str) -> list[str]:
        schema_map = schema.get("sheet_mappings", schema) if isinstance(schema, dict) else {}
        columns = schema_map.get(sheet_name) or []
        header_by_field = {col.get("field_name"): col.get("column_header") for col in columns if col.get("field_name")}
        return [header_by_field.get(header) or FIELD_MAPPINGS.get(header, header) for header in headers]

    def _headers_from_rows(self, rows: list[dict[str, Any]]) -> list[str]:
        headers: list[str] = []
        for row in rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)
        return headers or ["message"]

    def _normalize(self, value: str | None) -> str:
        return re.sub(r"\s+", " ", (value or "").strip()).lower()

    def _yes_no(self, value: bool) -> str:
        return "Yes" if value else "No"

    def _source_reference(
        self,
        pdf_reference: str | None,
        page_reference: int | None,
        source_reference: str | None,
    ) -> str:
        if source_reference:
            return source_reference
        if pdf_reference and page_reference:
            return f"{pdf_reference} (p.{page_reference})"
        if pdf_reference:
            return pdf_reference
        return f"p.{page_reference}" if page_reference else ""

    def _priority(self, is_critical: bool) -> str:
        return "Critical" if is_critical else "Normal"

    def _match_component_code_row(
        self,
        component: Component,
        code_lookup: list[dict[str, Any]],
    ) -> dict[str, Any] | None:
        g1 = self._normalize(component.group1)
        g2 = self._normalize(component.group2)
        mm = self._normalize(component.main_machinery)
        name = self._normalize(component.component_name)
        for row in code_lookup:
            if (
                self._normalize(row.get("group1_name")) == g1
                and self._normalize(row.get("group2_name")) == g2
                and self._normalize(row.get("machinery_name")) == mm
                and self._normalize(row.get("component_name")) == name
            ):
                return row
        for row in code_lookup:
            if self._normalize(row.get("component_name")) == name:
                return row
        return None

    def _build_component_loadsheet(
        self,
        vessel: VesselProject,
        components: list[Component],
        code_map: dict[uuid.UUID, dict[str, Any] | None],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for component in components:
            matched = code_map.get(component.id) or {}
            rows.append(
                {
                    "vesselName": vessel.name,
                    "ShipComponentName": component.component_name,
                    "HierarchyComponentCode": matched.get("machinery_code") or matched.get("group2_code") or "",
                    "ShipComponentCode": matched.get("component_code") or "",
                    "ComponentType": matched.get("component_type") or component.main_machinery or "",
                    "Priority": self._priority(component.is_critical),
                    "Status": "Active",
                    "Quantity": 1,
                    "Category": component.group1 or component.group2 or "",
                    "Capacity": component.machinery_particulars or "",
                    "Department": "",
                    "AccountCode": "",
                    "RunningAverage": "",
                    "LifeTimeHrs": "",
                    "InstallationDate": "",
                    "WorkLocation": component.location or "",
                    "StorageLocation": "",
                    "Critical": self._yes_no(component.is_critical),
                    "UnitBased": "No",
                    "Rotate": "No",
                    "Spare": "Yes" if component.spare_pages else "No",
                    "Instrument": "No",
                    "Inactive": "No",
                    "IsOwnerEssential": "No",
                    "Power": "",
                    "Poles": "",
                    "RPM": "",
                    "Volts": "",
                    "AMPs": "",
                    "Hz": "",
                    "FiringOrder": "",
                    "FrameType": "",
                    "Specifications": component.specification or component.machinery_particulars or "",
                    "Maker": component.maker or "",
                    "DateofManufacture": "",
                }
            )
        return rows

    def _build_job_master_loadsheet(
        self,
        vessel: VesselProject,
        jobs: list[Job],
        components: list[Component],
        code_map: dict[uuid.UUID, dict[str, Any] | None],
    ) -> list[dict[str, Any]]:
        components_by_id = {component.id: component for component in components}
        rows: list[dict[str, Any]] = []
        for job in jobs:
            component = components_by_id.get(job.component_id) if job.component_id else None
            procedure_code = self._procedure_code(job)
            reference_code = self._reference_code(job)
            rows.append(
                {
                    "vesselName": vessel.name,
                    "JobName": job.job_name,
                    "jobCategory": component.group1 if component else "",
                    "jobGroup": component.main_machinery if component else "",
                    "isInactive": "No",
                    "isCritical": self._yes_no(job.is_critical),
                    "jobDescription": job.job_description or job.job_name,
                    "responsibility": job.performing_rank or "",
                    "frequencyType": self._frequency_label(job.frequency_type.value if job.frequency_type else None),
                    "frequency": job.frequency or "",
                    "alternateFrequencyType": "",
                    "alternateFrequency": "",
                    "jobType": "Maintenance",
                    "SafetyProcedureCode": "",
                    "OperationalProcedureCode": "",
                    "DocumentReferenceCode": reference_code,
                    "ProcedureReferenceCode": procedure_code,
                    "windowStart": "",
                    "windowEnd": "",
                    "AsRequired": "No",
                    "IsConditionBased": "No",
                    "IsUnitBased": "No",
                    "RiskAssesment": "",
                    "SafetyChecklist": "",
                    "CompanyApproval": "",
                    "TechForm": "",
                    "MOC": "",
                    "IsPhotoMandatory": "No",
                }
            )
        return rows

    def _build_job_component_links(
        self,
        vessel: VesselProject,
        jobs: list[Job],
        components: list[Component],
        code_map: dict[uuid.UUID, dict[str, Any] | None],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for job in jobs:
            if not job.component_id:
                continue
            matched = code_map.get(job.component_id) or {}
            rows.append(
                {
                    "vesselName": vessel.name,
                    "ShipComponentCode": matched.get("component_code") or "",
                    "JobName": job.job_name,
                    "IsInactive": "No",
                    "IsCritical": self._yes_no(job.is_critical),
                    "JobDescription": job.job_description or job.job_name,
                    "Responsibility": job.performing_rank or "",
                    "FrequencyType": self._frequency_label(job.frequency_type.value if job.frequency_type else None),
                    "Frequency": job.frequency or "",
                    "AlternateFrequencyType": "",
                    "AlternateFrequency": "",
                    "LastDoneDate": "",
                    "lastCounterValue": "",
                    "WindowStart": "",
                    "WindowEnd": "",
                    "AsRequired": "No",
                    "CompanyApproval": "",
                    "TechForm": "",
                }
            )
        return rows

    def _build_job_spares_links(
        self,
        vessel: VesselProject,
        jobs: list[Job],
        spares: list[Spare],
        components: list[Component],
        code_map: dict[uuid.UUID, dict[str, Any] | None],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        spares_by_component: dict[uuid.UUID, list[Spare]] = {}
        for spare in spares:
            if spare.component_id:
                spares_by_component.setdefault(spare.component_id, []).append(spare)
        for job in jobs:
            if not job.component_id:
                continue
            linked_spares = spares_by_component.get(job.component_id, [])
            for spare in linked_spares:
                rows.append(
                    {
                        "vesselName": vessel.name,
                        "JobName": job.job_name,
                        "SpareCode": spare.part_number or spare.drawing_number or "",
                        "ReqQuantity": 1,
                    }
                )
        return rows

    def _build_maintenance_procedures(self, vessel: VesselProject, jobs: list[Job]) -> list[dict[str, Any]]:
        return [
            {
                "vesselName": vessel.name,
                "procedure": job.job_name,
                "procedureCode": self._procedure_code(job),
                "description": job.job_description or "",
            }
            for job in jobs
        ]

    def _build_maintenance_references(self, vessel: VesselProject, jobs: list[Job]) -> list[dict[str, Any]]:
        return [
            {
                "vesselName": vessel.name,
                "reference": job.job_name,
                "referenceCode": self._reference_code(job),
                "description": job.source_reference or job.pdf_reference or "",
            }
            for job in jobs
        ]

    def _build_spare_master_loadsheet(
        self,
        vessel: VesselProject,
        spares: list[Spare],
        components: list[Component],
        code_map: dict[uuid.UUID, dict[str, Any] | None],
        manual_lookup: dict[uuid.UUID, Manual],
    ) -> list[dict[str, Any]]:
        components_by_id = {component.id: component for component in components}
        rows: list[dict[str, Any]] = []
        for spare in spares:
            component = components_by_id.get(spare.component_id) if spare.component_id else None
            manual = manual_lookup.get(spare.source_manual_id) if spare.source_manual_id else None
            source_reference = self._source_reference(
                manual.original_filename if manual else None,
                spare.page_reference,
                None,
            )
            rows.append(
                {
                    "vesselName": vessel.name,
                    "inventoryCode": spare.part_number or source_reference or spare.drawing_number or "",
                    "inventoryName": spare.part_name,
                    "minimumLevel": "",
                    "priority": self._priority(spare.is_critical),
                    "partNo": spare.part_number or source_reference or "",
                    "makerReference": self._maker_reference(spare),
                    "capacity": spare.specification or "",
                    "category": component.component_name if component else "",
                    "critical": self._yes_no(spare.is_critical),
                    "spareAssembly": self._infer_spare_assembly(spare, component),
                    "ROB": "",
                    "reOrderQty": "",
                    "Location": component.location if component else "",
                    "inactive": "No",
                    "IsOwnerEssential": "No",
                    "instrument": "No",
                    "autoReOrder": "No",
                    "partpostionno": spare.drawing_position or "",
                    "PDF Reference": manual.original_filename if manual else "",
                    "Page Reference": spare.page_reference or "",
                    "Source Reference": source_reference,
                    "Specification / Particulars": spare.specification or "",
                }
            )
        return rows

    def _build_spare_assembly_loadsheet(self, vessel: VesselProject, spares: list[Spare]) -> list[dict[str, Any]]:
        seen: dict[str, dict[str, Any]] = {}
        for spare in spares:
            assembly = self._infer_spare_assembly(spare, None)
            if not assembly:
                continue
            seen.setdefault(
                assembly,
                {
                    "vesselName": vessel.name,
                    "spareAssembly": assembly,
                    "description": spare.assembly_description or spare.specification or assembly,
                },
            )
        return list(seen.values())

    def _frequency_label(self, value: str | None) -> str:
        if not value:
            return ""
        return value.replace("_", " ").title()

    def _procedure_code(self, job: Job) -> str:
        base = job.job_code or job.job_name
        compact = re.sub(r"[^A-Za-z0-9]+", "", base)[:24]
        return f"PROC-{compact}" if compact else f"PROC-{job.id.hex[:8]}"

    def _reference_code(self, job: Job) -> str:
        base = job.job_code or job.job_name
        compact = re.sub(r"[^A-Za-z0-9]+", "", base)[:24]
        return f"REF-{compact}" if compact else f"REF-{job.id.hex[:8]}"

    def _maker_reference(self, spare: Spare) -> str:
        values = [value for value in [spare.spare_maker, spare.spare_model, spare.drawing_number] if value]
        return " / ".join(values)

    def _infer_spare_assembly(self, spare: Spare, component: Component | None) -> str:
        return (
            spare.spare_assembly
            or spare.assembly_description
            or spare.spare_model
            or (component.component_name if component else "")
        )


export_service = ExportService()
